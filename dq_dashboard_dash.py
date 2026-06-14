from __future__ import annotations
import csv
import io
import json
import logging
import os
from datetime import datetime
from pathlib import Path
import dash
from dash import dcc, html, Input, Output, ALL, MATCH, ctx, State
import plotly.graph_objects as go
from flask import session as flask_session
import dq_auth
from dq_rules import get_all_rules
log = logging.getLogger("dq_dashboard")

#file paths
_DIR            = Path(__file__).parent
HISTORY_FILE    = _DIR / "dq_history.json"
CATEGORIES_FILE = _DIR / "le_book_categories.json"
ACTIVITY_FILE   = _DIR / "institution_activity.json"
PIPELINE_FILE        = _DIR / "pipeline_run.json"
PIPELINE_STATUS_FILE = _DIR / "pipeline_status.json"
WATERMARK_FILE       = _DIR / "watermark.json"
REPORTS_DIR     = _DIR / "reports"

# design tokens
BNR_GOLD = "#f7c35f"   # --thm-primary  (website gold accent)
BG       = "#f2ede9"   # --thm-gray     (website light warm background)
CARD     = "#FFFFFF"
TEXT     = "#1c1c27"   # --thm-black    (website near-black)
MUTED    = "#68686f"   # --thm-color    (website body text / secondary)
DIVIDER  = "#e7e1dc"   # --thm-border-color
C_GREEN  = "#B8860B"   # dark goldenrod  — good score
C_AMBER  = "#A0784A"   # mid warm tan    — medium score
C_RED    = "#7C3D1E"   # burnt rust      — poor score
BRAND    = "#753918"   # --thm-base     (website primary brand brown)
FONT     = "'Inter','Franklin Gothic Medium',Arial,sans-serif"

# one color per dimension
DIM_COLORS = {
    "completeness": "#753918",   # brand primary
    "accuracy":     "#B8860B",   # dark goldenrod
    "timeliness":   "#7C3D1E",   # burnt rust
    "validity":     "#C9956C",   # peach terra cotta
}

DIMS = ["completeness", "accuracy", "timeliness", "validity"]
DIM_LABELS = {
    "completeness": "Completeness",
    "accuracy":     "Accuracy",
    "timeliness":   "Timeliness",
    "validity":     "Validity",
}

# Internal category codes (kept for data access helpers)
CATEGORIES = ["ALL", "B", "MF", "SACCO", "OSACCO"]
CAT_LABELS = {
    "ALL":    "All Institutions",
    "B":      "Banks",
    "MF":     "Microfinance",
    "SACCO":  "SACCOs",
    "OSACCO": "OSACCOs",
}

# Landing page category definitions
# "SACCO" combines both SACCO and OSACCO institution types
LANDING_CATS = [
    {
        "code":     "B",
        "label":    "Banks",
        "subtitle": "Commercial & savings banks",
        "color":    "#753918",
        "types":    ["B"],
    },
    {
        "code":     "MF",
        "label":    "Microfinance",
        "subtitle": "Microfinance institutions",
        "color":    "#B8860B",
        "types":    ["MF"],
    },
    {
        "code":     "SACCO",
        "label":    "SACCO",
        "subtitle": "Savings & credit cooperatives (incl. OSACCOs)",
        "color":    "#A0784A",
        "types":    ["SACCO", "OSACCO"],
    },
]


# ── issue tracker (loaded fresh each render — SQLite is fast) ─────────────────
def _issue_summary() -> dict:
    try:
        from dq_issue_tracker import get_institution_issue_summary
        return get_institution_issue_summary()
    except Exception:
        return {}

def _institution_issues(le_book: str) -> list:
    try:
        from dq_issue_tracker import get_open_issues
        return get_open_issues(le_book)
    except Exception:
        return []

_URGENCY_COLORS = {
    "new":       "#A0784A",   # Mid Warm Tan — calm, not alarming
    "attention": "#B8860B",   # Dark Goldenrod — caution
    "urgent":    "#7C3D1E",   # Burnt Rust — stronger warning
    "critical":  "#DC2626",   # Red — universal critical
    "overdue":   "#991B1B",   # Deep Red — SLA breached
}
_URGENCY_LABELS = {
    "new":       "New (1–3d)",
    "attention": "Needs Attention (4–15d)",
    "urgent":    "Urgent (16–20d)",
    "critical":  "About to Breach (21–30d)",
    "overdue":   "⚠ Overdue — SLA Breached",
}

# ── data loading ───────────────────────────────────────────────────────────────

def _load_history() -> list:
    if not HISTORY_FILE.exists():
        return []
    try:
        data = json.loads(HISTORY_FILE.read_text(encoding="utf-8"))
        return data if isinstance(data, list) else []
    except Exception as exc:
        log.warning("History load failed: %s", exc)
        return []

def _load_pipeline_run() -> dict:
    if not PIPELINE_FILE.exists():
        return {}
    try:
        return json.loads(PIPELINE_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _load_pipeline_status() -> dict:
    if not PIPELINE_STATUS_FILE.exists():
        return {}
    try:
        return json.loads(PIPELINE_STATUS_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}

def _load_activity() -> dict:
    if not ACTIVITY_FILE.exists():
        return {}
    try:
        return json.loads(ACTIVITY_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}

def _load_all_categories() -> dict:
    if not CATEGORIES_FILE.exists():
        return {}
    try:
        return json.loads(CATEGORIES_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}

_HISTORY       = _load_history()
_PIPELINE      = _load_pipeline_run()
_history_mtime = HISTORY_FILE.stat().st_mtime  if HISTORY_FILE.exists()       else 0.0
_pipeline_mtime= PIPELINE_FILE.stat().st_mtime if PIPELINE_FILE.exists()      else 0.0


def _fresh_history() -> list:
    """Return history list, re-reading from disk only when the file has changed."""
    global _HISTORY, _history_mtime
    try:
        mtime = HISTORY_FILE.stat().st_mtime
        if mtime != _history_mtime:
            _HISTORY       = _load_history()
            _history_mtime = mtime
            log.info("History reloaded from disk (pipeline ran)")
    except Exception:
        pass
    return _HISTORY


def _fresh_pipeline() -> dict:
    """Return pipeline run dict, re-reading from disk only when the file has changed."""
    global _PIPELINE, _pipeline_mtime
    try:
        mtime = PIPELINE_FILE.stat().st_mtime
        if mtime != _pipeline_mtime:
            _PIPELINE       = _load_pipeline_run()
            _pipeline_mtime = mtime
    except Exception:
        pass
    return _PIPELINE


# ── data access helpers ────────────────────────────────────────────────────────

def _today_entry()      -> dict: h = _fresh_history(); return h[-1]         if h           else {}
def _yesterday_entry()  -> dict: h = _fresh_history(); return h[-2]         if len(h) >= 2 else {}
def _trend_entries(n=7) -> list: h = _fresh_history(); return h[-n:]        if h           else []


def _cat_scores(entry: dict, cat: str) -> dict:
    """Overall or per-category scores from one history entry.
    cat='SACCO' returns the average of SACCO+OSACCO (data stores these as OSACCO)."""
    if not entry:
        return {d: 0.0 for d in DIMS}
    if cat == "ALL":
        return entry.get("overall", {})
    if cat == "SACCO":
        by_cat = entry.get("by_category", {})
        sacco  = by_cat.get("SACCO",  {})
        osacco = by_cat.get("OSACCO", {})
        combined = {}
        for d in DIMS:
            vals = [float(src.get(d) or 0) for src in (sacco, osacco) if src]
            combined[d] = sum(vals) / len(vals) if vals else 0.0
        return combined
    return entry.get("by_category", {}).get(cat, {})


def _filter_institutions(entry: dict, cat: str) -> dict:
    """Return institutions dict filtered to the given category.
    cat='SACCO' includes both SACCO and OSACCO institution types."""
    inst = entry.get("by_institution", {}) if entry else {}
    if cat == "ALL":
        return inst
    if cat == "SACCO":
        return {lb: d for lb, d in inst.items()
                if d.get("category_type") in ("SACCO", "OSACCO")}
    return {lb: d for lb, d in inst.items() if d.get("category_type") == cat}


def _inst_scores(entry: dict, inst_code: str) -> dict:
    """Return dimension scores for a specific institution from one history entry."""
    if not entry or not inst_code:
        return {d: 0.0 for d in DIMS}
    d = entry.get("by_institution", {}).get(inst_code, {})
    return {dim: float(d.get(dim) or 0) for dim in DIMS}


def _category_counts(entry: dict) -> dict:
    counts = {c: 0 for c in CATEGORIES}
    counts["ALL"] = 0
    for data in entry.get("by_institution", {}).values():
        counts["ALL"] += 1
        ct = data.get("category_type", "")
        if ct in counts:
            counts[ct] += 1
    return counts


TABLE_NAMES_PRETTY: dict[str, str] = {
    "accounts":               "Accounts",
    "contract_loans":         "Contract Loans",
    "contract_schedules":     "Contract Schedules",
    "contracts_disburse":     "Contracts Disburse",
    "contracts_expanded":     "Contracts",
    "customers_expanded":     "Customers",
    "loan_applications_2":    "Loan Applications",
    "prev_loan_applications": "Prev Loan Apps",
}

ALL_TABLES = list(TABLE_NAMES_PRETTY.keys())


def _table_fails(entry: dict) -> dict[str, int]:
    """Return {table: failing_rule_count} from one history entry."""
    if not entry:
        return {}
    return entry.get("table_fail_counts", {})


def _table_nulls(entry: dict) -> dict[str, dict[str, int]]:
    """Return {table: {col: null_count}} for tables with null columns."""
    if not entry:
        return {}
    return entry.get("table_null_cols", {})


# ── score styling ──────────────────────────────────────────────────────────────

def _score_color(s: float) -> str:
    return C_GREEN if s >= 90 else C_AMBER if s >= 75 else C_RED

def _score_bg(s: float) -> str:
    return ("rgba(184,134,11,.10)"  if s >= 90 else
            "rgba(160,120,74,.10)"  if s >= 75 else
            "rgba(124,61,30,.10)")


# ── component builders ─────────────────────────────────────────────────────────

def _sparkline(values: list, color: str) -> dcc.Graph:
    r, g, b = int(color[1:3], 16), int(color[3:5], 16), int(color[5:7], 16)
    fig = go.Figure(go.Scatter(
        y=values or [0], mode="lines",
        line=dict(color=color, width=1.5),
        fill="tozeroy",
        fillcolor=f"rgba({r},{g},{b},0.12)",
    ))
    fig.update_layout(
        height=36,
        margin=dict(l=0, r=0, t=0, b=0),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        xaxis=dict(visible=False),
        yaxis=dict(visible=False, range=[0, 100]),
        showlegend=False,
    )
    return dcc.Graph(figure=fig, config={"displayModeBar": False},
                     style={"height": "36px", "marginTop": "8px"})


def _kpi_card(dim: str, score: float, delta: float, spark: list) -> html.Div:
    col     = _score_color(score)
    d_col   = C_GREEN if delta > 0 else C_RED if delta < 0 else MUTED
    d_icon  = "▲" if delta > 0 else "▼" if delta < 0 else "─"
    d_label = f"{d_icon} {abs(delta):.1f}%"

    return html.Div([
        html.Div(DIM_LABELS[dim], style={
            "fontSize": "11px", "fontWeight": "900",
            "color": MUTED, "letterSpacing": "0.06em",
            "textTransform": "uppercase", "lineHeight": "1.15",
        }),
        html.Div(f"{score:.1f}%", style={
            "fontSize": "30px", "fontWeight": "700",
            "color": col, "lineHeight": "1.1", "marginTop": "6px",
            "fontVariantNumeric": "tabular-nums",
        }),
        html.Div([
            html.Span(d_label, style={
                "color": d_col, "fontWeight": "700", "fontSize": "12px",
            }),
            html.Span(" vs yesterday", style={
                "color": MUTED, "fontSize": "11px",
            }),
        ], style={"marginTop": "4px", "lineHeight": "1.15"}),
        _sparkline(spark, col),
    ], style={
        "background":   CARD,
        "borderRadius": "8px",
        "padding":      "16px",
        "flex":         "1",
        "minWidth":     "150px",
        "borderTop":    f"3px solid {col}",
        "boxShadow":    "0 1px 4px rgba(117,57,24,0.08)",
    })


def _count_sparkline(values: list, color: str) -> dcc.Graph:
    """Sparkline for raw-count metrics (y-axis scales to data, not 0-100)."""
    r, g, b  = int(color[1:3], 16), int(color[3:5], 16), int(color[5:7], 16)
    max_val  = max(values) if values and max(values) > 0 else 1
    fig = go.Figure(go.Scatter(
        y=values or [0], mode="lines",
        line=dict(color=color, width=1.5),
        fill="tozeroy",
        fillcolor=f"rgba({r},{g},{b},0.12)",
    ))
    fig.update_layout(
        height=36,
        margin=dict(l=0, r=0, t=0, b=0),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        xaxis=dict(visible=False),
        yaxis=dict(visible=False, range=[0, max_val * 1.25]),
        showlegend=False,
    )
    return dcc.Graph(figure=fig, config={"displayModeBar": False},
                     style={"height": "36px", "marginTop": "8px"})


def _dup_card(count: int, delta: int, spark: list) -> html.Div:
    col    = C_GREEN if count == 0 else C_AMBER if count <= 10 else C_RED
    d_col  = C_RED if delta > 0 else C_GREEN if delta < 0 else MUTED
    d_icon = "▲" if delta > 0 else "▼" if delta < 0 else "─"
    return html.Div([
        html.Div("CUSTOMER DUPLICATES", style={
            "fontSize": "11px", "fontWeight": "900",
            "color": MUTED, "letterSpacing": "0.06em",
            "textTransform": "uppercase", "lineHeight": "1.15",
        }),
        html.Div(f"{count:,}", style={
            "fontSize": "30px", "fontWeight": "700",
            "color": col, "lineHeight": "1.1", "marginTop": "6px",
            "fontVariantNumeric": "tabular-nums",
        }),
        html.Div([
            html.Span(f"{d_icon} {abs(delta)}", style={
                "color": d_col, "fontWeight": "700", "fontSize": "12px",
            }),
            html.Span(" vs yesterday", style={
                "color": MUTED, "fontSize": "11px",
            }),
        ], style={"marginTop": "4px", "lineHeight": "1.15"}),
        _count_sparkline(spark, col),
    ], style={
        "background":   CARD,
        "borderRadius": "8px",
        "padding":      "16px",
        "flex":         "1",
        "minWidth":     "150px",
        "borderTop":    f"3px solid {col}",
        "boxShadow":    "0 1px 4px rgba(117,57,24,0.08)",
    })


def _table_card(table_name: str, fail_count: int, delta: int,
                spark: list, null_cols: dict) -> html.Div:
    """Per-table failing-rules card with collapsible mandatory-column null detail."""
    col    = C_GREEN if fail_count == 0 else C_AMBER if fail_count <= 3 else C_RED
    d_col  = C_RED   if delta > 0        else C_GREEN  if delta < 0        else MUTED
    d_icon = "▲"     if delta > 0        else "▼"      if delta < 0        else "─"
    pretty = TABLE_NAMES_PRETTY.get(table_name,
                                    table_name.replace("_", " ").title())

    null_section: list = []
    if null_cols:
        null_section = [
            html.Div([
                html.Span(
                    f"Mandatory columns with nulls: {len(null_cols)}",
                    style={"fontSize": "10px", "color": C_RED, "fontWeight": "700"},
                ),
                html.Span(
                    " ▶ show",
                    id={"type": "null-col-btn", "index": table_name},
                    n_clicks=0,
                    style={
                        "fontSize": "10px", "color": BRAND,
                        "cursor": "pointer", "marginLeft": "6px", "fontWeight": "700",
                    },
                ),
            ], style={"marginTop": "6px"}),
            html.Div(
                [html.Div(
                    f"└ {col}: {cnt:,} nulls",
                    style={"fontSize": "10px", "color": MUTED, "paddingLeft": "4px",
                           "lineHeight": "1.6"},
                ) for col, cnt in sorted(null_cols.items(), key=lambda x: -x[1])],
                id={"type": "null-col-body", "index": table_name},
                style={"display": "none", "marginTop": "4px"},
            ),
        ]

    other_count = max(0, fail_count - (1 if null_cols else 0))
    other_line  = html.Div(
        f"Other failing rules: {other_count}" if other_count > 0 else
        ("All rules passing" if fail_count == 0 else ""),
        style={
            "fontSize":   "10px",
            "color":      C_RED if other_count > 0 else C_GREEN if fail_count == 0 else MUTED,
            "marginTop":  "4px",
            "fontWeight": "700" if other_count > 0 else "400",
        },
    )

    return html.Div([
        html.Div(pretty.upper(), style={
            "fontSize": "10px", "fontWeight": "900", "color": MUTED,
            "letterSpacing": "0.06em", "textTransform": "uppercase",
        }),
        html.Div(str(fail_count), style={
            "fontSize": "28px", "fontWeight": "700", "color": col,
            "lineHeight": "1.1", "marginTop": "4px",
            "fontVariantNumeric": "tabular-nums",
        }),
        html.Div("failing rules", style={"fontSize": "10px", "color": MUTED}),
        html.Div([
            html.Span(f"{d_icon} {abs(delta)}", style={
                "color": d_col, "fontWeight": "700", "fontSize": "11px",
            }),
            html.Span(" vs yesterday", style={"color": MUTED, "fontSize": "10px"}),
        ], style={"marginTop": "3px"}),
        *null_section,
        other_line,
        _count_sparkline(spark, col),
    ], style={
        "background":   CARD,
        "borderRadius": "8px",
        "padding":      "14px 16px",
        "flex":         "1",
        "minWidth":     "140px",
        "borderTop":    f"3px solid {col}",
        "boxShadow":    "0 1px 4px rgba(117,57,24,0.08)",
    })


def _inst_dup_count(entry: dict, inst_code: str) -> int:
    if not entry or not inst_code:
        return 0
    return int(entry.get("by_institution", {}).get(inst_code, {}).get("customer_duplicates", 0))


def _cat_dup_count(entry: dict, cat: str) -> int:
    if not entry:
        return 0
    bc = entry.get("by_category", {})
    if cat == "SACCO":
        return int(bc.get("SACCO",  {}).get("customer_duplicates", 0)) + \
               int(bc.get("OSACCO", {}).get("customer_duplicates", 0))
    return int(bc.get(cat, {}).get("customer_duplicates", 0))


def _issue_counts_trend(n_days: int = 30,
                        le_books: set | None = None) -> list[dict]:
    """
    For each of the last n_days, compute:
      open     — issues open (detected but not yet resolved) on that date
      overdue  — open issues past their sla_deadline on that date
      resolved — issues resolved on that specific date

    Reconstructed from dq_open_issues dates — no separate history table needed.
    le_books: if given, restrict to those institution codes.
    """
    from datetime import date as _date, timedelta
    try:
        from dq_issue_tracker import get_issues
        all_issues = get_issues()          # all statuses
    except Exception:
        return []

    if le_books is not None:
        all_issues = [i for i in all_issues if str(i["le_book"]) in le_books]

    today = _date.today()
    rows  = []
    for offset in range(n_days - 1, -1, -1):
        day = today - timedelta(days=offset)
        ds  = day.isoformat()

        open_n = overdue_n = resolved_n = 0
        for iss in all_issues:
            det = iss.get("detected_at") or ""
            res = iss.get("resolved_at") or ""
            sla = iss.get("sla_deadline") or ""
            if not det or det > ds:
                continue
            if res and res <= ds:
                if res == ds:
                    resolved_n += 1
            else:
                open_n += 1
                if sla and sla < ds:
                    overdue_n += 1

        rows.append({"date": ds, "open": open_n,
                     "overdue": overdue_n, "resolved": resolved_n})
    return rows


def _issue_trend_figure(data: list) -> go.Figure:
    """30-day issue count trend: open, overdue, resolved per day."""
    dates    = [r["date"]    for r in data]
    open_v   = [r["open"]    for r in data]
    over_v   = [r["overdue"] for r in data]
    res_v    = [r["resolved"] for r in data]

    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=dates, y=open_v, name="Open",
        mode="lines+markers",
        line=dict(color=C_AMBER, width=2),
        marker=dict(size=4),
        hovertemplate="<b>Open</b>: %{y}<br>%{x}<extra></extra>",
    ))
    fig.add_trace(go.Scatter(
        x=dates, y=over_v, name="Overdue",
        mode="lines+markers",
        line=dict(color=_URGENCY_COLORS["overdue"], width=2, dash="dot"),
        marker=dict(size=4),
        hovertemplate="<b>Overdue</b>: %{y}<br>%{x}<extra></extra>",
    ))
    fig.add_trace(go.Scatter(
        x=dates, y=res_v, name="Resolved (daily)",
        mode="lines+markers",
        line=dict(color=C_GREEN, width=2),
        marker=dict(size=4),
        hovertemplate="<b>Resolved</b>: %{y}<br>%{x}<extra></extra>",
    ))
    fig.update_layout(
        height=250,
        paper_bgcolor=CARD, plot_bgcolor=CARD,
        margin=dict(l=8, r=8, t=36, b=8),
        font=dict(family=FONT, size=11, color=TEXT),
        legend=dict(orientation="h", yanchor="bottom", y=1.02,
                    xanchor="left", x=0, font=dict(size=11)),
        yaxis=dict(gridcolor=DIVIDER, tickfont=dict(size=10),
                   zeroline=False, rangemode="tozero"),
        xaxis=dict(gridcolor=DIVIDER, tickfont=dict(size=10), showgrid=False),
        hovermode="x unified",
    )
    return fig


def _trend_figure(trend: list, cat: str, inst_code: str | None = None) -> go.Figure:
    """Build trend chart. When inst_code is given, shows that institution's scores."""
    dates = [e.get("date", "") for e in trend]
    fig   = go.Figure()
    for dim in DIMS:
        if inst_code:
            scores = [float(_inst_scores(e, inst_code).get(dim, 0)) for e in trend]
        else:
            scores = [float(_cat_scores(e, cat).get(dim) or 0) for e in trend]
        fig.add_trace(go.Scatter(
            x=dates, y=scores,
            name=DIM_LABELS[dim],
            mode="lines+markers",
            line=dict(color=DIM_COLORS[dim], width=2),
            marker=dict(size=5, color=DIM_COLORS[dim]),
            hovertemplate=(
                f"<b>{DIM_LABELS[dim]}</b><br>"
                "%{x}<br>%{y:.1f}%<extra></extra>"
            ),
        ))
    fig.update_layout(
        height=250,
        paper_bgcolor=CARD, plot_bgcolor=CARD,
        margin=dict(l=8, r=8, t=36, b=8),
        font=dict(family=FONT, size=11, color=TEXT),
        legend=dict(
            orientation="h", yanchor="bottom", y=1.02,
            xanchor="left", x=0, font=dict(size=11),
        ),
        yaxis=dict(
            range=[0, 100], gridcolor=DIVIDER,
            ticksuffix="%", tickfont=dict(size=10),
            zeroline=False,
        ),
        xaxis=dict(gridcolor=DIVIDER, tickfont=dict(size=10), showgrid=False),
        hovermode="x unified",
    )
    return fig


def _institution_table(institutions: dict, issue_summary: dict | None = None) -> html.Div:
    if not institutions:
        return html.Div(
            "No institution data for this category.",
            style={"color": MUTED, "fontSize": "12px",
                   "padding": "32px", "textAlign": "center"},
        )

    rows = sorted(institutions.items(), key=lambda kv: kv[1].get("overall", 0))

    _isu  = issue_summary or {}
    H = {"fontSize": "11px", "fontWeight": "900", "color": MUTED,
         "textTransform": "uppercase", "letterSpacing": "0.05em",
         "lineHeight": "1.15", "flexShrink": "0"}
    COL_W  = "74px"
    DL_W   = "52px"
    NTFY_W = "36px"

    header = html.Div([
        html.Span("Institution", style={**H, "flex": "1", "flexShrink": "1"}),
        *[html.Span(DIM_LABELS[d][:5], style={**H, "width": COL_W, "textAlign": "center"})
          for d in DIMS],
        html.Span("Overall", style={**H, "width": COL_W, "textAlign": "center"}),
        html.Span("Report",  style={**H, "width": DL_W,  "textAlign": "center"}),
        html.Span("Alert",   style={**H, "width": NTFY_W, "textAlign": "center"}),
    ], style={
        "display": "flex", "alignItems": "center", "gap": "4px",
        "padding": "9px 14px",
        "borderBottom": "2px solid rgba(117,57,24,0.18)",
        "background": BG, "borderRadius": "8px 8px 0 0",
    })

    data_rows = []
    for i, (lb, d) in enumerate(rows):
        name    = (d.get("name") or lb).title()
        overall = float(d.get("overall") or 0)
        bg      = "#c9956c" if i % 2 == 0 else BG

        # urgency left-border based on tracked issues
        isu_info = _isu.get(lb, {})
        worst    = isu_info.get("worst_urgency")
        left_clr = _URGENCY_COLORS.get(worst, "transparent") if worst else "transparent"
        n_issues = isu_info.get("total", 0)

        # institution name cell with urgency dot
        name_children = []
        if worst:
            tip = f"{n_issues} open issue(s) — {_URGENCY_LABELS.get(worst, worst)}"
            name_children.append(html.Span("●", title=tip, style={
                "color": left_clr, "fontSize": "9px",
                "marginRight": "5px", "flexShrink": "0",
            }))
        name_children.append(html.Span(name, title=name, style={
            "fontSize": "12px", "color": TEXT, "lineHeight": "1.15",
            "overflow": "hidden", "textOverflow": "ellipsis",
            "whiteSpace": "nowrap",
        }))

        cells = [
            html.Div(name_children, style={
                "flex": "1", "flexShrink": "1", "display": "flex",
                "alignItems": "center", "minWidth": "0",
                "borderLeft": f"3px solid {left_clr}",
                "paddingLeft": "6px", "marginLeft": "-6px",
            }),
        ]
        for dim in DIMS:
            s = float(d.get(dim) or 0)
            cells.append(html.Span(f"{s:.0f}%", style={
                "width": COL_W, "textAlign": "center", "flexShrink": "0",
                "fontSize": "12px", "fontWeight": "700",
                "color": _score_color(s), "background": _score_bg(s),
                "borderRadius": "4px", "padding": "2px 0",
                "lineHeight": "1.15",
            }))
        cells.append(html.Span(f"{overall:.1f}%", style={
            "width": COL_W, "textAlign": "center", "flexShrink": "0",
            "fontSize": "12px", "fontWeight": "900",
            "color": _score_color(overall), "lineHeight": "1.15",
        }))

        # prefer CSV over XLSX for the date label
        csv_files = sorted(REPORTS_DIR.glob(f"*_{lb}_*.csv"), reverse=True) if REPORTS_DIR.exists() else []
        rpt_files = sorted(REPORTS_DIR.glob(f"{lb}_*.xlsx"),  reverse=True) if REPORTS_DIR.exists() else []
        has_report = bool(csv_files or rpt_files)

        if has_report:
            if csv_files:
                # extract month: {table}_{lb}_{YYYY-MM}.csv
                parts    = csv_files[0].stem.rsplit("_", 2)
                rpt_date = parts[2] if len(parts) == 3 and len(parts[2]) == 7 else "—"
                fmt_label = "CSV"
            else:
                parts    = rpt_files[0].stem.rsplit("_", 3)
                rpt_date = parts[-1] if len(parts) >= 2 and len(parts[-1]) == 10 else "—"
                fmt_label = "XLSX"
            dl_btn = html.Div(
                [html.Span("⬇ ", style={"fontSize": "13px"}),
                 html.Span(rpt_date, style={"fontSize": "9px", "opacity": "0.75"})],
                id={"type": "inst-dl-btn", "index": lb},
                n_clicks=0,
                title=f"Download {name} report ({fmt_label}, {rpt_date})",
                style={
                    "width": DL_W, "textAlign": "center", "flexShrink": "0",
                    "display": "flex", "alignItems": "center", "justifyContent": "center",
                    "cursor": "pointer", "color": BRAND, "userSelect": "none",
                    "lineHeight": "1.15",
                },
            )
        else:
            dl_btn = html.Span("—", style={
                "width": DL_W, "textAlign": "center", "flexShrink": "0",
                "fontSize": "11px", "color": MUTED, "lineHeight": "1.15",
            })
        cells.append(dl_btn)

        # Notify button — shown when institution has open issues
        if n_issues > 0:
            notify_btn = html.Div(
                "🔔",
                id={"type": "notify-btn", "index": lb},
                n_clicks=0,
                title=f"Send reminder to {name} ({n_issues} open issue(s))",
                style={
                    "width": NTFY_W, "textAlign": "center", "flexShrink": "0",
                    "fontSize": "13px", "lineHeight": "1.15",
                    "cursor": "pointer", "userSelect": "none",
                    "color": left_clr,
                },
            )
        else:
            notify_btn = html.Span("", style={"width": NTFY_W, "flexShrink": "0"})
        cells.append(notify_btn)

        data_rows.append(html.Div(cells, className="inst-row", style={
            "display": "flex", "alignItems": "center", "gap": "4px",
            "padding": "7px 14px", "background": bg,
            "borderBottom": "1px solid rgba(117,57,24,0.12)",
        }))

    table = html.Div(
        [header] + data_rows,
        style={"border": f"1px solid {DIVIDER}", "borderRadius": "8px",
               "overflow": "hidden"},
    )
    return html.Div([table])


def _watermark_date() -> str:
    """Return the earliest watermark date (the 'data as of' cut-off)."""
    try:
        wm = json.loads(WATERMARK_FILE.read_text(encoding="utf-8"))
        dates = [str(v)[:10] for v in wm.values() if v]
        return min(dates) if dates else "—"
    except Exception:
        return "—"


def _unscored_section(cat: str, all_categories: dict, scored_lbs: set,
                      activity: dict) -> html.Div:
    """Count card + collapsible table for institutions absent from today's run."""
    if cat == "SACCO":
        types = {"SACCO", "OSACCO"}
    else:
        types = {cat}

    unscored = {
        lb: info for lb, info in all_categories.items()
        if info.get("category_type") in types and lb not in scored_lbs
    }
    n = len(unscored)
    if n == 0:
        return html.Div()

    wm_date = _watermark_date()

    H = {"fontSize": "11px", "fontWeight": "900", "color": MUTED,
         "textTransform": "uppercase", "letterSpacing": "0.05em",
         "lineHeight": "1.15", "flexShrink": "0"}

    # ── count card ─────────────────────────────────────────────────────────────
    count_card = html.Div([
        html.Div("NOT SCORED THIS RUN", style={
            "fontSize": "11px", "fontWeight": "900",
            "color": MUTED, "letterSpacing": "0.06em",
            "textTransform": "uppercase", "lineHeight": "1.15",
        }),
        html.Div(str(n), style={
            "fontSize": "30px", "fontWeight": "700",
            "color": BRAND, "lineHeight": "1.1", "marginTop": "6px",
            "fontVariantNumeric": "tabular-nums",
        }),
        html.Div(
            f"institution{'s' if n != 1 else ''}",
            style={"fontSize": "12px", "color": MUTED, "lineHeight": "1.15"},
        ),
        html.Div([
            html.Span("No data changes since ", style={"color": MUTED}),
            html.Span(wm_date, style={"color": BRAND, "fontWeight": "700"}),
        ], style={"fontSize": "11px", "marginTop": "8px", "lineHeight": "1.4"}),
    ], style={
        "background":   CARD,
        "borderRadius": "8px",
        "padding":      "16px",
        "borderTop":    f"3px solid {MUTED}",
        "boxShadow":    "0 1px 4px rgba(117,57,24,0.08)",
        "display":      "inline-block",
        "minWidth":     "200px",
    })

    # ── table ──────────────────────────────────────────────────────────────────
    header = html.Div([
        html.Span("LE Book",       style={**H, "width": "76px"}),
        html.Span("Institution",   style={**H, "flex": "1"}),
        html.Span("Last Modified", style={**H, "width": "124px", "textAlign": "center"}),
        html.Span("Last Created",  style={**H, "width": "124px", "textAlign": "center"}),
    ], style={
        "display": "flex", "alignItems": "center", "gap": "8px",
        "padding": "9px 14px",
        "borderBottom": f"2px solid rgba(117,57,24,0.18)",
        "background": BG, "borderRadius": "8px 8px 0 0",
    })

    data_rows = []
    for i, (lb, info) in enumerate(
        sorted(unscored.items(), key=lambda kv: (kv[1].get("name") or kv[0]).lower())
    ):
        name     = (info.get("name") or lb).title()
        act      = activity.get(lb, {})
        last_mod = act.get("last_modified") or "—"
        last_cre = act.get("last_created")  or "—"
        bg       = "#c9956c" if i % 2 == 0 else BG

        data_rows.append(html.Div([
            html.Span(lb, style={
                "width": "76px", "flexShrink": "0",
                "fontSize": "11px", "fontWeight": "700",
                "color": BRAND, "fontFamily": "monospace",
            }),
            html.Span(name, style={
                "flex": "1", "fontSize": "12px", "color": TEXT,
                "overflow": "hidden", "textOverflow": "ellipsis", "whiteSpace": "nowrap",
            }),
            html.Span(last_mod, style={
                "width": "124px", "flexShrink": "0",
                "fontSize": "11px", "color": MUTED, "textAlign": "center",
            }),
            html.Span(last_cre, style={
                "width": "124px", "flexShrink": "0",
                "fontSize": "11px", "color": MUTED, "textAlign": "center",
            }),
        ], style={
            "display": "flex", "alignItems": "center", "gap": "8px",
            "padding": "7px 14px", "background": bg,
            "borderBottom": f"1px solid rgba(117,57,24,0.10)",
        }))

    table = html.Div([header] + data_rows, style={
        "border": f"1px solid {DIVIDER}", "borderRadius": "8px",
        "overflow": "hidden", "marginTop": "10px",
    })

    # ── toggle button ──────────────────────────────────────────────────────────
    toggle_btn = html.Div(
        id="unscored-toggle",
        n_clicks=0,
        children=f"▶  Show {n} unscored institution{'s' if n != 1 else ''}",
        style={
            "cursor": "pointer", "color": BRAND,
            "fontSize": "12px", "fontWeight": "700",
            "userSelect": "none", "marginTop": "14px",
            "display": "inline-block",
        },
    )

    return html.Div([
        # divider
        html.Div(style={
            "borderTop": f"1px solid {DIVIDER}", "margin": "24px 0 16px",
        }),
        html.Div([count_card], style={"marginBottom": "4px"}),
        toggle_btn,
        html.Div(
            id="unscored-body",
            children=table,
            style={"display": "none"},
        ),
    ])


def _stale_banner() -> html.Div | None:
    if not _HISTORY:
        return html.Div(
            "No detection data found. Run:  python dq_monthly_detection.py",
            style={
                "background": "#FEF2F2", "border": f"1px solid {C_RED}",
                "borderRadius": "6px", "padding": "10px 16px",
                "fontSize": "12px", "color": C_RED,
                "marginBottom": "16px", "lineHeight": "1.15",
            },
        )
    from datetime import date as _date
    last_date = _HISTORY[-1].get("date", "")
    try:
        days_since = (_date.today() - _date.fromisoformat(last_date)).days
    except Exception:
        days_since = 999
    if days_since > 35:
        return html.Div(
            f"⚠  Last monthly detection: {last_date} — {days_since} days ago. "
            "Run dq_monthly_detection.py to refresh.",
            style={
                "background": "#FFFBEB", "border": "1px solid #F59E0B",
                "borderRadius": "6px", "padding": "10px 16px",
                "fontSize": "12px", "color": "#92400E",
                "marginBottom": "16px", "lineHeight": "1.15",
            },
        )
    return None


def _landing_page(counts: dict) -> html.Div:
    """Full-screen landing page prompting the user to pick a category type."""
    today = _today_entry()
    all_inst = today.get("by_institution", {})

    cards = []
    for cat_def in LANDING_CATS:
        code    = cat_def["code"]
        label   = cat_def["label"]
        subtitle = cat_def["subtitle"]
        color   = cat_def["color"]
        types   = cat_def["types"]

        # count institutions whose category_type matches this landing card
        n_inst = sum(
            1 for d in all_inst.values()
            if d.get("category_type") in types
        )

        r = int(color[1:3], 16)
        g = int(color[3:5], 16)
        b = int(color[5:7], 16)

        cards.append(html.Div(
            id={"type": "cat-landing-btn", "index": code},
            n_clicks=0,
            children=[
                html.Div(label, style={
                    "fontSize": "28px",
                    "fontWeight": "900",
                    "color": color,
                    "lineHeight": "1.1",
                    "marginBottom": "8px",
                    "letterSpacing": "-0.01em",
                }),
                html.Div(subtitle, style={
                    "fontSize": "12px",
                    "color": MUTED,
                    "lineHeight": "1.5",
                    "marginBottom": "24px",
                    "minHeight": "36px",
                }),
                html.Div([
                    html.Span(str(n_inst), style={
                        "fontSize": "36px",
                        "fontWeight": "900",
                        "color": color,
                        "fontVariantNumeric": "tabular-nums",
                        "lineHeight": "1",
                    }),
                    html.Span(
                        " institution" + ("s" if n_inst != 1 else ""),
                        style={"fontSize": "13px", "color": MUTED, "marginLeft": "4px"},
                    ),
                ], style={"marginBottom": "24px"}),
                html.Div("View dashboard →", style={
                    "display": "inline-block",
                    "fontSize": "12px",
                    "fontWeight": "700",
                    "color": CARD,
                    "background": color,
                    "padding": "8px 18px",
                    "borderRadius": "6px",
                }),
            ],
            style={
                "background":   CARD,
                "border":       f"1px solid rgba({r},{g},{b},0.20)",
                "borderTop":    f"4px solid {color}",
                "borderRadius": "10px",
                "padding":      "32px 28px",
                "cursor":       "pointer",
                "flex":         "1",
                "minWidth":     "220px",
                "boxShadow":    "0 2px 8px rgba(117,57,24,0.07)",
                "userSelect":   "none",
                "textAlign":    "left",
                "transition":   "box-shadow .15s",
            },
        ))

    return html.Div([
        html.Div([
            html.Div("Select Category Type", style={
                "fontSize": "26px",
                "fontWeight": "900",
                "color": TEXT,
                "marginBottom": "8px",
                "letterSpacing": "-0.01em",
                "lineHeight": "1.15",
            }),
            html.Div(
                "Choose a financial institution category to explore its data quality metrics.",
                style={
                    "fontSize": "14px",
                    "color": MUTED,
                    "marginBottom": "48px",
                    "lineHeight": "1.5",
                },
            ),
            html.Div(cards, style={
                "display":  "flex",
                "gap":      "24px",
                "flexWrap": "wrap",
            }),
        ], style={
            "maxWidth": "960px",
            "margin":   "80px auto",
            "padding":  "0 24px",
        }),
    ])


def _dashboard_content(cat: str, inst: str | None) -> html.Div:
    """Renders the dashboard for a specific category, optionally filtered to one institution."""
    today        = _today_entry()
    yesterday    = _yesterday_entry()
    trend        = _trend_entries(7)
    banner       = _stale_banner()

    institutions = _filter_institutions(today, cat)
    cat_label    = CAT_LABELS.get(cat, cat)

    # Institution dropdown options — "All [Category]" first, then sorted by name
    inst_options = [{"label": f"All {cat_label}", "value": ""}]
    for code, data in sorted(institutions.items(),
                              key=lambda kv: (kv[1].get("name") or kv[0]).lower()):
        name = (data.get("name") or code).title()
        inst_options.append({"label": name, "value": code})

    # KPI dimension score cards + trend figure
    # (per-institution when one is selected, category-wide otherwise)
    if inst and inst in institutions:
        cards = []
        for dim in DIMS:
            now   = float(_inst_scores(today,     inst).get(dim, 0))
            prev  = float(_inst_scores(yesterday, inst).get(dim, 0))
            delta = round(now - prev, 1)
            spark = [float(_inst_scores(e, inst).get(dim, 0)) for e in trend]
            cards.append(_kpi_card(dim, now, delta, spark))
        dup_now   = _inst_dup_count(today, inst)
        dup_prev  = _inst_dup_count(yesterday, inst)
        dup_spark = [_inst_dup_count(e, inst) for e in trend]
        cards.append(_dup_card(dup_now, dup_now - dup_prev, dup_spark))
        issue_trend_data = _issue_counts_trend(30, le_books={str(inst)})
        display_insts    = {inst: institutions[inst]}
        inst_name        = (institutions[inst].get("name") or inst).title()
        table_title      = f"ISSUE REPORT — {inst_name.upper()}"
    else:
        cards = []
        for dim in DIMS:
            now   = float(_cat_scores(today,     cat).get(dim) or 0)
            prev  = float(_cat_scores(yesterday, cat).get(dim) or 0)
            delta = round(now - prev, 1)
            spark = [float(_cat_scores(e, cat).get(dim) or 0) for e in trend]
            cards.append(_kpi_card(dim, now, delta, spark))
        dup_now   = _cat_dup_count(today, cat)
        dup_prev  = _cat_dup_count(yesterday, cat)
        dup_spark = [_cat_dup_count(e, cat) for e in trend]
        cards.append(_dup_card(dup_now, dup_now - dup_prev, dup_spark))
        cat_lbs          = {lb for lb, d in institutions.items()}
        issue_trend_data = _issue_counts_trend(30, le_books=cat_lbs if cat_lbs else None)
        display_insts    = institutions
        n                = len(institutions)
        table_title      = f"ISSUE REPORTS — {cat_label.upper()}  ({n} institutions)"

    # issue count summary chips (latest day of trend)
    last = issue_trend_data[-1] if issue_trend_data else {}
    def _issue_chip(value, label, color):
        return html.Div([
            html.Span(str(value), style={"fontWeight": "900", "fontSize": "20px",
                                         "color": color}),
            html.Span(label, style={"fontSize": "10px", "color": MUTED,
                                    "marginTop": "2px"}),
        ], style={
            "display": "flex", "flexDirection": "column", "alignItems": "center",
            "background": CARD, "borderRadius": "6px", "padding": "8px 18px",
            "border": f"1px solid {DIVIDER}", "minWidth": "80px",
        })
    issue_chips = html.Div([
        _issue_chip(last.get("open",     0), "Open",            C_AMBER),
        _issue_chip(last.get("overdue",  0), "Overdue",         _URGENCY_COLORS["overdue"]),
        _issue_chip(last.get("resolved", 0), "Resolved Today",  C_GREEN),
    ], style={"display": "flex", "gap": "8px", "marginBottom": "10px"})

    issue_fig = _issue_trend_figure(issue_trend_data)

    return html.Div([
        banner if banner else html.Div(),

        # breadcrumb row
        html.Div([
            html.Span(
                "← All Categories",
                id={"type": "nav-action", "index": "back"},
                n_clicks=0,
                style={
                    "cursor":     "pointer",
                    "color":      BRAND,
                    "fontSize":   "12px",
                    "fontWeight": "700",
                    "userSelect": "none",
                },
            ),
            html.Span(" / ", style={
                "color": MUTED, "margin": "0 8px", "fontSize": "12px",
            }),
            html.Span(cat_label, style={"fontSize": "12px", "color": MUTED}),
        ], style={"marginBottom": "16px", "display": "flex", "alignItems": "center"}),

        # category header + institution filter row
        html.Div([
            html.Div([
                html.Div(cat_label, style={
                    "fontSize":   "20px",
                    "fontWeight": "900",
                    "color":      TEXT,
                    "lineHeight": "1.15",
                }),
                html.Div(
                    f"{len(institutions)} institution" + ("s" if len(institutions) != 1 else ""),
                    style={"fontSize": "12px", "color": MUTED, "marginTop": "3px"},
                ),
            ]),
            html.Div([
                html.Span("Filter by institution:", style={
                    "fontSize":      "11px",
                    "fontWeight":    "900",
                    "color":         MUTED,
                    "textTransform": "uppercase",
                    "letterSpacing": "0.05em",
                    "marginRight":   "10px",
                    "whiteSpace":    "nowrap",
                }),
                dcc.Dropdown(
                    id={"type": "inst-dd", "index": "main"},
                    options=inst_options,
                    value=inst or "",
                    clearable=False,
                    style={
                        "fontSize":   "12px",
                        "fontFamily": FONT,
                        "minWidth":   "280px",
                    },
                ),
            ], style={"display": "flex", "alignItems": "center"}),
        ], style={
            "display":        "flex",
            "alignItems":     "center",
            "justifyContent": "space-between",
            "flexWrap":       "wrap",
            "gap":            "12px",
            "marginBottom":   "20px",
            "background":     CARD,
            "padding":        "16px 20px",
            "borderRadius":   "8px",
            "border":         f"1px solid {DIVIDER}",
            "boxShadow":      "0 1px 4px rgba(117,57,24,0.06)",
        }),

        # KPI cards + trend chart
        html.Div([
            html.Div(id="kpi-row", children=cards, style={
                "display":      "flex",
                "gap":          "12px",
                "flexWrap":     "wrap",
                "marginBottom": "16px",
            }),
            html.Div([
                html.Div([
                    html.Span("30-DAY ISSUE TREND", style={
                        "fontSize":      "11px",
                        "fontWeight":    "900",
                        "color":         MUTED,
                        "letterSpacing": "0.06em",
                        "textTransform": "uppercase",
                        "lineHeight":    "1.15",
                    }),
                    html.Span(
                        " — open / overdue / resolved per day",
                        style={"fontSize": "10px", "color": MUTED},
                    ),
                ], style={"marginBottom": "8px"}),
                issue_chips,
                dcc.Graph(id="trend-graph", figure=issue_fig,
                          config={"displayModeBar": False}),
            ]),
        ], style={
            "background":   CARD,
            "padding":      "20px",
            "borderRadius": "8px",
            "boxShadow":    "0 2px 8px rgba(117,57,24,0.07)",
            "border":       f"1px solid {DIVIDER}",
            "marginBottom": "20px",
        }),

        # Issue reports table
        html.Div([
            html.Div(table_title, id="table-title", style={
                "fontSize":      "12px",
                "fontWeight":    "900",
                "color":         TEXT,
                "letterSpacing": "0.03em",
                "marginBottom":  "12px",
                "lineHeight":    "1.15",
            }),
            _institution_reports_table(display_insts),
        ]),

        # Unscored institutions section
        _unscored_section(
            cat,
            _load_all_categories(),
            set(institutions.keys()),
            _load_activity(),
        ),

    ])


def _institution_reports_table(display_insts: dict) -> html.Div:
    """Issue reports table — one row per institution, ZIP download button."""
    import zipfile as _zf

    issue_dir = Path(__file__).parent / "issue_reports"

    H = {"fontSize": "11px", "fontWeight": "900", "color": MUTED,
         "textTransform": "uppercase", "letterSpacing": "0.05em",
         "lineHeight": "1.15", "flexShrink": "0"}

    header = html.Div([
        html.Span("Code",        style={**H, "width": "56px"}),
        html.Span("Institution", style={**H, "flex": "1"}),
        html.Span("Tables with issues", style={**H, "width": "220px", "textAlign": "left"}),
        html.Span("ZIP size",    style={**H, "width": "74px", "textAlign": "right"}),
        html.Span("Report",      style={**H, "width": "66px", "textAlign": "center"}),
    ], style={
        "display": "flex", "alignItems": "center", "gap": "10px",
        "padding": "9px 14px",
        "borderBottom": "2px solid rgba(117,57,24,0.18)",
        "background": BG, "borderRadius": "8px 8px 0 0",
    })

    # Build rows — institutions with ZIPs first (sorted by size desc), then clean ones
    with_zip = []
    no_zip   = []

    for lb, data in display_insts.items():
        name = (data.get("name") or lb).title()
        zips = sorted(issue_dir.glob(f"{lb}_*.zip"), reverse=True) if issue_dir.exists() else []
        if zips:
            zp = zips[0]
            try:
                size_kb    = zp.stat().st_size // 1024
                namelist   = _zf.ZipFile(zp).namelist()
                table_names = [n.rsplit(".", 1)[0] for n in namelist]
            except Exception:
                size_kb, table_names = 0, []
            with_zip.append((lb, name, size_kb, table_names, zp))
        else:
            no_zip.append((lb, name))

    with_zip.sort(key=lambda x: -x[2])

    data_rows = []
    for i, (lb, name, size_kb, table_names, zp) in enumerate(with_zip):
        bg = "#c9956c" if i % 2 == 0 else BG

        # table chips
        table_chips = html.Div(
            [html.Span(t, style={
                "fontSize": "9px", "fontWeight": "700",
                "background": "rgba(117,57,24,0.10)", "color": BRAND,
                "borderRadius": "3px", "padding": "2px 5px",
                "whiteSpace": "nowrap",
            }) for t in table_names],
            style={"display": "flex", "flexWrap": "wrap", "gap": "3px",
                   "width": "220px", "flexShrink": "0"},
        )

        size_label = f"{size_kb:,} KB" if size_kb >= 1 else "< 1 KB"

        dl_btn = html.Div("⬇ ZIP",
            id={"type": "open-issue-dl-btn", "index": lb},
            n_clicks=0,
            title=f"Download issue report for {name}",
            style={
                "cursor": "pointer", "background": BRAND, "color": CARD,
                "fontSize": "10px", "fontWeight": "700",
                "padding": "3px 10px", "borderRadius": "4px",
                "userSelect": "none", "width": "66px",
                "textAlign": "center", "flexShrink": "0",
            },
        )

        data_rows.append(html.Div([
            html.Span(lb, style={"fontWeight": "900", "fontSize": "12px",
                "color": TEXT, "width": "56px", "flexShrink": "0"}),
            html.Span(name, style={"fontSize": "12px", "color": TEXT,
                "flex": "1", "overflow": "hidden",
                "textOverflow": "ellipsis", "whiteSpace": "nowrap"}),
            table_chips,
            html.Span(size_label, style={"fontSize": "11px", "color": MUTED,
                "width": "74px", "textAlign": "right", "flexShrink": "0"}),
            dl_btn,
        ], style={
            "display": "flex", "alignItems": "center", "gap": "10px",
            "padding": "8px 14px", "background": bg,
            "borderBottom": f"1px solid rgba(117,57,24,0.12)",
        }))

    # clean institutions (no ZIP) — greyed out at bottom
    for i, (lb, name) in enumerate(sorted(no_zip, key=lambda x: x[1])):
        bg = CARD if i % 2 == 0 else BG
        data_rows.append(html.Div([
            html.Span(lb, style={"fontWeight": "900", "fontSize": "12px",
                "color": MUTED, "width": "56px", "flexShrink": "0"}),
            html.Span(name, style={"fontSize": "12px", "color": MUTED,
                "flex": "1", "overflow": "hidden",
                "textOverflow": "ellipsis", "whiteSpace": "nowrap"}),
            html.Span("No issues found", style={"fontSize": "11px",
                "color": C_GREEN, "fontWeight": "700",
                "width": "220px", "flexShrink": "0"}),
            html.Span("—", style={"fontSize": "11px", "color": MUTED,
                "width": "74px", "textAlign": "right", "flexShrink": "0"}),
            html.Span("—", style={"fontSize": "11px", "color": MUTED,
                "width": "66px", "textAlign": "center", "flexShrink": "0"}),
        ], style={
            "display": "flex", "alignItems": "center", "gap": "10px",
            "padding": "8px 14px", "background": bg,
            "borderBottom": f"1px solid rgba(117,57,24,0.12)",
        }))

    if not data_rows:
        return html.Div("No institution data for this category.",
            style={"color": MUTED, "fontSize": "12px",
                   "padding": "32px", "textAlign": "center"})

    return html.Div(
        [header] + data_rows,
        style={"border": f"1px solid {DIVIDER}", "borderRadius": "8px",
               "overflow": "hidden"},
    )


# ── alerts page ───────────────────────────────────────────────────────────────

def _issues_to_xlsx(issues: list) -> bytes:
    """Build an in-memory XLSX workbook from a list of resolved issue dicts."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import date as _date

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resolved Issues"

    hdr_fill  = PatternFill("solid", fgColor="753918")
    hdr_font  = Font(color="FFFFFF", bold=True, size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers    = ["Institution", "Table", "Rule ID", "Dimension",
                  "Failing Rows", "Detected", "Resolved", "Days to Fix"]
    col_widths = [30, 28, 12, 14, 12, 12, 12, 12]

    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22

    for ri, iss in enumerate(issues, start=2):
        try:
            days_fix = (_date.fromisoformat(iss.get("resolved_at", "")) -
                        _date.fromisoformat(iss.get("detected_at", ""))).days
            days_str = f"{days_fix}d"
        except Exception:
            days_str = ""

        ws.append([
            (iss.get("institution_name") or iss["le_book"]).title(),
            iss.get("table_name", ""),
            iss.get("rule_id", ""),
            iss.get("dimension", "").title(),
            iss.get("failing_rows", ""),
            iss.get("detected_at", ""),
            iss.get("resolved_at", ""),
            days_str,
        ])
        if ri % 2 == 0:
            fill = PatternFill("solid", fgColor="F2EDE9")
            for ci in range(1, len(headers) + 1):
                ws.cell(row=ri, column=ci).fill = fill

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _freshness_section() -> html.Div:
    """Read watermark.json and render a per-table data freshness summary."""
    from datetime import date as _date
    import json as _json

    try:
        wm = _json.loads(WATERMARK_FILE.read_text())
    except Exception:
        return html.Div()

    today = _date.today()
    rows  = []
    for table, wm_val in sorted(wm.items()):
        try:
            last = _date.fromisoformat(str(wm_val)[:10])
            days = (today - last).days
        except Exception:
            continue

        if days <= 1:
            color, badge, bg = "#B8860B", "Fresh", "rgba(184,134,11,.07)"
        elif days <= 3:
            color, badge, bg = "#A0784A", f"{days}d ago", "rgba(160,120,74,.07)"
        else:
            color, badge, bg = "#7C3D1E", f"{days}d ago", "rgba(124,61,30,.07)"

        rows.append(html.Div([
            html.Span(table, style={
                "flex": "1", "fontSize": "12px", "color": TEXT,
                "fontFamily": "monospace", "padding": "7px 12px",
            }),
            html.Span(str(wm_val)[:10], style={
                "width": "110px", "fontSize": "11px", "color": MUTED,
                "padding": "7px 10px",
            }),
            html.Span(badge, style={
                "width": "80px", "fontSize": "11px", "fontWeight": "700",
                "color": color, "padding": "7px 10px", "textAlign": "center",
            }),
        ], style={
            "display": "flex", "alignItems": "center",
            "background": bg, "borderBottom": f"1px solid {DIVIDER}",
        }))

    any_stale = any(
        (today - _date.fromisoformat(str(v)[:10])).days > 1
        for v in wm.values()
        if str(v)[:10]
    )

    H = {"fontSize": "11px", "fontWeight": "900", "color": MUTED,
         "textTransform": "uppercase", "letterSpacing": "0.05em", "padding": "8px 12px"}
    header = html.Div([
        html.Span("Table",        style={**H, "flex": "1"}),
        html.Span("Last Updated", style={**H, "width": "110px"}),
        html.Span("Freshness",    style={**H, "width": "80px", "textAlign": "center"}),
    ], style={
        "display": "flex", "background": BG,
        "borderRadius": "8px 8px 0 0", "borderBottom": f"2px solid {DIVIDER}",
    })

    title_color = "#7C3D1E" if any_stale else "#B8860B"
    title_label = "⚠ Some tables have stale data" if any_stale else "✓ All tables up to date"

    return html.Div([
        html.Div([
            html.H3("Data Freshness", style={
                "fontSize": "14px", "fontWeight": "900", "color": TEXT,
                "margin": "0", "display": "inline",
            }),
            html.Span(f"  {title_label}", style={
                "fontSize": "11px", "color": title_color,
                "marginLeft": "10px", "fontWeight": "700",
            }),
        ], style={"marginBottom": "10px"}),
        html.Div([header, *rows], style={
            "background": CARD, "borderRadius": "8px",
            "border": f"1px solid {DIVIDER}", "marginBottom": "28px",
        }),
    ])


def _urgency_dot(band: str) -> html.Span:
    col = _URGENCY_COLORS.get(band, MUTED)
    label = {"new": "New", "attention": "Attention", "urgent": "Urgent",
             "critical": "Critical", "overdue": "⚠ Overdue"}.get(band, band.title())
    return html.Span([
        html.Span("●", style={"color": col, "fontSize": "9px", "marginRight": "4px"}),
        html.Span(label, style={"color": col, "fontSize": "11px", "fontWeight": "700"}),
    ])


def _build_table_issue_sections(issues_by_table: dict, status: str,
                                cat_filter: str = "") -> html.Div:
    """
    Render Alerts page: Table → Institution (mini score card) → Rules.
    cat_filter: '' = all, 'B' = banks, 'MF' = MFIs, 'SACCO' = SACCOs+OSACCOs.
    """
    if not issues_by_table:
        label = "open issues" if status == "open" else "resolved issues"
        return html.Div(f"No {label}.", style={"color": MUTED, "padding": "24px"})

    # load categories for type-lookup
    try:
        _cats = json.loads(CATEGORIES_FILE.read_text())
    except Exception:
        _cats = {}
    cat_type_map = {str(lb): (info.get("category_type") or "") for lb, info in _cats.items()}

    today_entry = _today_entry()
    _BO = ["new", "attention", "urgent", "critical", "overdue"]

    # issue_id → {cr_id, status} for active CRs
    try:
        from dq_change_request import get_issue_cr_map
        _cr_map = get_issue_cr_map()
    except Exception:
        _cr_map = {}

    _CR_STATUS_CLR = {
        "open":        "#2563EB",
        "in_progress": "#D97706",
        "submitted":   "#7C3D1E",
        "approved":    "#16A34A",
        "rejected":    "#DC2626",
    }

    def _worst(a: str, b: str) -> str:
        return a if _BO.index(a) >= _BO.index(b) else b

    sections = []
    for table, rules in issues_by_table.items():
        # ── pivot: table → {le_book: {name, rules, total_rows, worst_urgency}} ──
        inst_map: dict[str, dict] = {}
        for rule in rules:
            for inst in rule["institutions"]:
                lb  = inst["le_book"]
                ct  = cat_type_map.get(str(lb), "")
                # category filter  (SACCO covers OSACCO)
                if cat_filter:
                    if cat_filter == "SACCO" and ct not in ("SACCO", "OSACCO"):
                        continue
                    elif cat_filter != "SACCO" and ct != cat_filter:
                        continue
                if lb not in inst_map:
                    inst_map[lb] = {
                        "name":          inst["institution_name"],
                        "rules":         [],
                        "total_rows":    0,
                        "worst_urgency": "new",
                        "any_pending":   False,
                    }
                inst_map[lb]["rules"].append({
                    "rule_id":        rule["rule_id"],
                    "rule_name":      rule["rule_name"],
                    "dimension":      rule["dimension"],
                    "failing_rows":   inst["failing_rows"],
                    "urgency_band":   inst["urgency_band"],
                    "days_left":      inst["days_left"],
                    "issue_id":       inst.get("issue_id", ""),
                    "recurrence_count": inst.get("recurrence_count", 0),
                    "pending":        inst.get("pending", False),
                })
                inst_map[lb]["total_rows"]   += inst["failing_rows"]
                inst_map[lb]["worst_urgency"] = _worst(
                    inst_map[lb]["worst_urgency"], inst["urgency_band"])
                if inst.get("pending"):
                    inst_map[lb]["any_pending"] = True

        if not inst_map:
            continue

        total_rows  = sum(d["total_rows"] for d in inst_map.values())
        worst_tbl   = max(inst_map.values(), key=lambda x: _BO.index(x["worst_urgency"]))["worst_urgency"]
        band_col    = _URGENCY_COLORS.get(worst_tbl, MUTED)
        tbl_label   = TABLE_NAMES_PRETTY.get(table, table.replace("_", " ").title())
        n_inst      = len(inst_map)

        # ── table section header ─────────────────────────────────────────────
        tbl_header = html.Div([
            html.Span("●", style={"color": band_col, "fontSize": "10px",
                                   "marginRight": "8px"}),
            html.Span(tbl_label, style={
                "fontSize": "14px", "fontWeight": "900", "color": TEXT,
                "letterSpacing": "0.02em",
            }),
            html.Span(
                f"{n_inst} institution{'s' if n_inst != 1 else ''}",
                style={"fontSize": "11px", "color": MUTED, "marginLeft": "12px"},
            ),
            html.Span(f"{total_rows:,} failing rows", style={
                "fontSize": "11px", "color": C_RED, "fontWeight": "700",
                "marginLeft": "10px",
            }),
        ], style={
            "display": "flex", "alignItems": "center",
            "padding": "11px 18px",
            "background": BG,
            "borderLeft": f"4px solid {band_col}",
            "borderRadius": "6px 6px 0 0",
        })

        # ── institution rows (sorted by worst urgency then failing rows) ─────
        inst_divs = []
        sort_key  = lambda kv: (_BO.index(kv[1]["worst_urgency"]), -kv[1]["total_rows"])
        for lb, idata in sorted(inst_map.items(), key=sort_key, reverse=True):
            urg      = idata["worst_urgency"]
            urg_col  = _URGENCY_COLORS.get(urg, MUTED)
            cidx     = f"ait_{table}__{lb}".replace("-", "_")

            # mini dimension score chips
            i_scores  = _inst_scores(today_entry, lb)
            score_chips = []
            for dim in DIMS:
                s = float(i_scores.get(dim, 0))
                score_chips.append(html.Span([
                    html.Span(dim[:4].title(),
                              style={"fontSize": "9px", "color": MUTED,
                                     "marginRight": "2px"}),
                    html.Span(f"{s:.0f}%",
                              style={"fontSize": "11px", "fontWeight": "700",
                                     "color": _score_color(s)}),
                ], style={"marginRight": "12px", "whiteSpace": "nowrap"}))

            # badges: recurrence + pending-resolution
            badges = []
            max_recurrence = max(
                (r.get("recurrence_count", 0) for r in idata["rules"]), default=0
            )
            if max_recurrence > 0:
                badges.append(html.Span(
                    f"↺ #{max_recurrence + 1} recurrence",
                    style={"fontSize": "9px", "fontWeight": "700",
                           "color": C_RED, "background": "rgba(220,38,38,.10)",
                           "border": "1px solid rgba(220,38,38,.30)",
                           "borderRadius": "3px", "padding": "1px 5px",
                           "marginRight": "6px", "whiteSpace": "nowrap"},
                ))
            if idata.get("any_pending"):
                badges.append(html.Span(
                    "⟳ Confirming",
                    style={"fontSize": "9px", "fontWeight": "700",
                           "color": "#D97706", "background": "rgba(217,119,6,.10)",
                           "border": "1px solid rgba(217,119,6,.30)",
                           "borderRadius": "3px", "padding": "1px 5px",
                           "marginRight": "6px", "whiteSpace": "nowrap"},
                ))
            # CR badge: pick the most advanced active CR for any issue in this inst+table
            inst_issue_ids = [r.get("issue_id") for r in idata["rules"] if r.get("issue_id")]
            cr_info = next((_cr_map[iid] for iid in inst_issue_ids if iid in _cr_map), None)
            if cr_info:
                cr_clr = _CR_STATUS_CLR.get(cr_info["status"], MUTED)
                badges.append(html.Span(
                    f"CR: {cr_info['status'].replace('_', ' ').title()}",
                    style={"fontSize": "9px", "fontWeight": "700",
                           "color": cr_clr, "background": f"rgba(0,0,0,.05)",
                           "border": f"1px solid {cr_clr}",
                           "borderRadius": "3px", "padding": "1px 5px",
                           "marginRight": "6px", "whiteSpace": "nowrap"},
                ))

            inst_header = html.Div([
                _urgency_dot(urg),
                html.Span(idata["name"], style={
                    "fontSize": "13px", "fontWeight": "700",
                    "color": TEXT, "marginLeft": "8px",
                }),
                html.Span(f"({lb})", style={
                    "fontSize": "10px", "color": MUTED,
                    "fontFamily": "monospace", "marginLeft": "6px",
                    "marginRight": "10px",
                }),
                html.Div(badges, style={"display": "flex", "alignItems": "center",
                                        "flexShrink": "0"}),
                html.Span(style={"flex": "1"}),
                html.Div(score_chips, style={
                    "display": "flex", "alignItems": "center",
                    "marginRight": "18px",
                }),
                html.Span(f"{idata['total_rows']:,} rows", style={
                    "fontSize": "12px", "fontWeight": "700",
                    "color": C_RED, "marginRight": "14px",
                }),
                html.Div("⬇ ZIP",
                    id={"type": "open-issue-dl-btn", "index": lb},
                    n_clicks=0,
                    title=f"Download issue report for {idata['name']}",
                    style={
                        "cursor": "pointer", "background": BRAND, "color": CARD,
                        "fontSize": "10px", "fontWeight": "700",
                        "padding": "3px 10px", "borderRadius": "4px",
                        "userSelect": "none", "marginRight": "10px",
                        "flexShrink": "0",
                    },
                ),
                html.Div("▶", id={"type": "alert-inst-toggle", "index": cidx},
                         n_clicks=0,
                         style={"fontSize": "10px", "color": MUTED,
                                "cursor": "pointer", "userSelect": "none",
                                "width": "18px", "textAlign": "center"}),
            ], style={
                "display": "flex", "alignItems": "center",
                "padding": "9px 18px",
                "background": CARD,
                "borderBottom": f"1px solid {DIVIDER}",
                "borderLeft": f"3px solid {urg_col}",
                "cursor": "pointer",
            })

            # collapsed rule detail panel
            _RH = {"fontSize": "10px", "fontWeight": "900", "color": MUTED,
                   "textTransform": "uppercase", "letterSpacing": "0.04em",
                   "padding": "5px 12px", "whiteSpace": "nowrap"}
            rule_hdr = html.Div([
                html.Span("Rule",      style={**_RH, "width": "90px"}),
                html.Span("Issue",     style={**_RH, "flex": "1"}),
                html.Span("Dimension", style={**_RH, "width": "105px"}),
                html.Span("Rows",      style={**_RH, "width": "70px",
                                              "textAlign": "right"}),
                html.Span("SLA",       style={**_RH, "width": "90px",
                                              "textAlign": "right"}),
            ], style={"display": "flex", "background": BG,
                      "borderBottom": f"1px solid {DIVIDER}"})

            rule_rows = [rule_hdr]
            for j, r in enumerate(
                sorted(idata["rules"],
                       key=lambda x: _BO.index(x["urgency_band"]), reverse=True)
            ):
                dl  = r["days_left"]
                ov  = r["urgency_band"] == "overdue"
                dlc = C_RED if dl <= 5 or ov else TEXT
                dls = f"⚠ {abs(dl)}d over" if ov else f"{dl}d left"
                # per-rule: pending indicator + CR badge
                rule_badges = []
                if r.get("pending"):
                    rule_badges.append(html.Span(
                        "⟳", title="Pending full-scan confirmation",
                        style={"color": "#D97706", "fontSize": "10px",
                               "marginRight": "4px"},
                    ))
                rule_cr = _cr_map.get(r.get("issue_id", ""))
                if rule_cr:
                    rcr_clr = _CR_STATUS_CLR.get(rule_cr["status"], MUTED)
                    rule_badges.append(html.Span(
                        rule_cr["cr_id"],
                        style={"fontSize": "9px", "fontWeight": "700",
                               "color": rcr_clr, "marginRight": "4px"},
                    ))

                rule_rows.append(html.Div([
                    html.Span(r["rule_id"], style={
                        "width": "90px", "fontSize": "11px", "fontWeight": "900",
                        "color": BRAND, "fontFamily": "monospace",
                        "padding": "6px 12px",
                    }),
                    html.Div([
                        html.Span(r["rule_name"], style={
                            "fontSize": "12px", "color": TEXT,
                        }),
                        html.Div(rule_badges, style={"display": "inline-flex",
                                                      "alignItems": "center",
                                                      "marginLeft": "6px"}),
                    ], style={"flex": "1", "padding": "6px 12px",
                              "display": "flex", "alignItems": "center"}),
                    html.Div(_dim_pill(r["dimension"]),
                             style={"width": "105px", "padding": "4px 12px"}),
                    html.Span(f"{r['failing_rows']:,}", style={
                        "width": "70px", "fontSize": "12px", "fontWeight": "700",
                        "color": MUTED if r.get("pending") else C_RED,
                        "textAlign": "right", "padding": "6px 12px",
                    }),
                    html.Span(dls, style={
                        "width": "90px", "fontSize": "11px", "fontWeight": "700",
                        "color": dlc, "textAlign": "right", "padding": "6px 12px",
                    }),
                ], style={
                    "display": "flex", "alignItems": "center",
                    "background": "rgba(244,246,249,0.8)" if j % 2 == 0 else CARD,
                    "borderBottom": f"1px solid {DIVIDER}",
                    "opacity": "0.75" if r.get("pending") else "1",
                }))

            inst_divs.append(html.Div([
                inst_header,
                html.Div(
                    html.Div(rule_rows),
                    id={"type": "alert-inst-collapse", "index": cidx},
                    style={"display": "none"},
                ),
            ]))

        sections.append(html.Div(
            [tbl_header, *inst_divs],
            style={"marginBottom": "16px", "border": f"1px solid {DIVIDER}",
                   "borderRadius": "6px", "overflow": "hidden"},
        ))

    if not sections:
        return html.Div("No issues for the selected category.",
                        style={"color": MUTED, "padding": "24px"})
    return html.Div(sections)


def _build_resolved_rows(issues: list) -> html.Div:
    """Flat list for resolved issues."""
    from datetime import date as _date
    today = _date.today()
    if not issues:
        return html.Div("No resolved issues.", style={"color": MUTED, "padding": "20px"})

    H = {"fontSize": "11px", "fontWeight": "900", "color": MUTED,
         "textTransform": "uppercase", "letterSpacing": "0.05em", "padding": "8px 10px"}
    hdr = html.Div([
        html.Span("Institution", style={**H, "flex": "1"}),
        html.Span("Table",       style={**H, "width": "160px"}),
        html.Span("Rule",        style={**H, "width": "80px"}),
        html.Span("Dimension",   style={**H, "width": "100px"}),
        html.Span("Detected",    style={**H, "width": "96px"}),
        html.Span("Resolved",    style={**H, "width": "96px"}),
        html.Span("Days to Fix", style={**H, "width": "80px", "textAlign": "center"}),
        html.Span("Timing",      style={**H, "width": "76px", "textAlign": "center"}),
    ], style={"display": "flex", "background": BG,
              "borderRadius": "8px 8px 0 0", "borderBottom": f"2px solid {DIVIDER}"})

    rows = []
    for i, iss in enumerate(issues):
        detected  = iss.get("detected_at", "—")
        resolved  = iss.get("resolved_at",  "—")
        deadline  = iss.get("sla_deadline", "")
        try:
            days_fix = (_date.fromisoformat(resolved) - _date.fromisoformat(detected)).days
            fix_str  = f"{days_fix}d"
            fix_clr  = C_GREEN if days_fix <= 7 else (C_RED if days_fix >= 20 else TEXT)
        except Exception:
            fix_str, fix_clr = "—", MUTED
        try:
            on_time = _date.fromisoformat(resolved) <= _date.fromisoformat(deadline)
            timing_label = "On Time"
            timing_color = C_GREEN
        except Exception:
            on_time = None
            timing_label, timing_color = "—", MUTED
        if on_time is False:
            timing_label = "Late"
            timing_color = C_RED
        name = (iss.get("institution_name") or iss["le_book"]).title()
        rows.append(html.Div([
            html.Div([html.Span(name, style={"fontSize": "12px", "color": TEXT})],
                     style={"flex": "1", "padding": "7px 10px"}),
            html.Span(iss["table_name"],       style={"width": "160px", "fontSize": "11px", "color": MUTED, "padding": "7px 10px"}),
            html.Span(iss["rule_id"],           style={"width": "80px",  "fontSize": "11px", "fontWeight": "700", "color": TEXT, "padding": "7px 10px"}),
            html.Span(iss["dimension"].title(), style={"width": "100px", "fontSize": "11px", "color": MUTED, "padding": "7px 10px"}),
            html.Span(detected,                style={"width": "96px",  "fontSize": "11px", "color": MUTED, "padding": "7px 10px"}),
            html.Span(resolved,                style={"width": "96px",  "fontSize": "11px", "color": C_GREEN, "padding": "7px 10px"}),
            html.Span(fix_str,                 style={"width": "80px",  "fontSize": "12px", "fontWeight": "700", "color": fix_clr, "textAlign": "center", "padding": "7px 10px"}),
            html.Span(timing_label,            style={"width": "76px",  "fontSize": "11px", "fontWeight": "700", "color": timing_color, "textAlign": "center", "padding": "7px 10px"}),
        ], style={
            "display": "flex", "alignItems": "center",
            "background": "#C9956C" if i % 2 == 0 else BG,
            "borderBottom": f"1px solid {DIVIDER}",
        }))
    return html.Div([hdr, *rows], style={
        "background": CARD, "borderRadius": "8px",
        "border": f"1px solid {DIVIDER}", "marginBottom": "8px",
    })


def _build_resolved_by_institution(issues: list, cat_filter: str = "") -> html.Div:
    """Resolved issues grouped by institution with collapsible rule detail and per-inst ZIP download."""
    from datetime import date as _date
    from collections import defaultdict

    if not issues:
        return html.Div("No resolved issues.", style={"color": MUTED, "padding": "24px"})

    try:
        _cats = json.loads(CATEGORIES_FILE.read_text())
    except Exception:
        _cats = {}

    if cat_filter:
        _SACCO_TYPES = {"SACCO", "OSACCO"}
        def _matches(lb: str) -> bool:
            ct = (_cats.get(str(lb), {}).get("category_type") or "").upper()
            return ct in _SACCO_TYPES if cat_filter == "SACCO" else ct == cat_filter
        issues = [i for i in issues if _matches(i["le_book"])]

    if not issues:
        return html.Div("No resolved issues for the selected category.",
                        style={"color": MUTED, "padding": "24px"})

    by_inst: dict[str, list] = defaultdict(list)
    for iss in issues:
        by_inst[iss["le_book"]].append(iss)

    inst_order = sorted(
        by_inst.keys(),
        key=lambda lb: max((i.get("resolved_at") or "") for i in by_inst[lb]),
        reverse=True,
    )

    H = {"fontSize": "11px", "fontWeight": "900", "color": MUTED,
         "textTransform": "uppercase", "letterSpacing": "0.05em", "padding": "8px 10px"}

    _CAT_CLR = {"B": "#2563EB", "MF": "#753918", "SACCO": "#16A34A", "OSACCO": "#16A34A"}

    cards = []
    for cidx, lb in enumerate(inst_order):
        inst_issues = sorted(by_inst[lb],
                             key=lambda x: x.get("resolved_at") or "", reverse=True)
        inst_info   = _cats.get(str(lb), {})
        name        = (inst_issues[0].get("institution_name") or lb).title()
        cat_type    = (inst_info.get("category_type") or "").upper()
        tables      = sorted({i["table_name"] for i in inst_issues})
        latest_res  = max((i.get("resolved_at") or "") for i in inst_issues)

        fix_days = []
        for i in inst_issues:
            try:
                fix_days.append(
                    (_date.fromisoformat(i["resolved_at"]) -
                     _date.fromisoformat(i["detected_at"])).days
                )
            except Exception:
                pass
        avg_fix = f"{sum(fix_days) // len(fix_days)}d avg" if fix_days else "—"

        cat_clr = _CAT_CLR.get(cat_type, MUTED)

        header = html.Div([
            html.Div([
                html.Span("▶",
                          id={"type": "res-inst-toggle", "index": cidx},
                          n_clicks=0,
                          style={"cursor": "pointer", "color": MUTED, "fontSize": "10px",
                                 "marginRight": "10px", "userSelect": "none",
                                 "flexShrink": "0"}),
                html.Span(name, style={"fontWeight": "700", "fontSize": "13px",
                                       "color": TEXT}),
                html.Span(cat_type, style={
                    "fontSize": "10px", "fontWeight": "700", "color": cat_clr,
                    "border": f"1px solid {cat_clr}", "borderRadius": "3px",
                    "padding": "1px 6px", "marginLeft": "8px",
                }),
                html.Span(f"le_book: {lb}", style={
                    "fontSize": "10px", "color": MUTED, "marginLeft": "8px",
                }),
            ], style={"display": "flex", "alignItems": "center", "flex": "1",
                      "minWidth": "0"}),

            html.Div([
                html.Span(f"{len(inst_issues)} issue{'s' if len(inst_issues) != 1 else ''}",
                          style={"fontSize": "11px", "fontWeight": "700",
                                 "color": C_GREEN, "whiteSpace": "nowrap"}),
                html.Span("  ·  ", style={"color": DIVIDER, "fontSize": "11px"}),
                html.Span(
                    ", ".join(t.replace("_", " ") for t in tables[:3])
                    + ("…" if len(tables) > 3 else ""),
                    style={"fontSize": "11px", "color": MUTED, "whiteSpace": "nowrap"},
                ),
                html.Span("  ·  ", style={"color": DIVIDER, "fontSize": "11px"}),
                html.Span(f"resolved {latest_res}",
                          style={"fontSize": "11px", "color": C_GREEN,
                                 "whiteSpace": "nowrap"}),
                html.Span("  ·  ", style={"color": DIVIDER, "fontSize": "11px"}),
                html.Span(avg_fix, style={"fontSize": "11px", "color": MUTED,
                                          "whiteSpace": "nowrap"}),
            ], style={"display": "flex", "alignItems": "center", "gap": "2px",
                      "flexWrap": "nowrap", "overflow": "hidden"}),

            html.Div("⬇ ZIP",
                     id={"type": "res-dl-btn", "index": lb},
                     n_clicks=0,
                     style={"cursor": "pointer", "background": BRAND, "color": CARD,
                            "fontSize": "10px", "fontWeight": "700",
                            "padding": "4px 12px", "borderRadius": "4px",
                            "userSelect": "none", "marginLeft": "16px",
                            "flexShrink": "0"}),
        ], style={
            "display": "flex", "alignItems": "center",
            "padding": "10px 14px", "background": CARD, "gap": "12px",
        })

        tbl_hdr = html.Div([
            html.Span("Table",        style={**H, "flex": "1"}),
            html.Span("Rule",         style={**H, "width": "90px"}),
            html.Span("Dimension",    style={**H, "width": "110px"}),
            html.Span("Failing Rows", style={**H, "width": "100px",
                                             "textAlign": "right"}),
            html.Span("Detected",     style={**H, "width": "96px"}),
            html.Span("Resolved",     style={**H, "width": "96px"}),
            html.Span("Days to Fix",  style={**H, "width": "80px",
                                             "textAlign": "center"}),
            html.Span("Timing",       style={**H, "width": "70px",
                                             "textAlign": "center"}),
        ], style={"display": "flex", "background": BG,
                  "borderTop": f"1px solid {DIVIDER}"})

        detail_rows = []
        for j, iss in enumerate(inst_issues):
            try:
                days_fix = (_date.fromisoformat(iss["resolved_at"]) -
                            _date.fromisoformat(iss["detected_at"])).days
                fix_str  = f"{days_fix}d"
                fix_clr  = C_GREEN if days_fix <= 7 else (C_RED if days_fix >= 20 else TEXT)
            except Exception:
                fix_str, fix_clr = "—", MUTED
            try:
                on_time = (_date.fromisoformat(iss["resolved_at"]) <=
                           _date.fromisoformat(iss["sla_deadline"]))
                timing_label = "On Time"
                timing_color = C_GREEN
            except Exception:
                on_time = None
                timing_label, timing_color = "—", MUTED
            if on_time is False:
                timing_label = "Late"
                timing_color = C_RED

            detail_rows.append(html.Div([
                html.Span(iss["table_name"],
                          style={"flex": "1", "fontSize": "11px", "color": TEXT,
                                 "padding": "6px 10px", "fontFamily": "monospace"}),
                html.Span(iss["rule_id"],
                          style={"width": "90px", "fontSize": "11px", "fontWeight": "700",
                                 "color": TEXT, "padding": "6px 10px"}),
                html.Span(iss["dimension"].title(),
                          style={"width": "110px", "fontSize": "11px",
                                 "color": MUTED, "padding": "6px 10px"}),
                html.Span(f"{iss.get('failing_rows', 0):,}",
                          style={"width": "100px", "fontSize": "11px", "color": MUTED,
                                 "textAlign": "right", "padding": "6px 10px"}),
                html.Span(iss.get("detected_at", "—"),
                          style={"width": "96px", "fontSize": "11px",
                                 "color": MUTED, "padding": "6px 10px"}),
                html.Span(iss.get("resolved_at", "—"),
                          style={"width": "96px", "fontSize": "11px",
                                 "color": C_GREEN, "padding": "6px 10px"}),
                html.Span(fix_str,
                          style={"width": "80px", "fontSize": "12px", "fontWeight": "700",
                                 "color": fix_clr, "textAlign": "center",
                                 "padding": "6px 10px"}),
                html.Span(timing_label,
                          style={"width": "70px", "fontSize": "11px", "fontWeight": "700",
                                 "color": timing_color, "textAlign": "center",
                                 "padding": "6px 10px"}),
            ], style={
                "display": "flex", "alignItems": "center",
                "background": BG if j % 2 == 0 else CARD,
                "borderTop": f"1px solid {DIVIDER}",
            }))

        collapse = html.Div(
            [tbl_hdr, *detail_rows],
            id={"type": "res-inst-collapse", "index": cidx},
            style={"display": "none"},
        )

        cards.append(html.Div(
            [header, collapse],
            style={
                "background": CARD, "border": f"1px solid {DIVIDER}",
                "borderRadius": "8px", "marginBottom": "8px",
                "overflow": "hidden",
            },
        ))

    return html.Div(cards)


def _build_issue_rows(issues: list, status: str) -> html.Div:
    """Render issue table rows for the given status category."""
    from datetime import date as _date
    today = _date.today()

    if not issues:
        label = {"open": "open issues", "penalized": "delayed issues",
                 "resolved": "resolved issues"}.get(status, "issues")
        return html.Div(f"No {label}.", style={"color": MUTED, "padding": "20px"})

    H = {"fontSize": "11px", "fontWeight": "900", "color": MUTED,
         "textTransform": "uppercase", "letterSpacing": "0.05em", "padding": "8px 10px"}

    if status == "resolved":
        hdr = html.Div([
            html.Span("Institution",  style={**H, "flex": "1"}),
            html.Span("Table",        style={**H, "width": "160px"}),
            html.Span("Rule",         style={**H, "width": "90px"}),
            html.Span("Dimension",    style={**H, "width": "100px"}),
            html.Span("Detected",     style={**H, "width": "96px"}),
            html.Span("Resolved",     style={**H, "width": "96px"}),
            html.Span("Days to Fix",  style={**H, "width": "80px", "textAlign": "center"}),
        ], style={"display": "flex", "background": BG,
                  "borderRadius": "8px 8px 0 0", "borderBottom": f"2px solid {DIVIDER}"})
    elif status == "penalized":
        hdr = html.Div([
            html.Span("Institution",  style={**H, "flex": "1"}),
            html.Span("Table",        style={**H, "width": "160px"}),
            html.Span("Rule",         style={**H, "width": "90px"}),
            html.Span("Dimension",    style={**H, "width": "100px"}),
            html.Span("Failing Rows", style={**H, "width": "96px", "textAlign": "right"}),
            html.Span("Detected",     style={**H, "width": "96px"}),
            html.Span("SLA Deadline", style={**H, "width": "96px"}),
            html.Span("Days Over",    style={**H, "width": "76px", "textAlign": "center"}),
        ], style={"display": "flex", "background": BG,
                  "borderRadius": "8px 8px 0 0", "borderBottom": f"2px solid {DIVIDER}"})
    else:
        hdr = html.Div([
            html.Span("Institution",  style={**H, "flex": "1"}),
            html.Span("Table",        style={**H, "width": "160px"}),
            html.Span("Rule",         style={**H, "width": "90px"}),
            html.Span("Dimension",    style={**H, "width": "100px"}),
            html.Span("Failing Rows", style={**H, "width": "90px", "textAlign": "right"}),
            html.Span("Detected",     style={**H, "width": "90px"}),
            html.Span("Deadline",     style={**H, "width": "90px"}),
            html.Span("Remaining",    style={**H, "width": "76px", "textAlign": "center"}),
            html.Span("Notify",       style={**H, "width": "52px", "textAlign": "center"}),
        ], style={"display": "flex", "background": BG,
                  "borderRadius": "8px 8px 0 0", "borderBottom": f"2px solid {DIVIDER}"})

    rows = []
    for i, iss in enumerate(issues):
        band  = iss.get("urgency_band", "new")
        clr   = _URGENCY_COLORS.get(band, MUTED)
        lb    = iss["le_book"]
        name  = (iss.get("institution_name") or lb).title()
        bg    = "#C9956C" if i % 2 == 0 else BG

        inst_cell = html.Div([
            html.Span("●", style={"color": clr, "fontSize": "9px", "marginRight": "5px"}),
            html.Span(name, style={"fontSize": "12px", "color": TEXT}),
        ], style={"flex": "1", "display": "flex", "alignItems": "center",
                  "padding": "7px 10px", "borderLeft": f"3px solid {clr}"})

        if status == "resolved":
            detected  = iss.get("detected_at", "—")
            resolved  = iss.get("resolved_at",  "—")
            try:
                days_fix = (_date.fromisoformat(resolved) - _date.fromisoformat(detected)).days
                fix_str  = f"{days_fix}d"
                fix_clr  = C_GREEN if days_fix <= 7 else (C_RED if days_fix >= 20 else TEXT)
            except Exception:
                fix_str, fix_clr = "—", MUTED
            row_children = [
                inst_cell,
                html.Span(iss["table_name"],       style={"width": "160px", "fontSize": "11px", "color": MUTED, "padding": "7px 10px"}),
                html.Span(iss["rule_id"],           style={"width": "90px",  "fontSize": "11px", "fontWeight": "700", "color": TEXT, "padding": "7px 10px"}),
                html.Span(iss["dimension"].title(), style={"width": "100px", "fontSize": "11px", "color": MUTED, "padding": "7px 10px"}),
                html.Span(detected,                 style={"width": "96px",  "fontSize": "11px", "color": MUTED, "padding": "7px 10px"}),
                html.Span(resolved,                 style={"width": "96px",  "fontSize": "11px", "color": C_GREEN, "padding": "7px 10px"}),
                html.Span(fix_str,                  style={"width": "80px",  "fontSize": "12px", "fontWeight": "700", "color": fix_clr, "textAlign": "center", "padding": "7px 10px"}),
            ]
        elif status == "penalized":
            try:
                days_over = (today - _date.fromisoformat(iss["sla_deadline"])).days
                over_str  = f"+{days_over}d"
            except Exception:
                over_str = "—"
            row_children = [
                inst_cell,
                html.Span(iss["table_name"],            style={"width": "160px", "fontSize": "11px", "color": MUTED, "padding": "7px 10px"}),
                html.Span(iss["rule_id"],               style={"width": "90px",  "fontSize": "11px", "fontWeight": "700", "color": TEXT, "padding": "7px 10px"}),
                html.Span(iss["dimension"].title(),     style={"width": "100px", "fontSize": "11px", "color": MUTED, "padding": "7px 10px"}),
                html.Span(f"{iss['failing_rows']:,}",   style={"width": "96px",  "fontSize": "12px", "fontWeight": "700", "color": C_RED, "textAlign": "right", "padding": "7px 10px"}),
                html.Span(iss.get("detected_at", "—"), style={"width": "96px",  "fontSize": "11px", "color": MUTED, "padding": "7px 10px"}),
                html.Span(iss.get("sla_deadline", "—"),style={"width": "96px",  "fontSize": "11px", "color": C_RED, "padding": "7px 10px"}),
                html.Span(over_str,                    style={"width": "76px",  "fontSize": "12px", "fontWeight": "700", "color": C_RED, "textAlign": "center", "padding": "7px 10px"}),
            ]
        else:
            try:
                days_left = (_date.fromisoformat(iss["sla_deadline"]) - today).days
            except Exception:
                days_left = None
            is_overdue  = band == "overdue" or (isinstance(days_left, int) and days_left < 0)
            if is_overdue:
                over_days  = abs(days_left) if isinstance(days_left, int) else "?"
                days_str   = f"⚠ {over_days}d over"
                days_color = _URGENCY_COLORS["overdue"]
            elif isinstance(days_left, int) and days_left <= 5:
                days_str   = f"{days_left}d left"
                days_color = C_RED
            elif isinstance(days_left, int):
                days_str   = f"{days_left}d left"
                days_color = TEXT
            else:
                days_str   = "?"
                days_color = MUTED
            row_children = [
                inst_cell,
                html.Span(iss["table_name"],           style={"width": "160px", "fontSize": "11px", "color": MUTED, "padding": "7px 10px"}),
                html.Span(iss["rule_id"],              style={"width": "90px",  "fontSize": "11px", "fontWeight": "700", "color": TEXT, "padding": "7px 10px"}),
                html.Span(iss["dimension"].title(),    style={"width": "100px", "fontSize": "11px", "color": MUTED, "padding": "7px 10px"}),
                html.Span(f"{iss['failing_rows']:,}",  style={"width": "90px",  "fontSize": "12px", "fontWeight": "700", "color": TEXT, "textAlign": "right", "padding": "7px 10px"}),
                html.Span(iss["detected_at"],          style={"width": "90px",  "fontSize": "11px", "color": MUTED, "padding": "7px 10px"}),
                html.Span(iss["sla_deadline"],         style={"width": "90px",  "fontSize": "11px", "color": C_RED if is_overdue else MUTED, "padding": "7px 10px"}),
                html.Span(days_str,                    style={"width": "76px",  "fontSize": "12px", "fontWeight": "700", "color": days_color, "textAlign": "center", "padding": "7px 10px"}),
                html.Div("🔔",
                    id={"type": "notify-btn", "index": lb},
                    n_clicks=0,
                    title=f"Send reminder to {name}",
                    style={"width": "52px", "textAlign": "center", "fontSize": "14px",
                           "cursor": "pointer", "color": clr, "padding": "7px 0",
                           "userSelect": "none"},
                ),
            ]

        rows.append(html.Div(row_children, style={
            "display": "flex", "alignItems": "center",
            "background": bg, "borderBottom": f"1px solid {DIVIDER}",
        }))

    return html.Div([hdr, *rows], style={
        "background": CARD, "borderRadius": "8px",
        "border": f"1px solid {DIVIDER}", "marginBottom": "8px",
    })


def _alerts_page(cat: str = "") -> html.Div:
    cat_label = CAT_LABELS.get(cat, "All Categories") if cat else "All Categories"

    header_row = html.Div([
        html.Div([
            html.H2("Resolved Issues", style={
                "fontSize": "18px", "fontWeight": "900", "color": TEXT,
                "margin": "0 0 2px",
            }),
            html.Span(f"Showing: {cat_label}", style={
                "fontSize": "11px", "color": MUTED,
            }),
        ]),
        html.Div("⬇ Export XLSX", id="issues-download-btn", n_clicks=0,
                 style={"cursor": "pointer", "background": BRAND, "color": CARD,
                        "fontSize": "11px", "fontWeight": "700",
                        "padding": "6px 14px", "borderRadius": "5px",
                        "userSelect": "none", "whiteSpace": "nowrap",
                        "alignSelf": "center"}),
    ], style={"display": "flex", "justifyContent": "space-between",
              "alignItems": "flex-start", "marginBottom": "20px"})

    return html.Div([
        header_row,
        html.Div(id="alerts-summary-bar", style={"marginBottom": "24px"}),
        # Hidden RadioItems — preserves callback wiring; value drives filtered data
        dcc.RadioItems(
            id="alerts-cat-filter",
            options=[
                {"label": "All", "value": ""},
                {"label": "B",   "value": "B"},
                {"label": "MF",  "value": "MF"},
                {"label": "S",   "value": "SACCO"},
            ],
            value=cat or "",
            style={"display": "none"},
        ),
        html.Div(id="issue-list"),
        html.Div(id="notify-feedback", style={"display": "none"}),
    ], style={"padding": "28px 32px", "maxWidth": "1380px", "margin": "0 auto"})


# ── bootstrap values (rendered once at startup) ────────────────────────────────

_today_e   = _today_entry()
_counts    = _category_counts(_today_e)
_run_ts    = _PIPELINE.get("data_processed", "")
_run_date  = _PIPELINE.get("run_date", _today_e.get("date", "—"))
_run_label = (
    f"Last detection: {_run_date}"
    + (f"  ·  {_run_ts[11:16]} UTC" if len(_run_ts) >= 16 else "")
)


# ── Validations page helpers ───────────────────────────────────────────────────

_DIM_PILL_COLOR = {
    "completeness": "#753918",
    "accuracy":     "#B8860B",
    "timeliness":   "#7C3D1E",
    "validity":     "#C9956C",
    "uniqueness":   "#2563EB",
    "relationship": "#475569",
}


def _dim_pill(dim: str) -> html.Span:
    color = _DIM_PILL_COLOR.get(dim, MUTED)
    return html.Span(dim.capitalize(), style={
        "background":   CARD,
        "color":        color,
        "border":       f"1px solid {color}",
        "borderRadius": "4px",
        "padding":      "2px 7px",
        "fontSize":     "11px",
        "fontWeight":   "700",
        "whiteSpace":   "nowrap",
    })


_KNOWN_TABLES = [
    "accounts", "customers_expanded", "contracts_disburse",
    "contract_loans", "contract_schedules", "contracts_expanded",
    "loan_applications_2", "prev_loan_applications",
]

_CHECK_TYPES = [
    ("not_null",        "Field must not be null"),
    ("positive",        "Numeric field must be > 0"),
    ("non_negative",    "Numeric field must be ≥ 0"),
    ("date_not_future", "Date field must not be in the future"),
    ("domain",          "Field value must be in an allowed set"),
    ("range",           "Numeric field must be between min and max"),
    ("pattern",         "Field must match a regex pattern"),
]

_STATUS_STYLE = {
    "draft":   {"color": "#4A3728", "background": "rgba(74,55,40,.12)",
                "border": "1px solid rgba(74,55,40,.35)"},
    "pending": {"color": "#92400E", "background": "rgba(245,158,11,.12)",
                "border": "1px solid rgba(245,158,11,.35)"},
    "active":  {"color": "#065F46", "background": "rgba(16,185,129,.12)",
                "border": "1px solid rgba(16,185,129,.35)"},
    "error":   {"color": "#991B1B", "background": "rgba(239,68,68,.12)",
                "border": "1px solid rgba(239,68,68,.35)"},
}


def _rules_charts(builtin_rules: list[dict], user_rules: list[dict]) -> html.Div:
    """Single horizontal bar chart: total rules per table (no dimension breakdown)."""
    from collections import defaultdict

    pending = [r for r in user_rules if r.get("status") == "pending"]
    active  = [r for r in user_rules if r.get("status") == "active"]

    _TABLE_LABELS = {
        "accounts":              "Accounts",
        "contract_loans":        "Contract Loans",
        "contract_schedules":    "Contract Schedules",
        "contracts_disburse":    "Contracts Disburse",
        "contracts_expanded":    "Contracts",
        "customers_expanded":    "Customers",
        "loan_applications_2":   "Loan Applications",
        "prev_loan_applications": "Prev Loan Apps",
    }

    def _count_per_table(rule_list: list[dict]) -> dict[str, int]:
        counts: dict[str, int] = defaultdict(int)
        for r in rule_list:
            tables_str = r.get("tables", "")
            if "→" in tables_str:
                child = tables_str.split("→")[0].strip()
                if child:
                    counts[child] += 1
            else:
                for t in tables_str.split(","):
                    t = t.strip()
                    if t:
                        counts[t] += 1
        return counts

    run_counts     = _count_per_table(builtin_rules + active)
    pending_counts = _count_per_table(pending)

    all_tables = sorted(
        set(run_counts.keys()) | set(pending_counts.keys()),
        key=lambda t: run_counts.get(t, 0) + pending_counts.get(t, 0),
    )

    y_labels  = [_TABLE_LABELS.get(t, t.replace("_", " ").title()) for t in all_tables]
    x_run     = [run_counts.get(t, 0) for t in all_tables]
    x_pending = [pending_counts.get(t, 0) for t in all_tables]
    has_pending = any(x_pending)

    traces = [go.Bar(
        name="Built-in + active rules",
        y=y_labels,
        x=x_run,
        orientation="h",
        marker_color=BRAND,
        text=[str(v) for v in x_run],
        textposition="outside",
        hovertemplate="%{y}: %{x} rules<extra></extra>",
        showlegend=False,
    )]
    if has_pending:
        traces.append(go.Bar(
            name="Pending (not yet run)",
            y=y_labels,
            x=x_pending,
            orientation="h",
            marker=dict(
                color="rgba(148,163,184,0.35)",
                pattern=dict(shape="/", fgcolor="rgba(100,116,139,0.6)", size=6),
                line=dict(color="rgba(100,116,139,0.5)", width=1),
            ),
            text=[str(v) if v else "" for v in x_pending],
            textposition="outside",
            hovertemplate="Pending — %{y}: %{x} rules<extra></extra>",
        ))

    fig = go.Figure(traces)
    fig.update_layout(
        barmode="stack",
        height=max(260, 40 * len(all_tables) + 70),
        paper_bgcolor=CARD, plot_bgcolor=CARD,
        margin=dict(l=8, r=40, t=20, b=8),
        font=dict(family=FONT, size=11, color=TEXT),
        xaxis=dict(title="Number of rules", gridcolor=DIVIDER, zeroline=False,
                   tickfont=dict(size=10)),
        yaxis=dict(tickfont=dict(size=11), showgrid=False, automargin=True),
        bargap=0.3,
        legend=dict(orientation="h", yanchor="bottom", y=1.02,
                    xanchor="left", x=0, font=dict(size=10)),
        showlegend=has_pending,
    )

    return html.Div([
        html.Div("RULES BY TABLE", style={
            "fontSize": "11px", "fontWeight": "900", "color": MUTED,
            "textTransform": "uppercase", "letterSpacing": "0.06em",
            "marginBottom": "6px",
        }),
        dcc.Graph(figure=fig, config={"displayModeBar": False}),
    ], style={
        "background": CARD, "borderRadius": "8px",
        "padding": "16px 16px 8px",
        "boxShadow": "0 1px 4px rgba(117,57,24,0.08)",
        "border": f"1px solid {DIVIDER}",
        "marginBottom": "20px",
    })


def _rule_form(next_id: str) -> html.Div:
    """The Add Rule form. Always rendered in the DOM; toggled via display style."""
    inp = {
        "width": "100%", "padding": "7px 10px",
        "border": f"1px solid {DIVIDER}", "borderRadius": "5px",
        "fontSize": "12px", "color": TEXT, "fontFamily": FONT,
        "boxSizing": "border-box", "outline": "none",
    }
    lbl = {
        "fontSize": "11px", "fontWeight": "900", "color": MUTED,
        "textTransform": "uppercase", "letterSpacing": "0.05em",
        "marginBottom": "4px", "display": "block",
    }

    def _field(label: str, child) -> html.Div:
        return html.Div([html.Span(label, style=lbl), child],
                        style={"display": "flex", "flexDirection": "column"})

    def _dd(id_, opts, placeholder="Select…") -> dcc.Dropdown:
        return dcc.Dropdown(
            id=id_, options=opts, placeholder=placeholder,
            clearable=False,
            style={"fontSize": "12px", "fontFamily": FONT},
        )

    dim_opts        = [{"label": d.capitalize(), "value": d}
                       for d in ["completeness", "accuracy", "timeliness", "validity"]]
    table_opts      = [{"label": t, "value": t} for t in _KNOWN_TABLES]
    check_type_opts = [{"label": label, "value": val} for val, label in _CHECK_TYPES]

    return html.Div([
        html.Div([
            _field("Rule ID",
                dcc.Input(id="new-rule-id", value=next_id, debounce=False, style=inp)),
            _field("Dimension", _dd("new-rule-dim", dim_opts)),
        ], style={"display": "grid", "gridTemplateColumns": "1fr 1fr", "gap": "14px",
                  "marginBottom": "12px"}),

        html.Div([
            _field("Category",
                dcc.Input(id="new-rule-cat", placeholder="e.g. Format Validity",
                          debounce=False, style=inp)),
            _field("Rule Name / Description",
                dcc.Input(id="new-rule-name",
                          placeholder="e.g. Email address must be valid",
                          debounce=False, style=inp)),
        ], style={"display": "grid", "gridTemplateColumns": "1fr 2fr", "gap": "14px",
                  "marginBottom": "12px"}),

        html.Div([
            _field("Table",      _dd("new-rule-table", table_opts, "Select table…")),
            _field("Field (column)",
                dcc.Input(id="new-rule-field", placeholder="e.g. email_id",
                          debounce=False, style=inp)),
            _field("Check Type", _dd("new-rule-check-type", check_type_opts, "Select check…")),
        ], style={"display": "grid", "gridTemplateColumns": "1fr 1fr 1fr", "gap": "14px",
                  "marginBottom": "12px"}),

        html.Div([
            _field("Allowed Values (comma-separated)",
                dcc.Input(id="new-rule-domain-vals",
                          placeholder="e.g.  M, F, C",
                          debounce=False, style=inp)),
        ], id="param-domain", style={"marginBottom": "12px", "display": "none"}),

        html.Div([
            _field("Minimum",
                dcc.Input(id="new-rule-range-min", type="number",
                          placeholder="0", debounce=False, style=inp)),
            _field("Maximum",
                dcc.Input(id="new-rule-range-max", type="number",
                          placeholder="100", debounce=False, style=inp)),
        ], id="param-range",
           style={"display": "none", "marginBottom": "12px",
                  "gridTemplateColumns": "1fr 1fr", "gap": "14px"}),

        html.Div([
            _field("Regex Pattern",
                dcc.Input(id="new-rule-pattern",
                          placeholder="e.g.  ^[A-Z]{3}$",
                          debounce=False, style=inp)),
        ], id="param-pattern", style={"marginBottom": "12px", "display": "none"}),

        html.Div([
            html.Div(
                "Submit Rule",
                id="new-rule-submit", n_clicks=0,
                style={
                    "cursor": "pointer", "background": BRAND,
                    "color": CARD, "fontSize": "12px", "fontWeight": "700",
                    "padding": "9px 22px", "borderRadius": "6px",
                    "userSelect": "none", "display": "inline-block",
                },
            ),
            html.Div(id="new-rule-feedback", style={
                "fontSize": "12px", "lineHeight": "1.4",
                "marginLeft": "14px", "flex": "1",
            }),
        ], style={"display": "flex", "alignItems": "center"}),

    ], id="rule-form-panel", style={
        "background":    CARD,
        "border":        f"1px solid {DIVIDER}",
        "borderRadius":  "8px",
        "padding":       "20px",
        "marginBottom":  "20px",
        "display":       "none",
    })


def _rules_table_row(r: dict, i: int, is_user: bool = False) -> html.Div:
    bg     = "#C9956C" if i % 2 == 0 else BG
    status = r.get("status") if is_user else None
    cells  = [
        html.Span(r["rule_id"], style={
            "width": "80px", "flexShrink": "0",
            "fontSize": "12px", "fontWeight": "900",
            "color": BRAND, "fontFamily": "monospace", "lineHeight": "1.4",
        }),
        html.Div(_dim_pill(r["dimension"]), style={"width": "110px", "flexShrink": "0"}),
        html.Span(r.get("category") or "—", style={
            "width": "160px", "flexShrink": "0",
            "fontSize": "11px", "color": MUTED, "lineHeight": "1.4",
        }),
        html.Span(r["rule_name"], style={
            "flex": "1", "flexShrink": "1",
            "fontSize": "12px", "color": TEXT, "lineHeight": "1.4",
        }),
        html.Span(r["tables"], style={
            "width": "200px", "flexShrink": "0",
            "fontSize": "11px", "color": MUTED,
            "overflow": "hidden", "textOverflow": "ellipsis",
            "whiteSpace": "nowrap", "lineHeight": "1.4",
        }),
        html.Span(r.get("fields") or "—", style={
            "width": "180px", "flexShrink": "0",
            "fontSize": "11px", "color": MUTED,
            "overflow": "hidden", "textOverflow": "ellipsis",
            "whiteSpace": "nowrap", "lineHeight": "1.4",
        }),
    ]
    if is_user:
        sty = _STATUS_STYLE.get(status, _STATUS_STYLE["pending"])
        label = (status or "pending").upper()
        cells.append(html.Span(label, style={
            **sty,
            "width": "72px", "flexShrink": "0",
            "fontSize": "10px", "fontWeight": "900",
            "borderRadius": "4px", "padding": "2px 6px",
            "textAlign": "center", "lineHeight": "1.5",
        }))
    else:
        cells.append(html.Span("", style={"width": "72px", "flexShrink": "0"}))

    return html.Div(cells, style={
        "display": "flex", "alignItems": "center", "gap": "12px",
        "padding": "8px 16px", "background": bg,
        "borderBottom": f"1px solid {DIVIDER}",
    })


def _draft_review_section(draft_rules: list[dict]) -> html.Div | None:
    """Pending-review panel shown only when there are draft rules awaiting approval."""
    if not draft_rules:
        return None

    H = {"fontSize": "11px", "fontWeight": "900", "color": MUTED,
         "textTransform": "uppercase", "letterSpacing": "0.05em", "flexShrink": "0"}

    header = html.Div([
        html.Span("Rule ID",   style={**H, "width": "80px"}),
        html.Span("Dimension", style={**H, "width": "100px"}),
        html.Span("Rule",      style={**H, "flex": "1"}),
        html.Span("Table(s)",  style={**H, "width": "180px"}),
        html.Span("Type",      style={**H, "width": "110px"}),
        html.Span("Actions",   style={**H, "width": "170px", "textAlign": "center"}),
    ], style={
        "display": "flex", "alignItems": "center", "gap": "10px",
        "padding": "9px 14px",
        "borderBottom": f"2px solid {DIVIDER}",
        "background": "rgba(74,55,40,0.08)", "borderRadius": "8px 8px 0 0",
    })

    rows = []
    for i, r in enumerate(draft_rules):
        rid    = r["rule_id"]
        bg     = "#C9956C" if i % 2 == 0 else BG
        rows.append(html.Div([
            html.Span(rid, style={
                "width": "80px", "flexShrink": "0", "fontSize": "12px",
                "fontWeight": "900", "color": "#4A3728", "fontFamily": "monospace",
            }),
            html.Div(_dim_pill(r["dimension"]), style={"width": "100px", "flexShrink": "0"}),
            html.Span(r["rule_name"], style={
                "flex": "1", "fontSize": "12px", "color": TEXT,
                "overflow": "hidden", "textOverflow": "ellipsis", "whiteSpace": "nowrap",
            }),
            html.Span(r.get("tables", ""), style={
                "width": "180px", "flexShrink": "0", "fontSize": "11px",
                "color": MUTED, "overflow": "hidden", "textOverflow": "ellipsis",
                "whiteSpace": "nowrap",
            }),
            html.Span(r.get("check_type", ""), style={
                "width": "110px", "flexShrink": "0", "fontSize": "11px", "color": MUTED,
            }),
            html.Div([
                html.Div("Approve", id={"type": "approve-btn", "index": rid}, n_clicks=0,
                    style={
                        "cursor": "pointer", "background": C_GREEN, "color": CARD,
                        "fontSize": "11px", "fontWeight": "700", "padding": "5px 12px",
                        "borderRadius": "5px", "userSelect": "none", "marginRight": "6px",
                    }),
                html.Div("Delete", id={"type": "delete-draft-btn", "index": rid}, n_clicks=0,
                    style={
                        "cursor": "pointer", "background": C_RED, "color": CARD,
                        "fontSize": "11px", "fontWeight": "700", "padding": "5px 12px",
                        "borderRadius": "5px", "userSelect": "none",
                    }),
            ], style={"width": "170px", "flexShrink": "0", "display": "flex", "alignItems": "center"}),
        ], style={
            "display": "flex", "alignItems": "center", "gap": "10px",
            "padding": "8px 14px", "background": bg,
            "borderBottom": f"1px solid {DIVIDER}",
        }))

    return html.Div([
        html.Div([
            html.Div("PENDING ADMIN REVIEW", style={
                "fontSize": "12px", "fontWeight": "900", "color": "#4A3728",
                "letterSpacing": "0.04em",
            }),
            html.Div(
                f"{len(draft_rules)} rule{'s' if len(draft_rules) != 1 else ''} submitted — "
                "approve to queue for next pipeline run, or delete to reject.",
                style={"fontSize": "11px", "color": MUTED, "marginTop": "3px"},
            ),
        ], style={"marginBottom": "12px"}),
        html.Div(
            [header] + rows,
            style={
                "border": "1px solid rgba(74,55,40,0.30)",
                "borderRadius": "8px", "overflow": "hidden",
            },
        ),
    ], style={
        "background": "rgba(74,55,40,0.04)",
        "border": "1px solid rgba(74,55,40,0.18)",
        "borderRadius": "10px", "padding": "16px 16px 8px",
        "marginBottom": "20px",
    })


def _complex_rule_form(next_id: str) -> html.Div:
    """Form for rules that cannot be expressed with a simple check type."""
    inp = {
        "width": "100%", "padding": "7px 10px",
        "border": f"1px solid {DIVIDER}", "borderRadius": "5px",
        "fontSize": "12px", "color": TEXT, "fontFamily": FONT,
        "boxSizing": "border-box", "outline": "none",
    }
    ta = {**inp, "resize": "vertical", "minHeight": "72px", "fontFamily": "monospace"}
    lbl = {
        "fontSize": "11px", "fontWeight": "900", "color": MUTED,
        "textTransform": "uppercase", "letterSpacing": "0.05em",
        "marginBottom": "4px", "display": "block",
    }

    def _field(label, child):
        return html.Div([html.Span(label, style=lbl), child],
                        style={"display": "flex", "flexDirection": "column"})

    dim_opts = [{"label": d.capitalize(), "value": d}
                for d in ["completeness", "accuracy", "timeliness", "validity"]]

    return html.Div([
        html.Div(
            "Use this form for business rules that cannot be expressed with simple check "
            "types. If you provide a SQL condition it will be auto-evaluated by the "
            "pipeline (as a pandas query expression identifying failing rows). "
            "Otherwise the rule is tracked as manual.",
            style={"fontSize": "12px", "color": MUTED, "marginBottom": "16px",
                   "lineHeight": "1.5"},
        ),

        html.Div([
            _field("Rule ID",
                dcc.Input(id="cx-rule-id", value=f"CX-{next_id[4:]}",
                          debounce=False, style=inp)),
            _field("Dimension",
                dcc.Dropdown(id="cx-rule-dim", options=dim_opts, clearable=False,
                             style={"fontSize": "12px", "fontFamily": FONT})),
        ], style={"display": "grid", "gridTemplateColumns": "1fr 1fr",
                  "gap": "14px", "marginBottom": "12px"}),

        _field("Rule Name / Short Description",
            dcc.Input(id="cx-rule-name",
                      placeholder="e.g. Loan disbursement cannot exceed approved limit",
                      debounce=False, style={**inp, "marginBottom": "12px"})),

        html.Div([
            _field("Table(s) (comma-separated)",
                dcc.Input(id="cx-rule-tables",
                          placeholder="e.g. contracts_disburse, contract_loans",
                          debounce=False, style=inp)),
            _field("Field(s) (optional)",
                dcc.Input(id="cx-rule-fields",
                          placeholder="e.g. disbursed_amount, approved_amount",
                          debounce=False, style=inp)),
        ], style={"display": "grid", "gridTemplateColumns": "1fr 1fr",
                  "gap": "14px", "marginBottom": "12px"}),

        _field("Business Logic — describe what this rule checks and why",
            dcc.Textarea(id="cx-rule-logic",
                         placeholder="e.g. The disbursed amount on a contract must never "
                                     "exceed the originally approved amount. Breaches indicate "
                                     "control failures in the disbursement workflow.",
                         style={**ta, "marginBottom": "12px"})),

        html.Div([
            html.Span("SQL / Pandas Condition", style=lbl),
            html.Span(
                " (optional — a pandas df.query() expression that selects FAILING rows)",
                style={"fontSize": "10px", "color": MUTED, "marginLeft": "4px"},
            ),
        ], style={"display": "flex", "alignItems": "baseline", "marginBottom": "4px"}),
        dcc.Textarea(
            id="cx-rule-condition",
            placeholder="e.g.  disbursed_amount > approved_amount",
            style={**ta, "marginBottom": "12px"},
        ),

        html.Div([
            html.Div("Submit for Review", id="cx-rule-submit", n_clicks=0, style={
                "cursor": "pointer", "background": BRAND, "color": CARD,
                "fontSize": "12px", "fontWeight": "700", "padding": "9px 22px",
                "borderRadius": "6px", "userSelect": "none", "display": "inline-block",
            }),
            html.Div(id="cx-rule-feedback",
                     style={"fontSize": "12px", "lineHeight": "1.4",
                            "marginLeft": "14px", "flex": "1"}),
        ], style={"display": "flex", "alignItems": "center"}),

    ], id="complex-form-panel", style={
        "background": CARD, "border": f"1px solid {DIVIDER}", "borderRadius": "8px",
        "padding": "20px", "marginBottom": "20px", "display": "none",
    })


def _validations_page() -> html.Div:
    builtin_rules = get_all_rules()
    n_tables  = len({t.strip() for r in builtin_rules
                     for t in (r.get("tables") or "").replace("→", ",").split(",")
                     if t.strip()})
    subtitle  = f"{len(builtin_rules)} rules across {n_tables} tables"

    return html.Div([
        # ── header: title + download ──────────────────────────────────────────
        html.Div([
            html.Div([
                html.Div("VALIDATION RULES", style={
                    "fontSize": "13px", "fontWeight": "900",
                    "color": TEXT, "letterSpacing": "0.04em", "lineHeight": "1.15",
                }),
                html.Div(subtitle,
                         style={"fontSize": "11px", "color": MUTED, "marginTop": "3px"}),
            ]),
            html.Div("Download CSV", id="rules-download-btn", n_clicks=0, style={
                "cursor": "pointer", "background": BRAND, "color": CARD,
                "fontSize": "12px", "fontWeight": "700", "padding": "8px 18px",
                "borderRadius": "6px", "userSelect": "none",
            }),
        ], style={
            "display": "flex", "alignItems": "center",
            "justifyContent": "space-between", "marginBottom": "16px",
        }),

        # Hidden trigger buttons + form panels kept in DOM so callback IDs resolve
        html.Div(id="form-toggle-btn",         n_clicks=0, style={"display": "none"}),
        html.Div(id="complex-form-toggle-btn", n_clicks=0, style={"display": "none"}),
        _rule_form("USR-001"),
        _complex_rule_form("USR-001"),

        # ── chart + download ──────────────────────────────────────────────────
        _rules_charts(builtin_rules, []),
        dcc.Download(id="rules-download"),
    ])


# ── remediation page ──────────────────────────────────────────────────────────

def _build_cr_list(crs: list[dict], role: str = "bnr_admin") -> html.Div:
    """Render the change-request table.  Called at page load and after each CR action."""
    import dq_change_request as cr_mod
    import dq_auth as _auth

    is_bnr_admin = _auth.is_admin(role)
    is_inst      = role == "inst_user"

    if not crs:
        return html.Div(
            "No change requests found for this filter.",
            style={"color": MUTED, "padding": "24px", "textAlign": "center",
                   "fontSize": "12px"},
        )

    H = {"fontSize": "11px", "fontWeight": "900", "color": MUTED,
         "textTransform": "uppercase", "letterSpacing": "0.05em",
         "padding": "8px 10px", "flexShrink": "0"}

    hdr = html.Div([
        html.Span("CR ID",        style={**H, "width": "148px"}),
        html.Span("Institution",  style={**H, "flex": "1"}),
        html.Span("Title",        style={**H, "flex": "2"}),
        html.Span("Status",       style={**H, "width": "108px"}),
        html.Span("Issues",       style={**H, "width": "56px",  "textAlign": "center"}),
        html.Span("Failing",      style={**H, "width": "74px",  "textAlign": "right"}),
        html.Span("Target",       style={**H, "width": "86px"}),
        html.Span("Assigned To",  style={**H, "width": "154px"}),
        html.Span("Actions",      style={**H, "width": "228px"}),
    ], style={
        "display": "flex", "alignItems": "center",
        "background": BG, "borderRadius": "8px 8px 0 0",
        "borderBottom": f"2px solid {DIVIDER}",
    })

    def _action_btn(label: str, cr_id: str, action: str, bg_color: str) -> html.Div:
        return html.Div(
            label,
            id={"type": "cr-action-btn", "index": f"{cr_id}|{action}"},
            n_clicks=0,
            title=f"{label} — {cr_id}",
            style={
                "display":      "inline-block",
                "background":   bg_color,
                "color":        CARD,
                "padding":      "4px 9px",
                "borderRadius": "4px",
                "fontSize":     "11px",
                "fontWeight":   "700",
                "cursor":       "pointer",
                "userSelect":   "none",
                "marginRight":  "4px",
                "marginBottom": "2px",
                "whiteSpace":   "nowrap",
            },
        )

    rows = []
    for i, cr in enumerate(crs):
        bg     = "#C9956C" if i % 2 == 0 else BG
        status = cr["status"]
        clr    = cr_mod.STATUS_COLORS.get(status, MUTED)
        label  = cr_mod.STATUS_LABELS.get(status, status.title())
        r, g, b_val = int(clr[1:3], 16), int(clr[3:5], 16), int(clr[5:7], 16)

        try:
            n_issues = len(json.loads(cr.get("issue_ids") or "[]"))
        except Exception:
            n_issues = 0

        status_chip = html.Span(label, style={
            "background":   f"rgba({r},{g},{b_val},0.12)",
            "color":        clr,
            "border":       f"1px solid {clr}",
            "borderRadius": "4px",
            "padding":      "2px 7px",
            "fontSize":     "10px",
            "fontWeight":   "700",
            "whiteSpace":   "nowrap",
        })

        # Contextual action buttons — gated by role
        # Flow: BNR creates → institution starts + submits → BNR approves/rejects (table by table)
        try:
            cr_tables = json.loads(cr.get("tables") or "[]")
        except Exception:
            cr_tables = []
        try:
            table_approvals = json.loads(cr.get("table_approvals") or "{}")
        except Exception:
            table_approvals = {}

        action_btns: list = []
        if status == "open":
            if is_inst:
                action_btns.append(_action_btn("Start Work", cr["cr_id"], "in_progress", "#D97706"))
            if is_bnr_admin:
                action_btns.append(_action_btn("Cancel", cr["cr_id"], "closed", "#6B7280"))
        elif status == "in_progress":
            if is_inst:
                action_btns.append(_action_btn("Submit for Review", cr["cr_id"], "submitted", "#7C3D1E"))
            if is_bnr_admin:
                action_btns.append(_action_btn("Cancel", cr["cr_id"], "closed", "#6B7280"))
        elif status == "submitted":
            if is_bnr_admin:
                action_btns.append(_action_btn("Reject", cr["cr_id"], "rejected", "#DC2626"))
                if cr_tables:
                    from dq_issue_tracker import get_issues as _gi
                    try:
                        lb_open: dict[str, int] = {}
                        for _i in _gi(status="open"):
                            if _i["le_book"] == cr["le_book"]:
                                lb_open[_i["table_name"]] = lb_open.get(_i["table_name"], 0) + 1
                    except Exception:
                        lb_open = {}
                    tbl_rows = []
                    for tbl in cr_tables:
                        tbl_label  = TABLE_NAMES_PRETTY.get(tbl, tbl.replace("_", " ").title())
                        tbl_status = table_approvals.get(tbl, {}).get("status", "pending")
                        n_open     = lb_open.get(tbl, 0)
                        chip_color = "#16A34A" if tbl_status == "approved" else "#D97706"
                        chip_label = "Approved" if tbl_status == "approved" else "Pending"
                        approve_btn = html.Span() if tbl_status == "approved" else html.Div(
                            "✓ Approve Table",
                            id={"type": "cr-tbl-approve-btn", "index": f"{cr['cr_id']}|{tbl}"},
                            n_clicks=0,
                            style={"display": "inline-block", "background": "#16A34A",
                                   "color": CARD, "fontSize": "10px", "fontWeight": "700",
                                   "padding": "3px 10px", "borderRadius": "4px",
                                   "cursor": "pointer", "userSelect": "none", "marginLeft": "6px"},
                        )
                        tbl_rows.append(html.Div([
                            html.Span(tbl_label, style={"fontSize": "11px", "color": TEXT,
                                                        "fontWeight": "700", "width": "130px",
                                                        "flexShrink": "0"}),
                            html.Span(f"{n_open} open", style={"fontSize": "10px",
                                                               "color": C_RED if n_open else C_GREEN,
                                                               "width": "60px", "flexShrink": "0"}),
                            html.Span(chip_label, style={"fontSize": "10px", "fontWeight": "700",
                                                          "color": chip_color,
                                                          "border": f"1px solid {chip_color}",
                                                          "borderRadius": "4px", "padding": "1px 6px"}),
                            approve_btn,
                        ], style={"display": "flex", "alignItems": "center",
                                  "gap": "6px", "marginBottom": "4px"}))
                    action_btns.append(html.Div(tbl_rows, style={
                        "background": BG, "border": f"1px solid {DIVIDER}",
                        "borderRadius": "6px", "padding": "8px 12px",
                        "marginTop": "6px", "width": "100%",
                    }))
                else:
                    action_btns.append(_action_btn("Approve", cr["cr_id"], "approved", "#16A34A"))
        elif status == "rejected":
            if is_inst:
                action_btns.append(_action_btn("Reopen", cr["cr_id"], "in_progress", "#D97706"))
            if is_bnr_admin:
                action_btns.append(_action_btn("Close", cr["cr_id"], "closed", "#6B7280"))
        elif status == "approved":
            if is_bnr_admin:
                action_btns = [_action_btn("Close", cr["cr_id"], "closed", "#6B7280")]

        # Reviewer note shown under approved/rejected
        reviewer_note = html.Span()
        if status in ("approved", "rejected") and cr.get("reviewed_by"):
            reviewer_note = html.Div(
                f"by {cr['reviewed_by']}" +
                (f" — \"{cr['review_notes']}\"" if cr.get("review_notes") else ""),
                style={"fontSize": "10px", "color": MUTED, "marginTop": "2px",
                       "wordBreak": "break-word"},
            )

        rows.append(html.Div([
            html.Span(
                cr["cr_id"],
                style={"width": "148px", "fontSize": "11px", "fontWeight": "700",
                       "color": BRAND, "padding": "7px 10px", "flexShrink": "0"},
            ),
            html.Span(
                (cr.get("institution_name") or cr["le_book"]).title(),
                style={"flex": "1", "fontSize": "12px", "color": TEXT,
                       "padding": "7px 10px", "overflow": "hidden",
                       "textOverflow": "ellipsis", "whiteSpace": "nowrap"},
            ),
            html.Div([
                html.Span(cr["title"],
                          style={"fontSize": "12px", "color": TEXT,
                                 "display": "block", "lineHeight": "1.3"}),
                html.Span(cr.get("description") or "",
                          style={"fontSize": "10px", "color": MUTED,
                                 "display": "block", "lineHeight": "1.3",
                                 "overflow": "hidden", "textOverflow": "ellipsis",
                                 "whiteSpace": "nowrap",
                                 "maxWidth": "260px"})
                if cr.get("description") else html.Span(),
            ], style={"flex": "2", "padding": "5px 10px", "overflow": "hidden",
                      "minWidth": "0"}),
            html.Div(status_chip,
                     style={"width": "108px", "padding": "7px 10px",
                            "flexShrink": "0"}),
            html.Span(
                str(n_issues),
                style={"width": "56px", "textAlign": "center", "fontSize": "12px",
                       "color": TEXT, "padding": "7px 10px", "flexShrink": "0"},
            ),
            html.Span(
                f"{cr.get('failing_rows', 0):,}",
                style={"width": "74px", "textAlign": "right", "fontSize": "12px",
                       "fontWeight": "700", "color": TEXT,
                       "padding": "7px 10px", "flexShrink": "0"},
            ),
            html.Span(
                cr.get("target_date") or "—",
                style={"width": "86px", "fontSize": "11px", "color": MUTED,
                       "padding": "7px 10px", "flexShrink": "0"},
            ),
            html.Span(
                cr.get("assigned_to") or "—",
                style={"width": "154px", "fontSize": "11px", "color": MUTED,
                       "padding": "7px 10px", "flexShrink": "0",
                       "overflow": "hidden", "textOverflow": "ellipsis",
                       "whiteSpace": "nowrap"},
            ),
            html.Div(
                action_btns + ([reviewer_note] if reviewer_note.children else []),  # type: ignore[attr-defined]
                style={"width": "228px", "padding": "5px 10px",
                       "display": "flex", "alignItems": "flex-start",
                       "flexWrap": "wrap", "flexShrink": "0"},
            ),
        ], style={
            "display":      "flex",
            "alignItems":   "center",
            "background":   bg,
            "borderBottom": f"1px solid {DIVIDER}",
            "minWidth":     "0",
        }))

    return html.Div(
        [hdr] + rows,
        style={"border": f"1px solid {DIVIDER}", "borderRadius": "8px",
               "overflow": "hidden", "overflowX": "auto"},
    )


def _remediation_page(role: str = "bnr_admin") -> html.Div:
    """
    Full Data Quality Remediation page.

    Implements the SAP MDG DQR workflow:
      1. Specialist selects open issues (filtered by institution / urgency).
      2. Creates a Change Request (CR) linking the selected issues.
      3. Assigned data officer marks CR In Progress then Submitted.
      4. BNR specialist Approves or Rejects with review notes.
      5. Approved CRs are Closed once the next pipeline run confirms resolution.
    """
    import dq_change_request as cr_mod
    from dq_issue_tracker import get_open_issues, ensure_tables as _ensure_issues

    _ensure_issues()
    cr_mod.ensure_table()

    # Institution options for the create-CR form
    open_issues = get_open_issues()
    inst_seen: dict[str, str] = {}
    for iss in open_issues:
        lb = iss["le_book"]
        if lb not in inst_seen:
            inst_seen[lb] = (iss.get("institution_name") or lb).title()

    inst_options = [{"label": "Select institution…", "value": ""}] + [
        {"label": name, "value": lb}
        for lb, name in sorted(inst_seen.items(), key=lambda kv: kv[1])
    ]

    # Summary banner
    stats  = cr_mod.get_stats()
    chips  = []
    for key, lbl in cr_mod.STATUS_LABELS.items():
        n   = stats.get(key, 0)
        clr = cr_mod.STATUS_COLORS[key]
        chips.append(html.Div([
            html.Span(str(n), style={
                "fontSize": "24px", "fontWeight": "900",
                "color": clr, "lineHeight": "1",
            }),
            html.Span(lbl, style={
                "fontSize": "10px", "color": MUTED,
                "marginTop": "3px", "lineHeight": "1.2",
                "textAlign": "center",
            }),
        ], style={
            "display":        "flex",
            "flexDirection":  "column",
            "alignItems":     "center",
            "background":     CARD,
            "borderRadius":   "8px",
            "padding":        "12px 18px",
            "border":         f"2px solid {clr}",
            "minWidth":       "90px",
        }))

    summary_bar = html.Div(chips, id="cr-summary-bar", style={
        "display": "flex", "gap": "10px", "flexWrap": "wrap",
        "marginBottom": "24px",
    })

    INP = {
        "width": "100%", "padding": "8px 10px",
        "border": f"1px solid {DIVIDER}", "borderRadius": "6px",
        "fontSize": "12px", "fontFamily": FONT,
        "background": BG, "color": TEXT,
        "boxSizing": "border-box", "outline": "none",
    }

    # ── Create-CR form (hidden by default, toggled) ────────────────────────────
    form_panel = html.Div([
        html.Div("NEW CHANGE REQUEST", style={
            "fontSize": "11px", "fontWeight": "900", "color": MUTED,
            "letterSpacing": "0.06em", "textTransform": "uppercase",
            "marginBottom": "16px",
        }),

        # Row 1: institution selector
        html.Div([
            html.Label("Institution *", style={
                "fontSize": "11px", "fontWeight": "900", "color": MUTED,
                "display": "block", "marginBottom": "5px",
            }),
            dcc.Dropdown(
                id="cr-inst-filter",
                options=inst_options,
                value="",
                clearable=False,
                placeholder="Select institution…",
                style={"fontSize": "12px", "fontFamily": FONT},
            ),
            html.Div(
                "Only institutions with open issues appear here.",
                style={"fontSize": "10px", "color": MUTED, "marginTop": "4px"},
            ),
        ], style={"marginBottom": "16px"}),

        # Row 2: issue checklist (options filled by callback)
        html.Div([
            html.Label("Select Issues to Address *", style={
                "fontSize": "11px", "fontWeight": "900", "color": MUTED,
                "display": "block", "marginBottom": "6px",
            }),
            html.Div(
                dcc.Checklist(
                    id="cr-issue-checklist",
                    options=[],
                    value=[],
                    inputStyle={"marginRight": "6px"},
                    labelStyle={
                        "display": "block",
                        "fontSize": "12px",
                        "lineHeight": "1.8",
                        "color": TEXT,
                        "cursor": "pointer",
                    },
                ),
                id="cr-issue-list",
                style={
                    "background":   BG,
                    "padding":      "10px 14px",
                    "borderRadius": "6px",
                    "border":       f"1px solid {DIVIDER}",
                    "maxHeight":    "220px",
                    "overflowY":    "auto",
                    "minHeight":    "44px",
                },
            ),
            html.Div(
                "Select an institution above to see its tables with open issues.",
                id="cr-issue-hint",
                style={"fontSize": "10px", "color": MUTED, "marginTop": "4px"},
            ),
            html.Div(id="cr-table-dl-area", style={"marginTop": "6px"}),
        ], style={"marginBottom": "16px"}),

        # Row 3: title
        html.Div([
            html.Label("Change Request Title *", style={
                "fontSize": "11px", "fontWeight": "900", "color": MUTED,
                "display": "block", "marginBottom": "5px",
            }),
            dcc.Input(
                id="cr-title",
                type="text",
                placeholder="Brief description of the correction required…",
                debounce=False,
                style={**INP},
            ),
        ], style={"marginBottom": "14px"}),

        # Row 4: description
        html.Div([
            html.Label("Correction Plan / Description", style={
                "fontSize": "11px", "fontWeight": "900", "color": MUTED,
                "display": "block", "marginBottom": "5px",
            }),
            dcc.Textarea(
                id="cr-description",
                placeholder=(
                    "Describe what data needs correcting, which records are affected, "
                    "and the steps the institution must follow to resolve the issue…"
                ),
                style={**INP, "height": "80px", "resize": "vertical"},
            ),
        ], style={"marginBottom": "14px"}),

        # Row 5: assigned to + target date (side by side)
        html.Div([
            html.Div([
                html.Label("Assigned To", style={
                    "fontSize": "11px", "fontWeight": "900", "color": MUTED,
                    "display": "block", "marginBottom": "5px",
                }),
                dcc.Dropdown(
                    id="cr-assigned-to",
                    options=[],
                    value=None,
                    placeholder="Select an institution first…",
                    clearable=True,
                    style={"fontSize": "12px", "fontFamily": FONT},
                ),
            ], style={"flex": "1"}),
            html.Div([
                html.Label("Target Resolution Date", style={
                    "fontSize": "11px", "fontWeight": "900", "color": MUTED,
                    "display": "block", "marginBottom": "5px",
                }),
                dcc.Input(
                    id="cr-target-date",
                    type="date",
                    debounce=False,
                    style={**INP},
                ),
            ], style={"flex": "1"}),
        ], style={"display": "flex", "gap": "16px", "marginBottom": "18px"}),

        # Submit row
        html.Div([
            html.Div("Create Change Request", id="cr-create-btn", n_clicks=0,
                     style={
                         "display":      "inline-block",
                         "background":   BRAND,
                         "color":        CARD,
                         "padding":      "9px 20px",
                         "borderRadius": "6px",
                         "fontSize":     "12px",
                         "fontWeight":   "900",
                         "cursor":       "pointer",
                         "userSelect":   "none",
                         "letterSpacing": "0.03em",
                     }),
            html.Div(id="cr-feedback",
                     style={"marginTop": "8px", "fontSize": "12px",
                            "lineHeight": "1.4"}),
        ]),
    ], id="cr-form-panel", style={
        "background":   CARD,
        "border":       f"1px solid {DIVIDER}",
        "borderRadius": "8px",
        "padding":      "20px 24px",
        "marginBottom": "20px",
        "display":      "none",
    })

    # ── Review-notes box (shared; reviewer fills in before clicking Approve/Reject)
    review_box = html.Div([
        html.Div("REVIEWER NOTES", style={
            "fontSize": "11px", "fontWeight": "900", "color": MUTED,
            "letterSpacing": "0.06em", "textTransform": "uppercase",
            "marginBottom": "6px",
        }),
        html.P(
            "Fill in your notes here before clicking Approve or Reject on a submitted CR below.",
            style={"fontSize": "11px", "color": MUTED, "margin": "0 0 8px"},
        ),
        dcc.Textarea(
            id="cr-review-notes",
            placeholder="e.g. Verified in the source system — 47 customer records corrected.",
            style={
                **INP,
                "height":  "56px",
                "resize":  "vertical",
            },
        ),
        html.Div(id="cr-action-feedback",
                 style={"marginTop": "6px", "fontSize": "12px", "lineHeight": "1.4"}),
    ], style={
        "background":   CARD,
        "border":       f"1px solid {DIVIDER}",
        "borderRadius": "8px",
        "padding":      "16px 20px",
        "marginBottom": "20px",
    })

    # ── CR list section ────────────────────────────────────────────────────────
    status_options = [{"label": "All Statuses", "value": "all"}] + [
        {"label": lbl, "value": key}
        for key, lbl in cr_mod.STATUS_LABELS.items()
    ]

    cr_list_section = html.Div([
        html.Div([
            html.Div("CHANGE REQUESTS", style={
                "fontSize": "11px", "fontWeight": "900", "color": MUTED,
                "letterSpacing": "0.06em", "textTransform": "uppercase",
            }),
            dcc.Dropdown(
                id="cr-status-filter",
                options=status_options,
                value="all",
                clearable=False,
                style={"fontSize": "12px", "fontFamily": FONT, "minWidth": "170px"},
            ),
        ], style={
            "display": "flex", "alignItems": "center",
            "justifyContent": "space-between",
            "marginBottom": "14px",
        }),
        html.Div(id="cr-list-container"),
    ], style={
        "background":   CARD,
        "border":       f"1px solid {DIVIDER}",
        "borderRadius": "8px",
        "padding":      "20px",
    })

    return html.Div([

        # Page header
        html.Div([
            html.H2("Data Quality Remediation", style={
                "fontSize": "18px", "fontWeight": "900", "color": TEXT,
                "margin": "0", "lineHeight": "1.2",
            }),
        ], style={"marginBottom": "24px"}),

        # Status summary bar
        summary_bar,

        # + New Change Request toggle — BNR admin only
        *([] if not dq_auth.is_admin(role) else [html.Div(
            html.Div("+ New Change Request", id="cr-form-toggle-btn", n_clicks=0,
                     style={
                         "display":      "inline-block",
                         "background":   BRAND,
                         "color":        CARD,
                         "padding":      "9px 18px",
                         "borderRadius": "6px",
                         "fontSize":     "12px",
                         "fontWeight":   "900",
                         "cursor":       "pointer",
                         "userSelect":   "none",
                     }),
            style={"marginBottom": "16px"},
        )]),

        form_panel,
        review_box,
        cr_list_section,

    ], style={"padding": "28px 32px", "maxWidth": "1400px", "margin": "0 auto"})


# ── app ────────────────────────────────────────────────────────────────────────

app = dash.Dash(
    __name__,
    title="BNR Data Quality Monitoring",
    suppress_callback_exceptions=True,
)
server = app.server  # exposed for gunicorn: gunicorn dq_dashboard_dash:server
server.secret_key = os.environ.get("SECRET_KEY", os.urandom(32))


@server.route("/download/issue-report/<le_book>")
def _serve_issue_report(le_book):
    """Stream an institution's latest issue-report ZIP directly (no base64 inlining).
    Used by the download buttons so large reports (100 MB+) download reliably."""
    import re as _re
    from flask import abort, send_file as _flask_send_file
    if not _re.fullmatch(r"[A-Za-z0-9_-]{1,20}", le_book or ""):
        abort(400)
    d = _DIR / "issue_reports"
    zips = sorted(d.glob(f"{le_book}_*.zip"), reverse=True) if d.exists() else []
    if not zips:
        abort(404)
    return _flask_send_file(str(zips[0]), as_attachment=True, download_name=zips[0].name)


@server.route("/download/issue-report/<le_book>/<table>")
def _serve_issue_table(le_book, table):
    """Stream a single {table}.xlsx out of an institution's latest report ZIP."""
    import re as _re, io as _io, zipfile as _zip
    from flask import abort, send_file as _flask_send_file
    if (not _re.fullmatch(r"[A-Za-z0-9_-]{1,20}", le_book or "")
            or not _re.fullmatch(r"[A-Za-z0-9_]{1,40}", table or "")):
        abort(400)
    d = _DIR / "issue_reports"
    zips = sorted([z for z in d.glob(f"{le_book}_*.zip")
                   if not z.name.endswith("_resolved.zip")], reverse=True) if d.exists() else []
    if not zips:
        abort(404)
    member = f"{table}.xlsx"
    try:
        with _zip.ZipFile(zips[0]) as zf:
            if member not in zf.namelist():
                abort(404)
            data = zf.read(member)
    except Exception:
        abort(404)
    month = zips[0].stem.split("_", 1)[-1]
    return _flask_send_file(
        _io.BytesIO(data), as_attachment=True,
        download_name=f"{table}_{le_book}_{month}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


dq_auth.ensure_users_table()


def _login_page(error: str = "", login_type: str = "bnr") -> html.Div:
    inp = {
        "width": "100%", "padding": "10px 12px",
        "border": f"1px solid {DIVIDER}", "borderRadius": "6px",
        "fontSize": "14px", "fontFamily": FONT, "color": TEXT,
        "boxSizing": "border-box", "outline": "none", "background": CARD,
    }

    is_bnr = (login_type == "bnr")

    def _tab(label, tab_id, active):
        return html.Div(
            label,
            id=tab_id,
            n_clicks=0,
            style={
                "flex": "1", "textAlign": "center",
                "padding": "11px 0",
                "fontSize": "12px", "fontWeight": "900",
                "cursor": "pointer", "userSelect": "none",
                "letterSpacing": "0.04em",
                "background":   CARD                       if active else "rgba(117,57,24,0.10)",
                "color":        BRAND                      if active else MUTED,
                "borderBottom": f"3px solid {BRAND}"      if active else f"3px solid transparent",
                "transition":   "all 0.15s",
            },
        )

    tab_bar = html.Div([
        _tab("🏛  BNR Staff",      "login-tab-bnr",  is_bnr),
        _tab("🏦  Institution",    "login-tab-inst", not is_bnr),
    ], style={
        "display": "flex",
        "borderBottom": f"1px solid {DIVIDER}",
        "background": CARD,
    })

    placeholder = "your.name@bnr.rw" if is_bnr else "focal.point@yourbank.com"
    hint = (
        "BNR staff accounts only — @bnr.rw required."
        if is_bnr else
        "Use the email address provided to you by BNR."
    )
    btn_color = BRAND

    return html.Div([
        html.Div([
            # Card header
            html.Div([
                html.Img(src="/assets/bnr_img.png",
                         style={"height": "48px", "marginBottom": "10px"}),
                html.Div("DATA QUALITY PROGRAM", style={
                    "fontSize": "13px", "fontWeight": "900", "color": CARD,
                    "letterSpacing": "0.07em",
                }),
                html.Div("National Bank of Rwanda", style={
                    "fontSize": "11px", "color": "rgba(255,255,255,0.65)", "marginTop": "3px",
                }),
            ], style={
                "background": BRAND, "padding": "28px 32px 22px",
                "textAlign": "center", "borderRadius": "12px 12px 0 0",
            }),

            # Login type tabs
            tab_bar,

            # Form body
            html.Div([
                html.Div(
                    "BNR Staff Sign In" if is_bnr else "Institution Sign In",
                    style={
                        "fontSize": "15px", "fontWeight": "900", "color": TEXT,
                        "marginBottom": "22px", "textAlign": "center",
                    },
                ),

                html.Div("Email", style={
                    "fontSize": "11px", "fontWeight": "900", "color": MUTED,
                    "textTransform": "uppercase", "letterSpacing": "0.05em",
                    "marginBottom": "5px",
                }),
                dcc.Input(
                    id="login-email", type="email",
                    placeholder=placeholder,
                    debounce=False, n_submit=0,
                    style={**inp, "marginBottom": "16px"},
                ),

                html.Div("Password", style={
                    "fontSize": "11px", "fontWeight": "900", "color": MUTED,
                    "textTransform": "uppercase", "letterSpacing": "0.05em",
                    "marginBottom": "5px",
                }),
                dcc.Input(
                    id="login-password", type="password",
                    placeholder="••••••••",
                    debounce=False, n_submit=0,
                    style={**inp, "marginBottom": "8px"},
                ),

                html.Div(hint, id="login-hint", style={
                    "fontSize": "11px", "color": MUTED, "marginBottom": "20px",
                }),

                html.Div(
                    error,
                    id="login-error",
                    style={
                        "fontSize": "12px", "color": C_RED, "marginBottom": "14px",
                        "minHeight": "16px", "textAlign": "center",
                        "display": "block" if error else "none",
                    },
                ),

                html.Div(
                    "Sign In",
                    id="login-btn",
                    n_clicks=0,
                    style={
                        "width": "100%", "padding": "11px 0",
                        "background": btn_color, "color": CARD,
                        "fontSize": "14px", "fontWeight": "900",
                        "textAlign": "center", "borderRadius": "6px",
                        "cursor": "pointer", "userSelect": "none",
                        "letterSpacing": "0.04em",
                    },
                ),
            ], style={"padding": "28px 32px", "background": BG}),

        ], style={
            "background": CARD, "borderRadius": "12px",
            "boxShadow": "0 8px 32px rgba(0,0,0,0.14)",
            "width": "400px",
        }),
    ], style={
        "display": "flex", "alignItems": "center", "justifyContent": "center",
        "minHeight": "100vh", "background": BG,
    })


def _nav_tabs(active: str) -> html.Div:
    items = [("dashboard", "Dashboard"), ("alerts", "Alerts"), ("remediation", "Remediation"), ("validations", "Validations")]
    tabs = []
    for key, label in items:
        is_active = key == active
        tabs.append(html.Div(
            label,
            id={"type": "page-nav", "index": key},
            n_clicks=0,
            style={
                "cursor":       "pointer",
                "padding":      "11px 24px",
                "fontSize":     "13px",
                "fontWeight":   "900" if is_active else "400",
                "color":        CARD if is_active else "rgba(255,255,255,0.60)",
                "borderBottom": f"3px solid {BNR_GOLD}" if is_active
                                else "3px solid transparent",
                "whiteSpace":   "nowrap",
                "userSelect":   "none",
                "transition":   "color .15s, border-color .15s",
            },
        ))
    return html.Div(tabs, style={
        "display":    "flex",
        "background": "#753918",
        "padding":    "0 32px",
        "borderTop":  "1px solid rgba(255,255,255,0.12)",
    })


app.layout = html.Div([

    # ── header ────────────────────────────────────────────────────────────────
    html.Div([
        html.Div([
            html.Img(
                src="/assets/bnr_img.png",
                style={"height": "50px", "marginRight": "16px", "flexShrink": "0"},
            ),
            html.Div([
                html.Div("DATA QUALITY MONITORING", style={
                    "fontSize": "14px", "fontWeight": "700",
                    "color": CARD, "letterSpacing": "0.06em",
                    "lineHeight": "1.15",
                }),
                html.Div(
                    "National Bank of Rwanda — Data Quality Program",
                    style={
                        "fontSize": "11px", "fontWeight": "400",
                        "color": "rgba(255,255,255,0.65)",
                        "lineHeight": "1.15", "marginTop": "3px",
                    },
                ),
            ]),
        ], style={"display": "flex", "alignItems": "center"}),
        html.Div([
            html.Div(id="pipeline-status-banner", style={
                "textAlign": "right", "lineHeight": "1.5",
            }),
            html.Div(id="user-info-header"),
        ], style={"display": "flex", "alignItems": "center", "gap": "20px"}),
    ], style={
        "background":     "#753918",
        "padding":        "14px 32px",
        "display":        "flex",
        "alignItems":     "center",
        "justifyContent": "space-between",
        "boxShadow":      "0 2px 8px rgba(0,0,0,0.18)",
    }),

    # ── page nav ──────────────────────────────────────────────────────────────
    html.Div(id="page-nav-bar"),

    # ── page content ──────────────────────────────────────────────────────────
    html.Div(id="page-content", style={
        "maxWidth":   "1440px",
        "margin":     "0 auto",
        "padding":    "24px 32px",
        "fontFamily": FONT,
    }),

    # ── notification overlay (fixed, shown on top of page content) ────────────
    html.Div(id="notif-overlay", style={
        "position": "fixed", "top": "52px", "right": "80px",
        "zIndex": "500", "display": "none",
    }),

    # ── stores ────────────────────────────────────────────────────────────────
    # nav-state: {"cat": None|"B"|"MF"|"SACCO", "inst": None|"<code>"}
    # cat=None means landing page; inst=None means show all in category
    dcc.Interval(id="status-poll",   interval=30_000,  n_intervals=0),
    dcc.Interval(id="notif-poll",    interval=60_000,  n_intervals=0),
    dcc.Store(id="nav-state",        data={"cat": None, "inst": None}),
    dcc.Store(id="active-page",      data="dashboard"),
    dcc.Store(id="rules-version",    data=0),
    dcc.Store(id="cr-version",       data=0),
    dcc.Store(id="notify-status",    data={}),
    dcc.Store(id="auth-store",       data={}),
    dcc.Store(id="inst-active-page", data="inst_dashboard"),
    dcc.Store(id="inst-notif-show",  data=False),
    dcc.Store(id="login-type",       data="bnr"),
    dcc.Download(id="inst-download"),
    dcc.Download(id="issues-download"),
    dcc.Download(id="inst-csv-download"),
    dcc.Download(id="resolved-inst-download"),
    dcc.Download(id="open-issue-dl"),
    dcc.Location(id="dl-nav", refresh=True),  # drives large-file downloads via /download route
    dcc.Download(id="cr-tbl-dl"),
    dcc.Store(id="resolved-dl-lb",  data=None),
    dcc.Store(id="dl-preview-lb", data=None),

    # ── download preview modal ────────────────────────────────────────────────
    html.Div(
        id="dl-preview-modal",
        style={"display": "none"},
        children=[
            # backdrop
            html.Div(style={
                "position": "fixed", "inset": "0",
                "background": "rgba(28,28,39,0.55)", "zIndex": "900",
            }),
            # dialog
            html.Div([
                # header
                html.Div([
                    html.Div([
                        html.Span("⬇", style={"fontSize": "18px", "marginRight": "10px",
                                              "color": BRAND}),
                        html.Span(id="dl-modal-title",
                                  style={"fontSize": "15px", "fontWeight": "700",
                                         "color": TEXT}),
                    ], style={"display": "flex", "alignItems": "center"}),
                    html.Div(id="dl-modal-subtitle",
                             style={"fontSize": "12px", "color": MUTED,
                                    "marginTop": "2px"}),
                ], style={
                    "padding": "18px 22px 14px",
                    "borderBottom": f"1px solid {DIVIDER}",
                }),

                # issues table
                html.Div(
                    id="dl-modal-table",
                    style={
                        "maxHeight": "420px", "overflowY": "auto",
                        "padding": "0",
                    },
                ),

                # footer buttons
                html.Div([
                    html.Div("Cancel", id="dl-modal-cancel", n_clicks=0,
                             style={
                                 "padding": "8px 22px", "borderRadius": "5px",
                                 "border": f"1px solid {DIVIDER}",
                                 "fontSize": "13px", "fontWeight": "600",
                                 "color": MUTED, "cursor": "pointer",
                                 "background": CARD,
                             }),
                    html.Div("Download Report", id="dl-modal-confirm", n_clicks=0,
                             style={
                                 "padding": "8px 22px", "borderRadius": "5px",
                                 "fontSize": "13px", "fontWeight": "700",
                                 "color": CARD, "cursor": "pointer",
                                 "background": BRAND,
                             }),
                ], style={
                    "display": "flex", "justifyContent": "flex-end",
                    "gap": "10px", "padding": "14px 22px",
                    "borderTop": f"1px solid {DIVIDER}",
                }),
            ], style={
                "position": "fixed",
                "top": "50%", "left": "50%",
                "transform": "translate(-50%, -50%)",
                "zIndex": "901",
                "background": CARD,
                "borderRadius": "10px",
                "boxShadow": "0 8px 40px rgba(28,28,39,0.22)",
                "width": "min(860px, 92vw)",
                "fontFamily": FONT,
            }),
        ],
    ),

], style={"background": BG, "minHeight": "100vh", "fontFamily": FONT})


# ── callbacks ──────────────────────────────────────────────────────────────────

@app.callback(
    Output({"type": "null-col-body", "index": MATCH}, "style"),
    Output({"type": "null-col-btn",  "index": MATCH}, "children"),
    Input({"type": "null-col-btn",   "index": MATCH}, "n_clicks"),
    State({"type": "null-col-body",  "index": MATCH}, "style"),
    prevent_initial_call=True,
)
def _toggle_null_cols(_n, current_style):
    is_open  = (current_style or {}).get("display") != "none"
    new_style = {"display": "none" if is_open else "block", "marginTop": "4px"}
    label     = "▼ hide" if not is_open else "▶ show"
    return new_style, label


@app.callback(
    Output("pipeline-status-banner", "children"),
    Output("status-poll", "interval"),
    Input("status-poll", "n_intervals"),
)
def _update_pipeline_banner(_):
    """
    Refresh pipeline status. Poll every 5 s while running, 30 s otherwise.
    When running, shows live stage detection + log tail instead of a single line.
    """
    from datetime import datetime as _dt
    import re as _re

    run    = _fresh_pipeline()
    status = _load_pipeline_status()
    data_date = run.get("run_date", "—")
    s = status.get("status", "")

    # ── dynamic poll interval ──────────────────────────────────────────────────
    poll_ms = 5_000 if s == "running" else 30_000

    # ── not running: compact banner line ──────────────────────────────────────
    if s != "running":
        if s == "success":
            color, label = "#4ADE80", "Success"
            ts_raw = status.get("finished_at", "")
        elif s == "failed":
            color, label = "#F87171", "Failed"
            ts_raw = status.get("finished_at", "")
        else:
            color = label = ts_raw = ""

        ts_lbl = f"{ts_raw[11:16]}" if len(ts_raw or "") >= 16 else ""
        children = [
            html.Span(f"Data as of: {data_date}",
                      style={"color": "rgba(255,255,255,0.55)"}),
        ]
        if label:
            children += [
                html.Span("  ·  ", style={"color": "rgba(255,255,255,0.30)"}),
                html.Span(f"● {label}", style={"color": color, "fontWeight": "700"}),
                html.Span(f"  ({ts_lbl})" if ts_lbl else "",
                          style={"color": "rgba(255,255,255,0.40)"}),
            ]
        return html.Span(children, style={"fontSize": "11px"}), poll_ms

    # ── running: rich live panel ───────────────────────────────────────────────
    started_raw = status.get("started_at", "")
    try:
        started  = _dt.strptime(started_raw, "%Y-%m-%d %H:%M:%S")
        elapsed  = _dt.now() - started
        mins, sec = divmod(int(elapsed.total_seconds()), 60)
        elapsed_str = f"{mins}m {sec:02d}s"
    except Exception:
        elapsed_str = "—"

    # Read today's log file for stage + table detection
    from datetime import date as _date
    log_path = _DIR / "logs" / f"pipeline_{_date.today().isoformat()}.log"
    log_lines: list[str] = []
    stage_label = "Initialising…"
    current_table = ""
    try:
        with open(log_path, "r", errors="replace") as f:
            all_lines = f.readlines()
        log_lines = [l.rstrip() for l in all_lines if l.strip()][-6:]

        # Detect current stage from log content
        full_text = "".join(all_lines)
        if "Stage 3" in full_text or "Resolution scan" in full_text:
            stage_label = "Stage 3 — Resolution scanner"
        elif "Stage 2" in full_text or "RI checks" in full_text or "--reports" in full_text:
            stage_label = "Stage 2 — RI checks + XLSX reports"
        elif "Stage 1" in full_text or "--load" in full_text:
            stage_label = "Stage 1 — SQL engines + scoring"

        # Detect current table from most recent ━━ separator
        tables_seen = _re.findall(r"━━\s+([\w]+)", full_text)
        if tables_seen:
            current_table = tables_seen[-1].replace("_", " ")
    except Exception:
        pass

    # Compact log tail rows
    _LOG_STYLE = {
        "fontFamily": "monospace", "fontSize": "9px",
        "color": "rgba(255,255,255,0.55)", "lineHeight": "1.4",
        "display": "block", "whiteSpace": "nowrap",
        "overflow": "hidden", "textOverflow": "ellipsis",
        "maxWidth": "480px",
    }
    log_tail = [html.Span(l, style=_LOG_STYLE) for l in log_lines[-4:]]

    panel = html.Div([
        # Top row: animated dot + stage + elapsed
        html.Div([
            html.Span("⬤ ", style={
                "color": "#FCD34D", "fontSize": "10px",
                "animation": "pulse 1.2s ease-in-out infinite",
            }),
            html.Span(stage_label, style={
                "color": "#FCD34D", "fontWeight": "700", "fontSize": "11px",
            }),
            html.Span(f"  ·  {elapsed_str}", style={
                "color": "rgba(255,255,255,0.50)", "fontSize": "10px",
            }),
            *([ html.Span(f"  ·  {current_table}", style={
                    "color": "rgba(255,255,255,0.40)", "fontSize": "10px",
                    "fontStyle": "italic",
                })] if current_table else []),
        ], style={"marginBottom": "3px"}),
        # Log tail
        html.Div(log_tail),
    ], style={"textAlign": "right"})

    return panel, poll_ms


@app.callback(
    Output("active-page", "data"),
    Input({"type": "page-nav", "index": ALL}, "n_clicks"),
    prevent_initial_call=True,
)
def _on_page_nav(_n_clicks):
    triggered = ctx.triggered_id
    if isinstance(triggered, dict) and "index" in triggered:
        return triggered["index"]
    raise dash.exceptions.PreventUpdate


@app.callback(
    Output("page-nav-bar", "children"),
    Output("page-content",  "children"),
    Input("active-page",    "data"),
    Input("nav-state",      "data"),
    Input("rules-version",  "data"),
    Input("auth-store",     "data"),
)
def _render_page(page: str, nav_state, _rv, auth_data):
    auth = auth_data or {}
    if auth.get("tab_switch"):
        lt = auth.get("tab_switch", "bnr")
        return html.Div(), _login_page(login_type=lt)
    if auth.get("error") or not auth.get("email"):
        lt = auth.get("tab", "bnr")
        return html.Div(), _login_page(auth.get("error", ""), login_type=lt)

    role     = auth.get("role", "viewer")
    le_books = auth.get("le_books", [])

    # ── institution user → hand off to inst portal ────────────────────────────
    if role == "inst_user":
        import dq_inst_portal as inst_mod
        import json as _json
        cats_path = Path(__file__).parent / "le_book_categories.json"
        try:
            categories = _json.loads(cats_path.read_text())
        except Exception:
            categories = {}

        page = page or "inst_dashboard"
        uid  = auth.get("user_id", "")
        try:
            from dq_notifications import get_unread_count
            unread = get_unread_count(uid) if uid else 0
        except Exception:
            unread = 0

        nav_bar = inst_mod.inst_nav_bar(
            active_page=page,
            user_name=auth.get("name", ""),
            le_book=le_books[0] if le_books else "—",
            unread_count=unread,
        )

        if page == "inst_issues":
            content = inst_mod.inst_issues_page(le_books)
        elif page == "inst_remediation":
            content = inst_mod.inst_remediation_page(le_books, role=role)
        elif page == "inst_validations":
            content = _validations_page()
        else:
            content = inst_mod.inst_dashboard_page(le_books, categories)

        return nav_bar, content

    # ── BNR user → existing portal ────────────────────────────────────────────
    page = page or "dashboard"
    nav  = nav_state or {"cat": None, "inst": None}
    cat  = nav.get("cat")
    inst = nav.get("inst")

    nav_bar = _nav_tabs(page)

    if page == "validations":
        return nav_bar, _validations_page()

    if page == "alerts":
        return nav_bar, _alerts_page(cat=cat or "")

    if page == "remediation":
        return nav_bar, _remediation_page(role=role)

    if not cat:
        return nav_bar, _landing_page(_counts)

    return nav_bar, _dashboard_content(cat, inst)


@app.callback(
    Output("nav-state", "data"),
    Input({"type": "cat-landing-btn", "index": ALL}, "n_clicks"),
    Input({"type": "nav-action",      "index": ALL}, "n_clicks"),
    Input({"type": "inst-dd",         "index": ALL}, "value"),
    State("nav-state", "data"),
    prevent_initial_call=True,
)
def _nav_handler(landing_clicks, nav_action_clicks, inst_values, current_nav):
    """Single callback that owns all navigation state changes."""
    nav = dict(current_nav or {"cat": None, "inst": None})
    tid = ctx.triggered_id
    triggered_val = ctx.triggered[0]["value"] if ctx.triggered else None

    if isinstance(tid, dict):
        t = tid.get("type")

        if t == "nav-action" and tid.get("index") == "back":
            if triggered_val and triggered_val > 0:
                return {"cat": None, "inst": None}
            raise dash.exceptions.PreventUpdate

        if t == "cat-landing-btn":
            if triggered_val and triggered_val > 0:
                return {"cat": tid["index"], "inst": None}
            raise dash.exceptions.PreventUpdate

        if t == "inst-dd":
            new_inst = triggered_val or None
            new_nav  = {**nav, "inst": new_inst}
            # prevent spurious re-renders when the dropdown first appears in the DOM
            if new_nav == nav:
                raise dash.exceptions.PreventUpdate
            return new_nav

    raise dash.exceptions.PreventUpdate


@app.callback(
    Output("unscored-body",   "style"),
    Output("unscored-toggle", "children"),
    Input("unscored-toggle",  "n_clicks"),
    State("unscored-body",    "style"),
    State("unscored-toggle",  "children"),
    prevent_initial_call=True,
)
def _toggle_unscored(n_clicks, body_style, btn_text):
    visible = (body_style or {}).get("display") != "none"
    if visible:
        new_label = (btn_text or "").replace("▼", "▶").replace("Hide", "Show")
        return {"display": "none"}, new_label
    new_label = (btn_text or "").replace("▶", "▼").replace("Show", "Hide")
    return {"display": "block"}, new_label


@app.callback(
    Output("rule-form-panel", "style"),
    Input("form-toggle-btn",  "n_clicks"),
    State("rule-form-panel",  "style"),
    prevent_initial_call=True,
)
def _toggle_form(n_clicks, current_style):
    style = dict(current_style or {})
    style["display"] = "none" if style.get("display") != "none" else "block"
    return style


@app.callback(
    Output("param-domain",  "style"),
    Output("param-range",   "style"),
    Output("param-pattern", "style"),
    Input("new-rule-check-type", "value"),
    prevent_initial_call=True,
)
def _show_params(check_type):
    hidden  = {"display": "none"}
    visible = {"display": "block", "marginBottom": "12px"}
    grid_v  = {"display": "grid", "gridTemplateColumns": "1fr 1fr",
                "gap": "14px", "marginBottom": "12px"}
    return (
        visible  if check_type == "domain"  else hidden,
        grid_v   if check_type == "range"   else hidden,
        visible  if check_type == "pattern" else hidden,
    )


@app.callback(
    Output("new-rule-feedback", "children"),
    Output("rules-version",     "data"),
    Input("new-rule-submit",    "n_clicks"),
    State("new-rule-id",         "value"),
    State("new-rule-dim",        "value"),
    State("new-rule-cat",        "value"),
    State("new-rule-name",       "value"),
    State("new-rule-table",      "value"),
    State("new-rule-field",      "value"),
    State("new-rule-check-type", "value"),
    State("new-rule-domain-vals","value"),
    State("new-rule-range-min",  "value"),
    State("new-rule-range-max",  "value"),
    State("new-rule-pattern",    "value"),
    State("rules-version",       "data"),
    prevent_initial_call=True,
)
def _submit_rule(n_clicks, rule_id, dim, cat, name, table, field,
                 check_type, domain_vals, range_min, range_max, pattern, version):
    if not n_clicks:
        raise dash.exceptions.PreventUpdate

    errors = []
    if not (rule_id or "").strip():
        errors.append("Rule ID is required.")
    if not dim:
        errors.append("Dimension is required.")
    if not (name or "").strip():
        errors.append("Rule Name is required.")
    if not table:
        errors.append("Table is required.")
    if not (field or "").strip():
        errors.append("Field is required.")
    if not check_type:
        errors.append("Check Type is required.")
    if check_type == "domain" and not (domain_vals or "").strip():
        errors.append("Allowed Values are required for a domain check.")
    if check_type == "range" and (range_min is None or range_max is None):
        errors.append("Both Min and Max are required for a range check.")
    if check_type == "pattern" and not (pattern or "").strip():
        errors.append("Regex Pattern is required for a pattern check.")

    if errors:
        return (
            html.Span("  ".join(errors), style={"color": C_RED}),
            version,
        )

    import json as _json
    check_params = None
    if check_type == "domain":
        vals = [v.strip() for v in domain_vals.split(",") if v.strip()]
        check_params = _json.dumps({"values": vals})
    elif check_type == "range":
        check_params = _json.dumps({"min": float(range_min), "max": float(range_max)})
    elif check_type == "pattern":
        check_params = _json.dumps({"pattern": pattern.strip()})

    try:
        add_user_rule({
            "rule_id":      rule_id.strip(),
            "dimension":    dim,
            "category":     (cat or "").strip() or dim.capitalize(),
            "rule_name":    name.strip(),
            "tables":       table,
            "fields":       (field or "").strip(),
            "check_type":   check_type,
            "check_params": check_params,
        })
    except Exception as exc:
        return (
            html.Span(f"Could not save: {exc}", style={"color": C_RED}),
            version,
        )

    feedback = html.Span([
        html.Span("✓ ", style={"color": C_GREEN, "fontWeight": "900"}),
        html.Span(f"{rule_id.strip()} submitted for admin review. "
                  "Once approved it will run on the next pipeline.",
                  style={"color": C_GREEN}),
    ])
    return feedback, (version or 0) + 1


@app.callback(
    Output("dl-preview-lb",    "data"),
    Output("dl-preview-modal", "style"),
    Input({"type": "inst-dl-btn", "index": ALL}, "n_clicks"),
    Input("dl-modal-cancel", "n_clicks"),
    prevent_initial_call=True,
)
def _on_inst_dl_btn(n_clicks, _cancel):
    tid = ctx.triggered_id
    # close modal
    if tid == "dl-modal-cancel":
        return None, {"display": "none"}
    # open modal
    if not isinstance(tid, dict) or "index" not in tid:
        raise dash.exceptions.PreventUpdate
    if not any(n for n in (n_clicks or []) if n):
        raise dash.exceptions.PreventUpdate
    return tid["index"], {"display": "block"}


@app.callback(
    Output("dl-modal-title",    "children"),
    Output("dl-modal-subtitle", "children"),
    Output("dl-modal-table",    "children"),
    Input("dl-preview-lb", "data"),
    prevent_initial_call=True,
)
def _populate_dl_modal(le_book):
    if not le_book:
        raise dash.exceptions.PreventUpdate

    from dq_issue_tracker import get_issues

    issues = get_issues(le_book=le_book)

    # institution name from categories JSON
    try:
        cats = json.loads(CATEGORIES_FILE.read_text())
        info = cats.get(str(le_book), cats.get(le_book, {}))
        inst_name = (info.get("name") or str(le_book)).title()
    except Exception:
        inst_name = str(le_book)

    title    = f"{inst_name}  ({le_book})"
    n_issues = len(issues)
    tables   = len({i.get("table_name", "") for i in issues})
    subtitle = (
        f"{n_issues} issue{'s' if n_issues != 1 else ''} across "
        f"{tables} table{'s' if tables != 1 else ''} will be included in the report"
        if issues else "No open issues recorded — report will contain schema data only."
    )

    # ── table header ─────────────────────────────────────────────────────────
    _TH = {
        "padding": "8px 14px", "fontSize": "11px", "fontWeight": "700",
        "color": "#FFFFFF", "textTransform": "uppercase",
        "letterSpacing": "0.05em", "whiteSpace": "nowrap",
        "background": BRAND,
    }
    header = html.Div([
        html.Span("Table",        style={**_TH, "flex": "1.4"}),
        html.Span("Rule",         style={**_TH, "flex": "0.9"}),
        html.Span("Issue",        style={**_TH, "flex": "3"}),
        html.Span("Dimension",    style={**_TH, "flex": "1"}),
        html.Span("Failing Rows", style={**_TH, "flex": "0.8", "textAlign": "right"}),
        html.Span("Status",       style={**_TH, "flex": "0.7", "textAlign": "center"}),
    ], style={"display": "flex", "position": "sticky", "top": "0", "zIndex": "1"})

    # ── urgency colour map ────────────────────────────────────────────────────
    _URGENCY_CLR = {
        "overdue":  "#7C3D1E", "critical": "#A0784A",
        "urgent":   "#B8860B", "attention": "#68686f",
        "new":      "#68686f",
    }
    _STATUS_BG = {
        "open": "rgba(117,57,24,.10)", "penalized": "rgba(124,61,30,.15)",
        "resolved": "rgba(40,160,80,.10)",
    }

    rows = []
    for i, iss in enumerate(issues):
        bg  = "rgba(244,246,249,0.7)" if i % 2 == 0 else "#FFFFFF"
        st  = (iss.get("status") or "open").lower()
        urg = (iss.get("urgency_band") or "new").lower()
        _TD = {
            "padding": "8px 14px", "fontSize": "12px",
            "color": TEXT, "display": "flex", "alignItems": "center",
        }
        rows.append(html.Div([
            html.Span(iss.get("table_name", "—"),
                      style={**_TD, "flex": "1.4", "fontFamily": "monospace",
                             "fontSize": "11px"}),
            html.Span(iss.get("rule_id", "—"),
                      style={**_TD, "flex": "0.9", "fontFamily": "monospace",
                             "fontSize": "11px", "color": BRAND,
                             "fontWeight": "700"}),
            html.Span(iss.get("rule_name") or iss.get("dimension", "—"),
                      style={**_TD, "flex": "3", "color": MUTED}),
            html.Span((iss.get("dimension") or "").title(),
                      style={**_TD, "flex": "1", "fontSize": "11px"}),
            html.Span(f"{iss.get('failing_rows', '—'):,}" if isinstance(
                          iss.get('failing_rows'), int) else "—",
                      style={**_TD, "flex": "0.8", "textAlign": "right",
                             "justifyContent": "flex-end", "fontWeight": "700",
                             "color": _URGENCY_CLR.get(urg, TEXT)}),
            html.Div(st.title(), style={
                **_TD, "flex": "0.7", "justifyContent": "center",
                "fontSize": "10px", "fontWeight": "700",
                "borderRadius": "4px",
                "background": _STATUS_BG.get(st, "transparent"),
                "color": _URGENCY_CLR.get(urg, MUTED),
            }),
        ], style={
            "display": "flex", "background": bg,
            "borderBottom": f"1px solid {DIVIDER}",
        }))

    if not rows:
        rows = [html.Div("No issues on record for this institution.",
                         style={"padding": "18px 14px", "fontSize": "13px",
                                "color": MUTED, "textAlign": "center"})]

    table = html.Div([header] + rows)
    return title, subtitle, table


@app.callback(
    Output("inst-download", "data"),
    Output("dl-nav", "href", allow_duplicate=True),
    Input("dl-modal-confirm", "n_clicks"),
    State("dl-preview-lb", "data"),
    prevent_initial_call=True,
)
def _on_inst_download_confirm(n_clicks, le_book):
    """
    Download the latest issue report for this institution.
    Preference order:
      1. issue_reports/{le_book}_{YYYY-MM}.zip  (monthly detection ZIP, via /download route)
      2. reports/ CSV bundle (legacy pipeline CSVs)
      3. reports/ XLSX fallback
    """
    if not n_clicks or not le_book:
        raise dash.exceptions.PreventUpdate

    # ── 1. monthly detection issue ZIP — streamed via Flask route (any size) ──
    issue_reports_dir = _DIR / "issue_reports"
    if issue_reports_dir.exists() and sorted(issue_reports_dir.glob(f"{le_book}_*.zip")):
        return dash.no_update, f"/download/issue-report/{le_book}?t={n_clicks}"

    # ── 2. legacy pipeline CSVs ───────────────────────────────────────────────
    if REPORTS_DIR.exists():
        import zipfile, io as _io
        csv_files = sorted(REPORTS_DIR.glob(f"*_{le_book}_*.csv"), reverse=True)
        if csv_files:
            latest_month = None
            for f in csv_files:
                parts = f.stem.rsplit("_", 2)
                if len(parts) == 3 and len(parts[2]) == 7:
                    m = parts[2]
                    if latest_month is None or m > latest_month:
                        latest_month = m
            month_files = [f for f in csv_files
                           if f.stem.endswith(f"_{latest_month}")]
            if len(month_files) == 1:
                return dcc.send_file(str(month_files[0])), dash.no_update
            if len(month_files) > 1:
                buf = _io.BytesIO()
                with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
                    for f in month_files:
                        zf.write(f, arcname=f.name)
                buf.seek(0)
                return (dcc.send_bytes(buf.read(),
                                       filename=f"dq_report_{le_book}_{latest_month}.zip"),
                        dash.no_update)

        # ── 3. XLSX fallback ──────────────────────────────────────────────────
        matches = sorted(REPORTS_DIR.glob(f"{le_book}_*.xlsx"), reverse=True)
        if matches:
            return dcc.send_file(str(matches[0])), dash.no_update

    raise dash.exceptions.PreventUpdate


@app.callback(
    Output("rules-download", "data"),
    Input("rules-download-btn", "n_clicks"),
    prevent_initial_call=True,
)
def _on_rules_download(n_clicks):
    if not n_clicks:
        raise dash.exceptions.PreventUpdate
    rules = get_all_rules()
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=["rule_id", "dimension", "category", "rule_name", "tables", "fields"])
    writer.writeheader()
    writer.writerows(rules)
    return dict(content=buf.getvalue(), filename="dq_validation_rules.csv")


# ── admin review: approve / delete draft rules ────────────────────────────────

@app.callback(
    Output("rules-version", "data", allow_duplicate=True),
    Input({"type": "approve-btn", "index": ALL}, "n_clicks"),
    State("rules-version", "data"),
    prevent_initial_call=True,
)
def _approve_draft(clicks, version):
    if not any(c for c in (clicks or []) if c):
        raise dash.exceptions.PreventUpdate
    tid = ctx.triggered_id
    if not isinstance(tid, dict) or tid.get("type") != "approve-btn":
        raise dash.exceptions.PreventUpdate
    if not (ctx.triggered[0]["value"] or 0) > 0:
        raise dash.exceptions.PreventUpdate
    approve_draft_rule(tid["index"])
    return (version or 0) + 1


@app.callback(
    Output("rules-version", "data", allow_duplicate=True),
    Input({"type": "delete-draft-btn", "index": ALL}, "n_clicks"),
    State("rules-version", "data"),
    prevent_initial_call=True,
)
def _delete_draft(clicks, version):
    if not any(c for c in (clicks or []) if c):
        raise dash.exceptions.PreventUpdate
    tid = ctx.triggered_id
    if not isinstance(tid, dict) or tid.get("type") != "delete-draft-btn":
        raise dash.exceptions.PreventUpdate
    if not (ctx.triggered[0]["value"] or 0) > 0:
        raise dash.exceptions.PreventUpdate
    delete_draft_rule(tid["index"])
    return (version or 0) + 1


# ── complex rule form toggle + submit ─────────────────────────────────────────

@app.callback(
    Output("complex-form-panel", "style"),
    Input("complex-form-toggle-btn", "n_clicks"),
    State("complex-form-panel", "style"),
    prevent_initial_call=True,
)
def _toggle_complex_form(n_clicks, current_style):
    style = dict(current_style or {})
    style["display"] = "none" if style.get("display") != "none" else "block"
    return style


@app.callback(
    Output("cx-rule-feedback", "children"),
    Output("rules-version", "data", allow_duplicate=True),
    Input("cx-rule-submit",     "n_clicks"),
    State("cx-rule-id",         "value"),
    State("cx-rule-dim",        "value"),
    State("cx-rule-name",       "value"),
    State("cx-rule-tables",     "value"),
    State("cx-rule-fields",     "value"),
    State("cx-rule-logic",      "value"),
    State("cx-rule-condition",  "value"),
    State("rules-version",      "data"),
    prevent_initial_call=True,
)
def _submit_complex_rule(n_clicks, rule_id, dim, name,
                         tables, fields, logic, condition, version):
    if not n_clicks:
        raise dash.exceptions.PreventUpdate

    errors = []
    if not (rule_id or "").strip():
        errors.append("Rule ID is required.")
    if not dim:
        errors.append("Dimension is required.")
    if not (name or "").strip():
        errors.append("Rule Name is required.")
    if not (tables or "").strip():
        errors.append("Table(s) is required.")
    if not (logic or "").strip():
        errors.append("Business Logic description is required.")

    if errors:
        return html.Span("  ".join(errors), style={"color": C_RED}), version

    import json as _json
    check_type   = "sql_condition" if (condition or "").strip() else "description"
    check_params = None
    if check_type == "sql_condition":
        check_params = _json.dumps({"condition": condition.strip(), "logic": (logic or "").strip()})
    else:
        check_params = _json.dumps({"logic": (logic or "").strip()})

    try:
        add_user_rule({
            "rule_id":      rule_id.strip(),
            "dimension":    dim,
            "category":     "Complex Rule",
            "rule_name":    name.strip(),
            "tables":       (tables or "").strip(),
            "fields":       (fields or "").strip(),
            "check_type":   check_type,
            "check_params": check_params,
        })
    except Exception as exc:
        return html.Span(f"Could not save: {exc}", style={"color": C_RED}), version

    return html.Span([
        html.Span("✓ ", style={"color": C_GREEN, "fontWeight": "900"}),
        html.Span(
            f"{rule_id.strip()} submitted for admin review"
            + (" (will auto-evaluate via SQL condition)."
               if check_type == "sql_condition"
               else " (manual evaluation — no SQL condition provided)."),
            style={"color": C_GREEN},
        ),
    ]), (version or 0) + 1


# ── login tab switching ───────────────────────────────────────────────────────

@app.callback(
    Output("auth-store",  "data", allow_duplicate=True),
    Output("login-type",  "data"),
    Input("login-tab-bnr",  "n_clicks"),
    Input("login-tab-inst", "n_clicks"),
    prevent_initial_call=True,
)
def _switch_login_tab(n_bnr, n_inst):
    tid = ctx.triggered_id
    lt  = "inst" if tid == "login-tab-inst" else "bnr"
    # Return a "tab-switch" marker so _render_page re-draws the login page
    return {"tab_switch": lt}, lt


# ── login callback ───────────────────────────────────────────────────────────

@app.callback(
    Output("auth-store", "data", allow_duplicate=True),
    Input("login-btn",      "n_clicks"),
    Input("login-password", "n_submit"),
    State("login-email",    "value"),
    State("login-password", "value"),
    State("login-type",     "data"),
    prevent_initial_call=True,
)
def _do_login(n_clicks, n_submit, email, password, login_type):
    if not (n_clicks or n_submit):
        raise dash.exceptions.PreventUpdate

    email      = (email    or "").strip().lower()
    password   = (password or "").strip()
    login_type = login_type or "bnr"

    # Enforce @bnr.rw for the BNR tab before even hitting the DB
    if login_type == "bnr" and not dq_auth.is_valid_bnr_email(email):
        return {"error": "BNR Staff login requires a @bnr.rw email address.", "tab": login_type}

    # Reject inst_user credentials on the BNR tab
    user = dq_auth.verify_credentials(email, password)
    if not user:
        return {"error": "Incorrect email or password.", "tab": login_type}

    if login_type == "bnr" and user["role"] == "inst_user":
        return {"error": "Institution accounts must use the Institution login.", "tab": login_type}

    if login_type == "inst" and user["role"] != "inst_user":
        return {"error": "BNR staff accounts must use the BNR Staff login.", "tab": login_type}

    flask_session["user_email"] = user["email"]
    flask_session["user_name"]  = user["name"]
    flask_session["user_role"]  = user["role"]
    flask_session.permanent     = True

    le_books = dq_auth.get_user_institutions(user["user_id"]) if user["role"] == "inst_user" else []

    return {
        "email":    user["email"],
        "name":     user["name"],
        "role":     user["role"],
        "user_id":  user["user_id"],
        "le_books": le_books,
    }


# ── logout callback ───────────────────────────────────────────────────────────

@app.callback(
    Output("auth-store", "data", allow_duplicate=True),
    Input("logout-btn",  "n_clicks"),
    prevent_initial_call=True,
)
def _do_logout(n):
    if not n:
        raise dash.exceptions.PreventUpdate
    flask_session.clear()
    return {}   # empty auth-store → triggers login page re-render


# ── user info header ──────────────────────────────────────────────────────────

@app.callback(
    Output("user-info-header", "children"),
    Input("auth-store", "data"),
    prevent_initial_call=True,
)
def _update_user_header(auth_data):
    auth  = auth_data or {}
    email = auth.get("email", "")
    name  = auth.get("name",  "")
    if not email:
        return html.Div()
    display = name.split()[0] if name else email.split("@")[0]
    return html.Div([
        html.Span(display, style={
            "fontSize": "12px", "color": "rgba(255,255,255,0.85)",
            "marginRight": "10px",
        }),
        html.Div(
            "Sign out",
            id="logout-btn",
            n_clicks=0,
            style={
                "fontSize": "11px", "fontWeight": "700",
                "color": "rgba(255,255,255,0.60)",
                "cursor": "pointer", "userSelect": "none",
                "border": "1px solid rgba(255,255,255,0.25)",
                "borderRadius": "4px", "padding": "4px 10px",
            },
        ),
    ], style={"display": "flex", "alignItems": "center"})


# ── notify button callback ────────────────────────────────────────────────────

@app.callback(
    Output("issue-list",        "children"),
    Output("alerts-summary-bar","children"),
    Input("alerts-cat-filter",  "value"),
    Input("notif-poll",         "n_intervals"),
)
def _refresh_issue_list(cat_filter, _poll):
    from dq_issue_tracker import get_issues
    from datetime import date as _date

    cat_filter = cat_filter or ""

    # Build le_book → category_type map for filtering
    try:
        _cats = json.loads(CATEGORIES_FILE.read_text())
    except Exception:
        _cats = {}

    def _matches(lb: str) -> bool:
        if not cat_filter:
            return True
        ct = (_cats.get(str(lb), {}).get("category_type") or "")
        if cat_filter == "SACCO":
            return ct in ("SACCO", "OSACCO")
        return ct == cat_filter

    all_resolved_raw = get_issues(status="resolved")
    all_open_raw     = get_issues(status="open")

    all_resolved = [i for i in all_resolved_raw if _matches(i["le_book"])]
    all_open     = [i for i in all_open_raw     if _matches(i["le_book"])]

    # ── summary chips ─────────────────────────────────────────────────────────
    today         = _date.today()
    this_month    = today.strftime("%Y-%m")
    month_res     = sum(1 for i in all_resolved
                        if (i.get("resolved_at") or "").startswith(this_month))
    overdue_count = sum(1 for i in all_open if i.get("urgency_band") == "overdue")
    on_time       = sum(1 for i in all_resolved
                        if i.get("sla_deadline") and i.get("resolved_at")
                        and i["resolved_at"] <= i["sla_deadline"])
    late          = len(all_resolved) - on_time

    fix_days = []
    for i in all_resolved:
        try:
            fix_days.append(
                (_date.fromisoformat(i["resolved_at"]) -
                 _date.fromisoformat(i["detected_at"])).days
            )
        except Exception:
            pass
    avg_fix = f"{sum(fix_days) // len(fix_days)}d" if fix_days else "—"

    def _chip(value, label, color=None, border_color=None):
        bc = border_color or DIVIDER
        return html.Div([
            html.Span(str(value), style={"fontWeight": "900", "fontSize": "22px",
                                         "color": color or TEXT}),
            html.Span(label,      style={"fontSize": "11px", "color": MUTED,
                                         "marginTop": "2px"}),
        ], style={
            "display": "flex", "flexDirection": "column", "alignItems": "center",
            "background": CARD, "borderRadius": "8px", "padding": "12px 24px",
            "border": f"1px solid {bc}", "minWidth": "110px",
        })

    summary_bar = html.Div([
        _chip(len(all_open),     "Open Issues",         C_AMBER),
        _chip(overdue_count,     "Overdue",             _URGENCY_COLORS["overdue"],
              border_color=_URGENCY_COLORS["overdue"]),
        _chip(month_res,         "Resolved This Month", C_GREEN),
        _chip(len(all_resolved), "Total Resolved"),
        _chip(on_time,           "Resolved On Time",    C_GREEN),
        _chip(late,              "Resolved Late",       C_RED if late else MUTED),
        _chip(avg_fix,           "Avg Days to Fix"),
    ], style={"display": "flex", "gap": "10px", "flexWrap": "wrap"})

    # ── issue list ─────────────────────────────────────────────────────────────
    all_resolved_raw.sort(key=lambda x: x.get("resolved_at") or "", reverse=True)
    issue_list = _build_resolved_by_institution(all_resolved_raw, cat_filter)

    return issue_list, summary_bar


@app.callback(
    Output({"type": "alert-inst-collapse", "index": MATCH}, "style"),
    Output({"type": "alert-inst-toggle",   "index": MATCH}, "children"),
    Input({"type": "alert-inst-toggle",    "index": MATCH}, "n_clicks"),
    State({"type": "alert-inst-collapse",  "index": MATCH}, "style"),
    prevent_initial_call=True,
)
def _toggle_alert_inst(n_clicks, current_style):
    hidden    = (current_style or {}).get("display") == "none"
    new_style = {"display": "block"} if hidden else {"display": "none"}
    return new_style, "▼" if hidden else "▶"


@app.callback(
    Output({"type": "res-inst-collapse", "index": MATCH}, "style"),
    Output({"type": "res-inst-toggle",   "index": MATCH}, "children"),
    Input({"type": "res-inst-toggle",    "index": MATCH}, "n_clicks"),
    State({"type": "res-inst-collapse",  "index": MATCH}, "style"),
    prevent_initial_call=True,
)
def _toggle_res_inst(n_clicks, current_style):
    hidden    = (current_style or {}).get("display") == "none"
    new_style = {"display": "block"} if hidden else {"display": "none"}
    return new_style, "▼" if hidden else "▶"


@app.callback(
    Output("resolved-dl-lb", "data"),
    Input({"type": "res-dl-btn", "index": ALL}, "n_clicks"),
    prevent_initial_call=True,
)
def _resolved_dl_select(clicks):
    if not any(c for c in (clicks or []) if c):
        raise dash.exceptions.PreventUpdate
    tid = ctx.triggered_id
    if not isinstance(tid, dict) or tid.get("type") != "res-dl-btn":
        raise dash.exceptions.PreventUpdate
    if not ctx.triggered[0]["value"]:
        raise dash.exceptions.PreventUpdate
    return tid["index"]


@app.callback(
    Output("resolved-inst-download", "data"),
    Input("resolved-dl-lb", "data"),
    prevent_initial_call=True,
)
def _resolved_dl_generate(le_book):
    if not le_book:
        raise dash.exceptions.PreventUpdate

    import zipfile, io as _io
    import pandas as pd
    from datetime import date as _date
    from dq_issue_tracker import get_issues
    from dq_rules import COMP_RULE_META

    resolved = [i for i in get_issues(status="resolved") if i["le_book"] == le_book]
    if not resolved:
        raise dash.exceptions.PreventUpdate

    issue_reports_dir = _DIR / "issue_reports"
    comp_rule_ids     = set(COMP_RULE_META.keys())
    today_str         = _date.today().strftime("%Y-%m")
    out_buf           = _io.BytesIO()
    found_any         = False

    # group resolved issues by (detected month, table)
    by_month_table: dict[tuple, list] = {}
    for iss in resolved:
        month = (iss.get("detected_at") or "")[:7]
        if month:
            by_month_table.setdefault((month, iss["table_name"]), []).append(iss)

    with zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED) as out_zf:
        # process each (month, table) group using the stored detection ZIP
        for (month, table), issues in sorted(by_month_table.items()):
            zip_path = issue_reports_dir / f"{le_book}_{month}.zip"
            if not zip_path.exists():
                continue

            try:
                with zipfile.ZipFile(zip_path) as src_zf:
                    if f"{table}.xlsx" not in src_zf.namelist():
                        continue
                    xls = pd.ExcelFile(_io.BytesIO(src_zf.read(f"{table}.xlsx")))
            except Exception:
                continue

            # Reports are now one Excel sheet per issue (sheet name = issue type).
            # Map each sheet to the resolved issue it belongs to: completeness issues
            # cover the "Missing *" sheets; dimension issues match the sheet whose
            # name starts with their rule_id ("VAL-001 ...", "ACC-002 ...", etc.).
            comp_present = any(i["rule_id"] in comp_rule_ids for i in issues)
            dim_issues   = [i for i in issues if i["rule_id"] not in comp_rule_ids]

            sheet_issue: dict[str, dict] = {}
            for sheet in xls.sheet_names:
                name = str(sheet)
                if comp_present and name.startswith("Missing "):
                    sheet_issue[sheet] = next(
                        (i for i in issues if i["rule_id"] in comp_rule_ids), {})
                else:
                    iss = next((i for i in dim_issues if name.startswith(i["rule_id"])), None)
                    if iss is not None:
                        sheet_issue[sheet] = iss
            if not sheet_issue:
                continue

            frames = []
            for sheet, iss in sheet_issue.items():
                try:
                    sdf = xls.parse(sheet)
                except Exception:
                    continue
                if sdf.empty:
                    continue
                sdf["issue_type"]   = sheet
                sdf["resolved_at"]  = iss.get("resolved_at", "")
                sdf["sla_deadline"] = iss.get("sla_deadline", "")
                frames.append(sdf)
            if not frames:
                continue

            filtered = pd.concat(frames, ignore_index=True)
            filtered["on_time"] = filtered.apply(
                lambda r: (
                    "On Time" if r["resolved_at"] and r["sla_deadline"]
                    and r["resolved_at"] <= r["sla_deadline"] else
                    "Late"    if r["resolved_at"] and r["sla_deadline"] else ""
                ), axis=1,
            )

            fname = f"{table}_{le_book}_{month}_resolved.csv"
            out_zf.writestr(fname, filtered.to_csv(index=False, encoding="utf-8-sig"))
            found_any = True

    if not found_any:
        raise dash.exceptions.PreventUpdate

    out_buf.seek(0)
    return dcc.send_bytes(out_buf.read(),
                          filename=f"dq_resolved_{le_book}_{today_str}.zip")


@app.callback(
    Output("notify-status", "data"),
    Input({"type": "notify-btn", "index": ALL}, "n_clicks"),
    State("notify-status", "data"),
    prevent_initial_call=True,
)
def _on_notify(clicks, current):
    if not any(c for c in (clicks or []) if c):
        raise dash.exceptions.PreventUpdate
    tid = ctx.triggered_id
    if not isinstance(tid, dict) or tid.get("type") != "notify-btn":
        raise dash.exceptions.PreventUpdate
    if not ctx.triggered[0]["value"]:
        raise dash.exceptions.PreventUpdate

    lb = tid["index"]
    try:
        from dq_issue_tracker import (
            get_open_issues, send_notification, get_contact,
        )
        issues = get_open_issues(lb)
        if not issues:
            result = "no_issues"
        else:
            inst_name = (issues[0].get("institution_name") or lb).title()
            sent = send_notification(lb, inst_name, issues, force=True)
            result = "sent" if sent else "no_email"
    except Exception as exc:
        log.error("Notify callback error: %s", exc)
        result = "error"

    updated = dict(current or {})
    updated[lb] = result
    return updated


@app.callback(
    Output("notify-feedback", "children"),
    Input("notify-status", "data"),
    prevent_initial_call=True,
)
def _show_notify_feedback(status_data):
    if not status_data:
        raise dash.exceptions.PreventUpdate

    msgs = []
    for lb, result in status_data.items():
        if result == "sent":
            msgs.append(html.Span(
                f"✓ Reminder sent for institution {lb}.",
                style={"color": C_GREEN, "fontSize": "12px", "marginRight": "12px"},
            ))
        elif result == "no_email":
            msgs.append(html.Span(
                f"⚠ No contact email configured for {lb}. "
                "Set one in dq_institution_contacts (SQLite) or configure SMTP.",
                style={"color": C_AMBER, "fontSize": "12px", "marginRight": "12px"},
            ))
        elif result == "no_issues":
            msgs.append(html.Span(
                f"No open issues found for {lb}.",
                style={"color": MUTED, "fontSize": "12px", "marginRight": "12px"},
            ))
        elif result == "error":
            msgs.append(html.Span(
                f"Error sending notification for {lb}. Check logs.",
                style={"color": C_RED, "fontSize": "12px", "marginRight": "12px"},
            ))
    return html.Div(msgs) if msgs else dash.no_update


# ── remediation callbacks ──────────────────────────────────────────────────────

@app.callback(
    Output("cr-form-panel", "style"),
    Input("cr-form-toggle-btn", "n_clicks"),
    State("cr-form-panel", "style"),
    prevent_initial_call=True,
)
def _toggle_cr_form(n_clicks, current_style):
    style = dict(current_style or {})
    style["display"] = "none" if style.get("display") != "none" else "block"
    return style


@app.callback(
    Output("cr-issue-checklist", "options"),
    Output("cr-issue-checklist", "value"),
    Output("cr-issue-hint",      "children"),
    Output("cr-table-dl-area",   "children"),
    Input("cr-inst-filter", "value"),
    prevent_initial_call=True,
)
def _update_issue_checklist(le_book):
    """Populate the table selector when the specialist picks an institution."""
    _empty = ([], [], "Select an institution above to see its tables with open issues.", html.Div())
    if not le_book:
        return _empty
    from dq_issue_tracker import get_open_issues
    issues = get_open_issues(le_book)
    if not issues:
        return [], [], "No open issues found for this institution.", html.Div()

    _BO = {"new": 0, "attention": 1, "urgent": 2, "critical": 3, "overdue": 4}
    by_table: dict[str, list] = {}
    for iss in issues:
        by_table.setdefault(iss["table_name"], []).append(iss)

    options = []
    for tbl in sorted(by_table.keys()):
        tbl_issues = by_table[tbl]
        tbl_label  = TABLE_NAMES_PRETTY.get(tbl, tbl.replace("_", " ").title())
        n_issues   = len(tbl_issues)
        worst_band = max(
            (iss.get("urgency_band") or "new" for iss in tbl_issues),
            key=lambda b: _BO.get(b, 0),
        )
        options.append({"label": f"{tbl_label}  —  {n_issues} issue(s)  —  {worst_band.upper()}", "value": tbl})

    hint = f"{len(by_table)} table(s) with open issues — select those to include in this Change Request."

    dl_buttons = html.Div([
        html.Div(
            "⬇ " + TABLE_NAMES_PRETTY.get(tbl, tbl),
            id={"type": "cr-tbl-dl-btn", "index": f"{le_book}|{tbl}"},
            n_clicks=0,
            style={"display": "inline-block", "background": BRAND, "color": CARD,
                   "fontSize": "10px", "fontWeight": "700", "padding": "3px 10px",
                   "borderRadius": "4px", "cursor": "pointer", "userSelect": "none"},
        )
        for tbl in sorted(by_table.keys())
    ], style={"display": "flex", "gap": "6px", "flexWrap": "wrap"})

    return options, [], hint, dl_buttons


@app.callback(
    Output("cr-assigned-to", "options"),
    Output("cr-assigned-to", "value"),
    Output("cr-assigned-to", "placeholder"),
    Input("cr-inst-filter",  "value"),
    prevent_initial_call=True,
)
def _populate_assigned_to(le_book):
    if not le_book:
        return [], None, "Select an institution first…"

    import json as _json
    from dq_auth import get_users_by_le_book

    users = get_users_by_le_book(le_book)
    if not users:
        # Get institution name for a helpful placeholder
        cats_path = Path(__file__).parent / "le_book_categories.json"
        try:
            cats = _json.loads(cats_path.read_text())
            inst_name = (cats.get(le_book, {}).get("name") or le_book).title()
        except Exception:
            inst_name = le_book
        return [], None, f"No users registered under {inst_name}"

    options = [
        {"label": f"{u['name']} ({u['email']})", "value": u["email"]}
        for u in users
    ]
    return options, None, "Select a user…"


@app.callback(
    Output("cr-feedback",  "children"),
    Output("cr-version",   "data", allow_duplicate=True),
    Input("cr-create-btn", "n_clicks"),
    State("cr-issue-checklist", "value"),
    State("cr-title",           "value"),
    State("cr-description",     "value"),
    State("cr-assigned-to",     "value"),
    State("cr-target-date",     "value"),
    State("cr-inst-filter",     "value"),
    State("cr-version",         "data"),
    State("auth-store",         "data"),
    prevent_initial_call=True,
)
def _create_cr(n_clicks, tables, title, description,
               assigned_to, target_date, le_book, version, auth_data):
    if not n_clicks:
        raise dash.exceptions.PreventUpdate

    _err = lambda msg: (html.Span(msg, style={"color": C_RED}), version)

    if not le_book:
        return _err("Select an institution first.")
    if not tables:
        return _err("Select at least one table to include in this change request.")
    if not (title or "").strip():
        return _err("A title is required.")

    import dq_change_request as cr_mod
    from dq_issue_tracker import get_open_issues

    issues     = get_open_issues(le_book)
    tables_set = set(tables)
    inst_name  = ""
    total_fail = 0
    dims: set[str] = set()
    for iss in issues:
        if iss["table_name"] not in tables_set:
            continue
        inst_name  = (iss.get("institution_name") or le_book).title()
        total_fail += int(iss.get("failing_rows") or 0)
        if iss.get("dimension"):
            dims.add(iss["dimension"])

    dimension  = next(iter(dims)) if len(dims) == 1 else ("Multiple" if dims else "")
    created_by = (auth_data or {}).get("email", "")

    try:
        cr_id = cr_mod.create_cr(
            tables           = list(tables),
            le_book          = le_book,
            institution_name = inst_name or le_book,
            title            = title.strip(),
            description      = (description or "").strip(),
            assigned_to      = (assigned_to  or "").strip(),
            created_by       = created_by,
            target_date      = (target_date  or "").strip(),
            dimension        = dimension,
            failing_rows     = total_fail,
        )
    except Exception as exc:
        return _err(f"Could not save change request: {exc}")

    return (
        html.Span([
            html.Span("✓ ", style={"color": C_GREEN, "fontWeight": "900"}),
            html.Span(
                f"{cr_id} created for {inst_name} "
                f"({len(tables)} table(s), {total_fail:,} failing rows).",
                style={"color": C_GREEN},
            ),
        ]),
        (version or 0) + 1,
    )


@app.callback(
    Output("cr-version",        "data", allow_duplicate=True),
    Output("cr-action-feedback", "children"),
    Input({"type": "cr-action-btn", "index": ALL}, "n_clicks"),
    State("cr-review-notes", "value"),
    State("auth-store",       "data"),
    State("cr-version",       "data"),
    prevent_initial_call=True,
)
def _cr_action(clicks, review_notes, auth_data, version):
    if (auth_data or {}).get("role") == "inst_user":
        raise dash.exceptions.PreventUpdate
    if not any(c for c in (clicks or []) if c):
        raise dash.exceptions.PreventUpdate
    tid = ctx.triggered_id
    if not isinstance(tid, dict) or tid.get("type") != "cr-action-btn":
        raise dash.exceptions.PreventUpdate
    if not (ctx.triggered[0]["value"] or 0) > 0:
        raise dash.exceptions.PreventUpdate

    import dq_change_request as cr_mod

    raw = tid["index"]          # e.g. "CR-20260519-0001|approved"
    if "|" not in raw:
        raise dash.exceptions.PreventUpdate
    cr_id, new_status = raw.split("|", 1)

    actor = (auth_data or {}).get("email", "system")
    notes = (review_notes or "").strip()

    ok, msg = cr_mod.update_status(cr_id, new_status, actor=actor, notes=notes)

    if ok:
        label = cr_mod.STATUS_LABELS.get(new_status, new_status)
        feedback = html.Span([
            html.Span("✓ ", style={"color": C_GREEN, "fontWeight": "900"}),
            html.Span(f"{cr_id} updated to \"{label}\".", style={"color": C_GREEN}),
        ])
        return (version or 0) + 1, feedback
    else:
        return version, html.Span(msg, style={"color": C_RED})


@app.callback(
    Output("cr-list-container", "children"),
    Input("cr-version",      "data"),
    Input("cr-status-filter", "value"),
    Input("notif-poll",       "n_intervals"),
    State("auth-store",       "data"),
)
def _refresh_cr_list(version, status_filter, _poll, auth_data):
    import dq_change_request as cr_mod
    role = (auth_data or {}).get("role", "viewer")
    crs  = cr_mod.get_crs(status=status_filter if status_filter != "all" else None)
    return _build_cr_list(crs, role=role)


@app.callback(
    Output("cr-summary-bar", "children"),
    Input("cr-version",  "data"),
    Input("notif-poll",  "n_intervals"),
    prevent_initial_call=False,
)
def _refresh_cr_stats(version, _poll):
    import dq_change_request as cr_mod
    stats = cr_mod.get_stats()
    chips = []
    for key, lbl in cr_mod.STATUS_LABELS.items():
        n   = stats.get(key, 0)
        clr = cr_mod.STATUS_COLORS[key]
        chips.append(html.Div([
            html.Span(str(n), style={
                "fontSize": "24px", "fontWeight": "900",
                "color": clr, "lineHeight": "1",
            }),
            html.Span(lbl, style={
                "fontSize": "10px", "color": MUTED,
                "marginTop": "3px", "lineHeight": "1.2",
                "textAlign": "center",
            }),
        ], style={
            "display":       "flex",
            "flexDirection": "column",
            "alignItems":    "center",
            "background":    CARD,
            "borderRadius":  "8px",
            "padding":       "12px 18px",
            "border":        f"2px solid {clr}",
            "minWidth":      "90px",
        }))
    return chips


# ── institution portal callbacks ──────────────────────────────────────────────

@app.callback(
    Output("active-page", "data", allow_duplicate=True),
    Input({"type": "inst-nav-tab", "index": ALL}, "n_clicks"),
    prevent_initial_call=True,
)
def _inst_nav_click(clicks):
    if not any(c for c in (clicks or []) if c):
        raise dash.exceptions.PreventUpdate
    tid = ctx.triggered_id
    if isinstance(tid, dict):
        return tid["index"]
    raise dash.exceptions.PreventUpdate


@app.callback(
    Output("inst-issue-list", "children"),
    Input("inst-issue-filter",  "value"),
    Input("inst-table-filter",  "value"),
    Input("notif-poll",         "n_intervals"),
    State("auth-store", "data"),
    prevent_initial_call=False,
)
def _inst_issue_list(status, table_filter, _poll, auth_data):
    from dq_issue_tracker import get_issues_by_table, get_issues
    from datetime import date as _date

    le_books = set(str(lb) for lb in (auth_data or {}).get("le_books", []))
    le_book  = next(iter(le_books), None)
    status   = status or "open"

    if status != "open":
        issues = [i for i in get_issues(status=status) if i["le_book"] in le_books]
        label  = {"penalized": "Delayed", "resolved": "Resolved"}.get(status, status.title())
        if not issues:
            return html.Div(f"No {label.lower()} issues.",
                            style={"color": MUTED, "padding": "20px", "fontSize": "12px"})
        return _build_resolved_rows(issues)

    # ── open: one row per affected table ──────────────────────────────────────
    all_by_table = get_issues_by_table(status="open")

    # apply table filter
    selected_tables = set(table_filter) if table_filter else set()

    # aggregate per table for this institution
    _BO = ["new", "attention", "urgent", "critical", "overdue"]
    table_summary: list[dict] = []
    for table, rules in all_by_table.items():
        if selected_tables and table not in selected_tables:
            continue
        total_rows  = 0
        n_rules     = 0
        worst       = "new"
        earliest_dl = None
        for rule in rules:
            my_insts = [i for i in rule["institutions"] if i["le_book"] in le_books]
            if not my_insts:
                continue
            total_rows += sum(i["failing_rows"] for i in my_insts)
            n_rules    += 1
            urg = rule["worst_urgency"]
            if _BO.index(urg) > _BO.index(worst):
                worst = urg
            for inst in my_insts:
                dl = inst.get("sla_deadline")
                if dl and (earliest_dl is None or dl < earliest_dl):
                    earliest_dl = dl
        if n_rules:
            table_summary.append({
                "table":       table,
                "label":       TABLE_NAMES_PRETTY.get(table, table.replace("_", " ").title()),
                "n_rules":     n_rules,
                "total_rows":  total_rows,
                "worst":       worst,
                "deadline":    earliest_dl or "—",
            })

    if not table_summary:
        return html.Div("No open issues.",
                        style={"color": MUTED, "padding": "20px", "fontSize": "12px"})

    # sort worst-urgency first, then most failing rows
    table_summary.sort(key=lambda r: (_BO.index(r["worst"]), -r["total_rows"]), reverse=True)

    # ── find the latest institution report file (mirrors _on_inst_download_confirm order) ──
    rpt_file  = None
    rpt_date  = "—"
    if le_book:
        _issue_dir = _DIR / "issue_reports"
        if _issue_dir.exists():
            _zips = sorted(_issue_dir.glob(f"{le_book}_*.zip"), reverse=True)
            if _zips:
                rpt_file = _zips[0]
                rpt_date = rpt_file.stem.split("_", 1)[-1]  # YYYY-MM
        if rpt_file is None and REPORTS_DIR.exists():
            _csvs = sorted(REPORTS_DIR.glob(f"*_{le_book}_*.csv"), reverse=True)
            if _csvs:
                rpt_file = _csvs[0]
                parts    = rpt_file.stem.rsplit("_", 2)
                rpt_date = parts[2] if len(parts) == 3 and len(parts[2]) == 7 else "—"
            if rpt_file is None:
                _xlsxs = sorted(REPORTS_DIR.glob(f"{le_book}_*.xlsx"), reverse=True)
                if _xlsxs:
                    rpt_file = _xlsxs[0]
                    stem     = rpt_file.stem.rsplit("_", 1)
                    rpt_date = stem[-1] if len(stem) == 2 and len(stem[-1]) == 10 else "—"

    # ── table header ──────────────────────────────────────────────────────────
    _H = {"fontSize": "10px", "fontWeight": "900", "color": MUTED,
          "textTransform": "uppercase", "letterSpacing": "0.05em",
          "padding": "7px 14px", "whiteSpace": "nowrap"}
    hdr = html.Div([
        html.Span("Table",        style={**_H, "flex": "1"}),
        html.Span("Issues",       style={**_H, "width": "68px", "textAlign": "center"}),
        html.Span("Failing Rows", style={**_H, "width": "110px", "textAlign": "right"}),
        html.Span("Urgency",      style={**_H, "width": "90px", "textAlign": "center"}),
        html.Span("SLA Deadline", style={**_H, "width": "100px"}),
        html.Span("Report",       style={**_H, "width": "80px", "textAlign": "center"}),
    ], style={"display": "flex", "background": BG,
              "borderRadius": "8px 8px 0 0",
              "borderBottom": f"2px solid {DIVIDER}"})

    _UC = _URGENCY_COLORS
    rows = []
    for i, row in enumerate(table_summary):
        urg     = row["worst"]
        urg_col = _UC.get(urg, MUTED)
        ov      = urg == "overdue"
        bg      = "rgba(244,246,249,0.7)" if i % 2 == 0 else CARD

        # urgency chip
        urg_chip = html.Span(
            ("⚠ " if ov else "") + urg.title(),
            style={"fontSize": "10px", "fontWeight": "700", "color": urg_col,
                   "background": f"rgba({','.join(str(int(urg_col.lstrip('#')[j:j+2],16)) for j in (0,2,4))},.10)",
                   "border": f"1px solid {urg_col}",
                   "borderRadius": "4px", "padding": "2px 7px",
                   "whiteSpace": "nowrap"},
        )

        # download icon — links to the institution's report file
        if rpt_file:
            dl_icon = html.Div(
                "⬇",
                id={"type": "inst-dl-btn", "index": le_book or ""},
                n_clicks=0,
                title=f"Download full report ({rpt_date})",
                style={"fontSize": "16px", "color": BRAND,
                       "cursor": "pointer", "textAlign": "center",
                       "userSelect": "none"},
            )
        else:
            dl_icon = html.Span("—", style={"color": MUTED, "fontSize": "12px",
                                             "textAlign": "center", "display": "block"})

        rows.append(html.Div([
            html.Div([
                html.Span("●", style={"color": urg_col, "fontSize": "9px",
                                       "marginRight": "8px"}),
                html.Span(row["label"], style={"fontSize": "13px", "fontWeight": "700",
                                                "color": TEXT}),
            ], style={"flex": "1", "display": "flex", "alignItems": "center",
                      "padding": "10px 14px",
                      "borderLeft": f"3px solid {urg_col}"}),
            html.Span(
                str(row["n_rules"]),
                style={"width": "68px", "textAlign": "center",
                       "fontSize": "13px", "color": MUTED, "padding": "10px 0"},
            ),
            html.Span(
                f"{row['total_rows']:,}",
                style={"width": "110px", "textAlign": "right",
                       "fontSize": "13px", "fontWeight": "700",
                       "color": C_RED, "padding": "10px 14px"},
            ),
            html.Div(urg_chip,
                     style={"width": "90px", "padding": "10px 6px",
                            "display": "flex", "justifyContent": "center"}),
            html.Span(
                row["deadline"],
                style={"width": "100px", "fontSize": "11px",
                       "color": C_RED if ov else MUTED, "padding": "10px 14px"},
            ),
            html.Div(dl_icon,
                     style={"width": "80px", "display": "flex",
                            "justifyContent": "center", "alignItems": "center"}),
        ], style={
            "display": "flex", "alignItems": "center",
            "background": bg,
            "borderBottom": f"1px solid {DIVIDER}",
        }))

    return html.Div(
        [hdr, *rows],
        style={"background": CARD, "borderRadius": "8px",
               "border": f"1px solid {DIVIDER}"},
    )


@app.callback(
    Output("inst-csv-download", "data"),
    Input("inst-csv-dl-btn",   "n_clicks"),
    State("inst-issue-filter", "value"),
    State("inst-table-filter", "value"),
    State("auth-store",        "data"),
    prevent_initial_call=True,
)
def _inst_download_csv(n_clicks, status_filter, table_filter, auth_data):
    """Generate a filtered CSV of open issues for the institution."""
    if not n_clicks:
        raise dash.exceptions.PreventUpdate

    from dq_issue_tracker import get_issues
    from datetime import date as _date
    import io as _io
    import csv as _csv

    le_books = set(str(lb) for lb in (auth_data or {}).get("le_books", []))
    inst_name = (auth_data or {}).get("institution_name", "institution")
    status    = status_filter or "open"
    today     = _date.today()

    issues = [i for i in get_issues(status=status) if i["le_book"] in le_books]

    # apply table filter
    if table_filter:
        issues = [i for i in issues if i["table_name"] in set(table_filter)]

    _PRETTY = {
        "accounts":               "Accounts",
        "contract_loans":         "Contract Loans",
        "contract_schedules":     "Contract Schedules",
        "contracts_disburse":     "Contracts Disburse",
        "contracts_expanded":     "Contracts",
        "customers_expanded":     "Customers",
        "loan_applications_2":    "Loan Applications",
        "prev_loan_applications": "Prev Loan Apps",
    }

    buf = _io.StringIO()
    w   = _csv.writer(buf)
    w.writerow([
        "Table", "Rule ID", "Issue", "Dimension",
        "Failing Rows", "Prev Count", "Detected", "SLA Deadline",
        "Days Remaining", "Urgency", "Status", "Recurrence",
    ])

    for iss in sorted(issues, key=lambda x: (x.get("table_name",""), x.get("sla_deadline",""))):
        try:
            days = (_date.fromisoformat(iss["sla_deadline"]) - today).days
            days_str = f"{days}d" if days >= 0 else f"⚠ {abs(days)}d overdue"
        except Exception:
            days_str = "—"

        recur = int(iss.get("recurrence_count") or 0)
        prev  = iss.get("last_failing_rows")

        w.writerow([
            _PRETTY.get(iss["table_name"], iss["table_name"]),
            iss.get("rule_id", ""),
            iss.get("rule_name") or iss.get("rule_id", ""),
            (iss.get("dimension") or "").title(),
            iss.get("failing_rows", 0),
            prev if prev is not None else "",
            iss.get("detected_at", ""),
            iss.get("sla_deadline", ""),
            days_str,
            (iss.get("urgency_band") or "").title(),
            (iss.get("status") or "").replace("_", " ").title(),
            f"#{recur + 1}" if recur > 0 else "1st occurrence",
        ])

    tbl_suffix  = "_" + "_".join(table_filter) if table_filter else ""
    safe_name   = inst_name.lower().replace(" ", "_").replace("/", "")[:20]
    filename    = f"issues_{safe_name}{tbl_suffix}_{today.isoformat()}.csv"
    return dcc.send_string(buf.getvalue(), filename=filename)


@app.callback(
    Output("inst-cr-list", "children"),
    Input("inst-cr-status-filter", "value"),
    Input("notif-poll",            "n_intervals"),
    State("auth-store", "data"),
    prevent_initial_call=False,
)
def _inst_cr_list(status_filter, _poll, auth_data):
    import dq_change_request as cr_mod
    le_books = set((auth_data or {}).get("le_books", []))
    role     = (auth_data or {}).get("role", "inst_user")
    all_crs  = cr_mod.get_crs(status=status_filter if status_filter != "all" else None)
    my_crs   = [c for c in all_crs if c["le_book"] in le_books]
    return _build_cr_list(my_crs, role=role)


@app.callback(
    Output("inst-cr-list",     "children", allow_duplicate=True),
    Output("inst-cr-feedback", "children"),
    Input({"type": "cr-action-btn", "index": ALL}, "n_clicks"),
    State("inst-cr-status-filter", "value"),
    State("auth-store", "data"),
    State("cr-version", "data"),
    prevent_initial_call=True,
)
def _inst_cr_action(clicks, status_filter, auth_data, version):
    if not any(c for c in (clicks or []) if c):
        raise dash.exceptions.PreventUpdate
    tid = ctx.triggered_id
    if not isinstance(tid, dict) or tid.get("type") != "cr-action-btn":
        raise dash.exceptions.PreventUpdate
    if not ctx.triggered[0]["value"]:
        raise dash.exceptions.PreventUpdate

    import dq_change_request as cr_mod
    raw     = tid["index"]
    cr_id, new_status = raw.split("|", 1)
    actor    = (auth_data or {}).get("email", "inst_user")
    le_books = set((auth_data or {}).get("le_books", []))
    role     = (auth_data or {}).get("role", "inst_user")

    # Verify the CR belongs to this institution before mutating
    owned_cr = cr_mod.get_cr(cr_id)
    if not owned_cr or owned_cr.get("le_book") not in le_books:
        return _build_cr_list([], role=role), html.Span(
            "Unauthorized: this CR does not belong to your institution.",
            style={"color": C_RED},
        )

    ok, msg = cr_mod.update_status(cr_id, new_status, actor=actor)
    all_crs  = cr_mod.get_crs(status=status_filter if status_filter != "all" else None)
    my_crs   = [c for c in all_crs if c["le_book"] in le_books]
    cr_list  = _build_cr_list(my_crs, role=role)

    if ok:
        label    = cr_mod.STATUS_LABELS.get(new_status, new_status)
        feedback = html.Span([
            html.Span("✓ ", style={"color": C_GREEN, "fontWeight": "900"}),
            html.Span(f"{cr_id} moved to \"{label}\".", style={"color": C_GREEN}),
        ])
    else:
        feedback = html.Span(msg, style={"color": C_RED})

    return cr_list, feedback


# ── notification bell callbacks ───────────────────────────────────────────────

@app.callback(
    Output("inst-notif-show", "data"),
    Input("inst-bell-btn",     "n_clicks"),
    Input("page-content",      "n_clicks"),
    State("inst-notif-show",   "data"),
    prevent_initial_call=True,
)
def _toggle_notif_panel(bell_clicks, page_clicks, currently_shown):
    trig = ctx.triggered_id
    if trig == "inst-bell-btn" and bell_clicks:
        return not bool(currently_shown)
    if trig == "page-content" and currently_shown:
        return False
    raise dash.exceptions.PreventUpdate


@app.callback(
    Output("notif-overlay", "children"),
    Output("notif-overlay", "style"),
    Input("inst-notif-show", "data"),
    Input("notif-poll",      "n_intervals"),
    State("auth-store",      "data"),
    prevent_initial_call=False,
)
def _notif_panel_render(show, _poll, auth_data):
    import dq_inst_portal as inst_mod

    hidden = {"position": "fixed", "top": "52px", "right": "80px",
              "zIndex": "500", "display": "none"}
    shown  = {"position": "fixed", "top": "52px", "right": "80px",
              "zIndex": "500", "display": "block"}

    auth    = auth_data or {}
    role    = auth.get("role", "")
    user_id = auth.get("user_id", "")

    if role != "inst_user" or not user_id:
        return html.Div(), hidden

    if show:
        return inst_mod.inst_notification_panel(user_id), shown

    return html.Div(), hidden


@app.callback(
    Output("inst-notif-show", "data", allow_duplicate=True),
    Input("inst-mark-all-read", "n_clicks"),
    State("auth-store", "data"),
    prevent_initial_call=True,
)
def _mark_all_read(n, auth_data):
    if not n:
        raise dash.exceptions.PreventUpdate
    from dq_notifications import mark_all_read
    user_id = (auth_data or {}).get("user_id", "")
    if user_id:
        mark_all_read(user_id)
    return True  # keep panel open, re-render with cleared notifications


# ── issues XLSX download (BNR alerts page) ────────────────────────────────────

@app.callback(
    Output("issues-download", "data"),
    Input("issues-download-btn", "n_clicks"),
    State("alerts-cat-filter",   "value"),
    prevent_initial_call=True,
)
def _on_issues_download(n_clicks, cat_filter):
    if not n_clicks:
        raise dash.exceptions.PreventUpdate
    from dq_issue_tracker import get_issues
    issues = get_issues(status="resolved")
    issues.sort(key=lambda x: x.get("resolved_at") or "", reverse=True)
    return dcc.send_bytes(_issues_to_xlsx(issues), "dq_resolved_issues.xlsx")



# ── CR: per-table CSV download ────────────────────────────────────────────────

@app.callback(
    Output("dl-nav", "href", allow_duplicate=True),
    Input({"type": "cr-tbl-dl-btn", "index": ALL}, "n_clicks"),
    prevent_initial_call=True,
)
def _cr_table_dl(clicks):
    if not any(c for c in (clicks or []) if c):
        raise dash.exceptions.PreventUpdate
    tid = ctx.triggered_id
    if not isinstance(tid, dict) or tid.get("type") != "cr-tbl-dl-btn":
        raise dash.exceptions.PreventUpdate
    if not ctx.triggered[0]["value"]:
        raise dash.exceptions.PreventUpdate

    raw = tid["index"]
    if "|" not in raw:
        raise dash.exceptions.PreventUpdate
    le_book, table_name = raw.split("|", 1)

    issue_reports_dir = _DIR / "issue_reports"
    zips = (sorted([z for z in issue_reports_dir.glob(f"{le_book}_*.zip")
                    if not z.name.endswith("_resolved.zip")], reverse=True)
            if issue_reports_dir.exists() else [])
    if not zips:
        raise dash.exceptions.PreventUpdate

    # stream the single {table}.xlsx via the Flask route (no base64 inlining)
    return f"/download/issue-report/{le_book}/{table_name}?t={ctx.triggered[0]['value']}"


# ── CR: table-level approval ───────────────────────────────────────────────────

@app.callback(
    Output("cr-version",         "data", allow_duplicate=True),
    Output("cr-action-feedback", "children", allow_duplicate=True),
    Input({"type": "cr-tbl-approve-btn", "index": ALL}, "n_clicks"),
    State("auth-store", "data"),
    State("cr-version", "data"),
    prevent_initial_call=True,
)
def _approve_table_cb(clicks, auth_data, version):
    role = (auth_data or {}).get("role", "viewer")
    if not _auth.is_admin(role):
        raise dash.exceptions.PreventUpdate
    if not any(c for c in (clicks or []) if c):
        raise dash.exceptions.PreventUpdate
    tid = ctx.triggered_id
    if not isinstance(tid, dict) or tid.get("type") != "cr-tbl-approve-btn":
        raise dash.exceptions.PreventUpdate
    if not ctx.triggered[0]["value"]:
        raise dash.exceptions.PreventUpdate

    raw = tid["index"]
    if "|" not in raw:
        raise dash.exceptions.PreventUpdate
    cr_id, table_name = raw.split("|", 1)

    import dq_change_request as cr_mod
    actor  = (auth_data or {}).get("email", "")
    ok, msg = cr_mod.approve_table(cr_id, table_name, actor=actor)

    feedback = (
        html.Span([html.Span("✓ ", style={"color": C_GREEN, "fontWeight": "900"}),
                   html.Span(msg, style={"color": C_GREEN})])
        if ok else html.Span(msg, style={"color": C_RED})
    )
    return (version or 0) + 1 if ok else version, feedback


# ── open-issue ZIP download (alerts page, per institution) ────────────────────

@app.callback(
    Output("dl-nav", "href", allow_duplicate=True),
    Input({"type": "open-issue-dl-btn", "index": ALL}, "n_clicks"),
    prevent_initial_call=True,
)
def _on_open_issue_dl(clicks):
    if not any(c for c in (clicks or []) if c):
        raise dash.exceptions.PreventUpdate
    tid = ctx.triggered_id
    if not isinstance(tid, dict) or tid.get("type") != "open-issue-dl-btn":
        raise dash.exceptions.PreventUpdate
    if not ctx.triggered[0]["value"]:
        raise dash.exceptions.PreventUpdate

    lb = tid["index"]
    issue_reports_dir = _DIR / "issue_reports"
    if issue_reports_dir.exists() and sorted(issue_reports_dir.glob(f"{lb}_*.zip")):
        # stream via the Flask route (handles 100 MB+ reports; no base64 inlining).
        # ?t= makes the href change each click so dcc.Location re-navigates.
        return f"/download/issue-report/{lb}?t={ctx.triggered[0]['value']}"

    raise dash.exceptions.PreventUpdate


# ── dev server ─────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8050, debug=False)
