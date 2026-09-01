#get failing rows for each institution and write to excel sheet
from __future__ import annotations
import io
import logging
import os
import re
import tempfile
import zipfile
from datetime import date as _date, datetime as _datetime
from decimal import Decimal
from pathlib import Path
from sqlalchemy import text

from dq.rules.accuracy import (
    ACC_RULE_META, ACC_TABLE_RULES,
    VALID_ACCOUNT_STATUS, VALID_PERFORMANCE_CLASS, VALID_GENDER,
    VALID_ACCOUNT_TYPE, CORPORATE_LEGAL_STATUS,
)
from dq.rules.completeness import MANDATORY_COLUMNS
from dq.rules.timeliness import TIM_RULE_META, TIM_TABLE_RULES, FRESHNESS_WINDOW_DAYS
from dq.rules.uniqueness import UNI_RULE_META, UNI_TABLE_RULES
from dq.rules.validity import (
    VAL_RULE_META, VAL_TABLE_RULES,
    MIN_PHONE_DIGITS, MIN_NATIONAL_ID, INTEREST_RATE_MAX, MIN_AGE_AT_OPEN,
)
from dq.sql.metadata import all_columns

log = logging.getLogger("dq.exports.failing_rows")

SCRIPT_DIR= Path(__file__).resolve().parents[2]
# SCRIPT_DIR_=Path(__file__)
# print("SCRIPT_DIR_:",SCRIPT_DIR_)
# print("SCRIPT_DIR",SCRIPT_DIR)
ISSUE_REPORTS_DIR = SCRIPT_DIR / "issue_reports"

# When a single issue exceeds this, overflow into consecutive numbered sheets.
_EXCEL_MAX_ROWS = 1_048_575

#referential integrity rules 
_REL_RULE_IDS = {rid for rid, m in ACC_RULE_META.items() if "parent_table" in m} #must contain a parent table

# Identifier/context columns shown on the left so a record can be located.
IDENTIFIER_COLS: dict[str, list[str]] = {
    "customers_expanded":     ["le_book", "customer_id", "customer_name"],
    "accounts":               ["le_book", "account_no", "customer_id", "account_name"],
    "contracts_disburse":     ["le_book", "contract_id", "business_date"],
    "contract_loans":         ["le_book", "contract_sequence_number"],
    "contract_schedules":     ["le_book", "contract_sequence_number", "schedule_date"],
    "contracts_expanded":     ["le_book", "contract_sequence_number", "customer_id"],
    "loan_applications_2":    ["le_book", "loan_application_id", "customer_name"],
    "prev_loan_applications": ["le_book", "loan_application_id"],
}

#cast types for excel 
def _coerce(v):
    if v is None or isinstance(v, (str, int, float, bool, _date, _datetime)):
        return v
    if isinstance(v, Decimal):
        return float(v)
    return str(v)


def _affected_cols_for_rule(rule_id: str) -> list[str]:
    """Return the affected column(s) for a rule from metadata."""
    if rule_id in ACC_RULE_META:
        meta = ACC_RULE_META[rule_id]
        if "child_col" in meta:
            return [meta["child_col"]]
        return list(meta.get("fields", []))
    if rule_id in VAL_RULE_META:
        return list(VAL_RULE_META[rule_id].get("fields", []))
    if rule_id in UNI_RULE_META:
        return list(UNI_RULE_META[rule_id].get("fields", []))
    return []

#string lateral quoting for sql
def _sqlstr(s: str) -> str:
    """Quote a Python string as a SQL string literal."""
    return "'" + s.replace("'", "''") + "'"


def _acc_invalid(rule_id: str, existing: set) -> tuple[str, list[str]] | None:
    """Row-level predicate selecting accuracy-INVALID rows, + the failing column(s)."""
    def has(*cols: str) -> bool:
        return all(c in existing for c in cols)

    if rule_id == "ACC-002":
        if not has("account_status"):
            return None
        vals = ", ".join(f"'{v}'" for v in sorted(str(x) for x in VALID_ACCOUNT_STATUS))
        return (f'"account_status" IS NOT NULL AND "account_status"::TEXT NOT IN ({vals})',
                ["account_status"])
    if rule_id == "ACC-003":
        if not has("performance_class"):
            return None
        vals = ", ".join(f"'{v}'" for v in sorted(VALID_PERFORMANCE_CLASS))
        return (f'"performance_class" IS NOT NULL AND UPPER(TRIM("performance_class"::TEXT)) NOT IN ({vals})',
                ["performance_class"])
    if rule_id == "ACC-004":
        if not has("customer_gender"):
            return None
        vals = ", ".join(f"'{v}'" for v in sorted(VALID_GENDER))
        return (f'"customer_gender" IS NOT NULL AND UPPER(TRIM("customer_gender"::TEXT)) NOT IN ({vals})',
                ["customer_gender"])
    if rule_id == "ACC-005":
        if not has("account_type"):
            return None
        vals = ", ".join(f"'{v}'" for v in sorted(VALID_ACCOUNT_TYPE))
        return (f'"account_type" IS NOT NULL AND UPPER(TRIM("account_type"::TEXT)) NOT IN ({vals})',
                ["account_type"])
    if rule_id == "ACC-010":
        if not has("customer_gender", "legal_status"):
            return None
        corp = ", ".join(f"'{v}'" for v in sorted(str(x) for x in CORPORATE_LEGAL_STATUS))
        return (f'"customer_gender" IS NOT NULL AND "legal_status" IS NOT NULL '
                f'AND "legal_status"::TEXT IN ({corp}) AND UPPER(TRIM("customer_gender"::TEXT)) <> \'C\'',
                ["customer_gender", "legal_status"])
    if rule_id == "ACC-012":
        if not has("customer_gender", "marital_status"):
            return None
        return (f'"customer_gender" IS NOT NULL AND "marital_status" IS NOT NULL '
                f'AND UPPER(TRIM("customer_gender"::TEXT)) = \'C\' AND UPPER(TRIM("marital_status"::TEXT)) <> \'NA\'',
                ["customer_gender", "marital_status"])
    return None


def _val_invalid(rule_id: str, existing: set) -> tuple[str, list[str]] | None:
    """Row-level predicate selecting validity-INVALID rows, + the failing column(s)."""
    def has(*cols: str) -> bool:
        return all(c in existing for c in cols)

    if rule_id == "VAL-001":
        if not has("email_id"):
            return None
        return (r"""("email_id" IS NOT NULL AND "email_id"::TEXT !~ '^[^@\s]+@[^@\s]+\.[^@\s]+$')""",
                ["email_id"])
    if rule_id == "VAL-002":
        cols = [c for c in ("work_telephone", "home_telephone") if c in existing]
        if not cols:
            return None
        pred = " OR ".join(
            f'("{c}" IS NOT NULL AND TRIM("{c}"::TEXT) <> \'\' '
            f'AND LENGTH(REGEXP_REPLACE("{c}"::TEXT, \'[^0-9]\', \'\', \'g\')) < {MIN_PHONE_DIGITS})'
            for c in cols)
        return (f"({pred})", cols)
    if rule_id == "VAL-003":
        cols = [c for c in ("currency", "mis_currency") if c in existing]
        if not cols:
            return None
        pred = " OR ".join(
            f'("{c}" IS NOT NULL AND TRIM("{c}"::TEXT) <> \'\' AND TRIM("{c}"::TEXT) !~ \'^[A-Z]{{3}}$\')'
            for c in cols)
        return (f"({pred})", cols)
    if rule_id == "VAL-004":
        if not has("national_id_type", "national_id_number"):
            return None
        return (f'"national_id_type" IS NOT NULL AND TRIM("national_id_type"::TEXT) <> \'\' '
                f'AND LENGTH(TRIM(COALESCE("national_id_number"::TEXT, \'\'))) < {MIN_NATIONAL_ID}',
                ["national_id_number"])
    if rule_id == "VAL-010":
        if not has("interest_rate_dr"):
            return None
        return (f'"interest_rate_dr" IS NOT NULL AND ("interest_rate_dr"::NUMERIC < 0 '
                f'OR "interest_rate_dr"::NUMERIC > {INTEREST_RATE_MAX})', ["interest_rate_dr"])
    if rule_id == "VAL-011":
        if not has("interest_rate_cr"):
            return None
        return (f'"interest_rate_cr" IS NOT NULL AND ("interest_rate_cr"::NUMERIC < 0 '
                f'OR "interest_rate_cr"::NUMERIC > {INTEREST_RATE_MAX})', ["interest_rate_cr"])
    if rule_id == "VAL-012":
        cols = [c for c in ("current_disbursed_amt", "previous_disbursed_amt") if c in existing]
        if not cols:
            return None
        pred = " OR ".join(f'("{c}" IS NOT NULL AND "{c}"::NUMERIC < 0)' for c in cols)
        return (f"({pred})", cols)
    if rule_id == "VAL-013":
        if not has("emi_amount"):
            return None
        return ('"emi_amount" IS NOT NULL AND "emi_amount"::NUMERIC <= 0', ["emi_amount"])
    if rule_id == "VAL-014":
        cols = [c for c in ("outstanding_amount_lcy", "outstanding_amount",
                            "principal_amount_due", "int_amount_due",
                            "due_amount", "principal_amount_lcy") if c in existing]
        if not cols:
            return None
        pred = " OR ".join(f'("{c}" IS NOT NULL AND "{c}"::NUMERIC < 0)' for c in cols)
        return (f"({pred})", cols)
    if rule_id == "VAL-015":
        if not has("applied_amount_lcy"):
            return None
        return ('"applied_amount_lcy" IS NOT NULL AND "applied_amount_lcy"::NUMERIC <= 0',
                ["applied_amount_lcy"])
    if rule_id == "VAL-016":
        if not has("num_of_instalments"):
            return None
        return ('"num_of_instalments" IS NOT NULL AND "num_of_instalments"::NUMERIC < 1',
                ["num_of_instalments"])
    if rule_id == "VAL-020":
        if not has("num_instalments_paid", "num_of_instalments"):
            return None
        return ('"num_instalments_paid" IS NOT NULL AND "num_of_instalments" IS NOT NULL '
                'AND "num_instalments_paid"::NUMERIC > "num_of_instalments"::NUMERIC',
                ["num_instalments_paid", "num_of_instalments"])
    if rule_id == "VAL-021":
        if not has("approved_amount_lcy", "applied_amount_lcy"):
            return None
        return ('"approved_amount_lcy" IS NOT NULL AND "applied_amount_lcy" IS NOT NULL '
                'AND "approved_amount_lcy"::NUMERIC > "applied_amount_lcy"::NUMERIC',
                ["approved_amount_lcy", "applied_amount_lcy"])
    if rule_id == "VAL-022":
        if not has("date_of_birth", "customer_open_date"):
            return None
        return (f'"date_of_birth" IS NOT NULL AND "customer_open_date" IS NOT NULL '
                f'AND ("customer_open_date"::DATE - "date_of_birth"::DATE) < {MIN_AGE_AT_OPEN * 365}',
                ["date_of_birth", "customer_open_date"])
    if rule_id == "VAL-030":
        if not has("currency", "current_disbursed_amt"):
            return None
        return (
            '"currency" IS NOT NULL AND UPPER(TRIM("currency"::TEXT)) = \'RWF\' '
            'AND "current_disbursed_amt" IS NOT NULL '
            'AND "current_disbursed_amt"::NUMERIC < 1000',
            ["currency", "current_disbursed_amt"],
        )
    if rule_id == "VAL-031":
        if not has("currency", "current_disbursed_amt"):
            return None
        return (
            '"currency" IS NOT NULL AND UPPER(TRIM("currency"::TEXT)) <> \'RWF\' '
            'AND "current_disbursed_amt" IS NOT NULL '
            'AND "current_disbursed_amt"::NUMERIC < 1',
            ["currency", "current_disbursed_amt"],
        )
    if rule_id == "VAL-032":
        if not has("disbursed_amount"):
            return None
        return (
            '"disbursed_amount" IS NOT NULL AND "disbursed_amount"::NUMERIC < 1000',
            ["disbursed_amount"],
        )
    if rule_id == "VAL-033":
        if not has("principal_amount_lcy"):
            return None
        return (
            '"principal_amount_lcy" IS NOT NULL AND "principal_amount_lcy"::NUMERIC < 1000',
            ["principal_amount_lcy"],
        )
    if rule_id == "VAL-034":
        if not has("application_date", "business_date"):
            return None
        return (
            '"application_date" IS NOT NULL AND "business_date" IS NOT NULL '
            'AND "application_date"::DATE > "business_date"::DATE',
            ["application_date", "business_date"],
        )
    if rule_id == "VAL-048":
        if not has("national_id_number", "customer_gender"):
            return None
        return (
            '"national_id_number" IS NOT NULL AND TRIM("national_id_number"::TEXT) != \'\' '
            'AND "customer_gender" IN (\'M\', \'F\') '
            "AND NOT (TRIM(\"national_id_number\"::TEXT) ~ '^[0-9]{16}$' "
            'AND (("customer_gender" = \'F\' AND RIGHT(TRIM("national_id_number"::TEXT), 1) = \'7\') '
            '  OR ("customer_gender" = \'M\' AND RIGHT(TRIM("national_id_number"::TEXT), 1) = \'8\')))',
            ["national_id_number", "customer_gender"],
        )
    return None


def _tim_invalid(rule_id: str, existing: set) -> tuple[str, list[str]] | None:
    """Row-level predicate selecting timeliness-INVALID rows, + the failing column(s).
    """
    def has(*cols: str) -> bool:
        return all(c in existing for c in cols)

    _NO_FUTURE: dict[str, str] = {
        "TIM-001": "customer_open_date",
        "TIM-003": "account_open_date",
        "TIM-004": "date_creation",
        "TIM-005": "business_date",
        "TIM-006": "approval_date",
        "TIM-007": "application_date",
    }
    if rule_id in _NO_FUTURE:
        c = _NO_FUTURE[rule_id]
        if not has(c):
            return None
        return (f'"{c}" IS NOT NULL AND "{c}"::DATE > CURRENT_DATE', [c])

    if rule_id == "TIM-002":
        if not has("date_of_birth"):
            return None
        return (
            '"date_of_birth" IS NOT NULL AND '
            '("date_of_birth"::DATE < \'1900-01-01\'::DATE OR "date_of_birth"::DATE > CURRENT_DATE)',
            ["date_of_birth"],
        )

    # Logical date-order rules: valid requires a <= b (or a < b for strict=True).
    # Failing = both present AND a > b (or a >= b for strict).
    _ORDER: dict[str, tuple[str, str, bool]] = {
        "TIM-010": ("date_creation",  "date_last_modified",  False),
        "TIM-011": ("start_date",     "maturity_date",       True),
        "TIM-012": ("schedule_date",  "payment_date",        False),
        "TIM-013": ("commence_date",  "benefit_expiry_date", False),
        "TIM-014": ("commence_date",  "ins_expiry_date",     False),
    }
    if rule_id in _ORDER:
        a, b, strict = _ORDER[rule_id]
        if not has(a, b):
            return None
        op = ">=" if strict else ">"
        return (
            f'"{a}" IS NOT NULL AND "{b}" IS NOT NULL AND "{a}"::DATE {op} "{b}"::DATE',
            [a, b],
        )

    if rule_id == "TIM-020":
        if not has("date_last_modified"):
            return None
        return (
            f'"date_last_modified" IS NOT NULL '
            f'AND "date_last_modified"::DATE < CURRENT_DATE - INTERVAL \'{FRESHNESS_WINDOW_DAYS} days\'',
            ["date_last_modified"],
        )

    return None

#get failing columns that are also present in db
def _failing_columns(table: str, existing: set) -> list[str]:
    cols: set[str] = {c for c in MANDATORY_COLUMNS.get(table, []) if c in existing}
    for rid in ACC_TABLE_RULES.get(table, []):
        cols |= {c for c in ACC_RULE_META.get(rid, {}).get("fields", []) if c in existing}
    for rid in VAL_TABLE_RULES.get(table, []):
        cols |= {c for c in VAL_RULE_META.get(rid, {}).get("fields", []) if c in existing}
    for rid in UNI_TABLE_RULES.get(table, []):
        cols |= {c for c in UNI_RULE_META.get(rid, {}).get("fields", []) if c in existing}
    for rid in TIM_TABLE_RULES.get(table, []):
        cols |= {c for c in TIM_RULE_META.get(rid, {}).get("fields", []) if c in existing}
    return sorted(cols) 


#skip them for now ( too many rows )
_SKIP_COMPLETENESS_TABLES: frozenset[str] = frozenset({"contracts_expanded","customers_expanded"})


def build_failing_union(schema: str, table: str, existing: set,
                        valid_le_books: frozenset, limit: int = 0,
                        extra_where: str = "", per_issue_cap: int = 0):
    """Return (sql, output_columns, issue_cols) selecting failing rows across all
    dimensions for `table`, or None if nothing is applicable.

    output_columns order: identifiers → 'issue_type' → failing columns (rightmost).
    issue_cols: {issue_label: [affected column(s)]} — for red-highlighting per sheet.
    Rows are ORDER BY le_book, issue_type so the caller can stream one (institution,
    issue) group at a time. per_issue_cap > 0 caps rows per (institution, issue)
    server-side (one Excel sheet's worth).

    Tables in _SKIP_COMPLETENESS_TABLES omit the MANDATORY_COLUMNS completeness
    branches to avoid a CROSS JOIN row explosion on wide tables.
    """
    if "le_book" not in existing:
        return None

    id_cols   = [c for c in IDENTIFIER_COLS.get(table, ["le_book"]) if c in existing]
    if "le_book" not in id_cols:
        id_cols = ["le_book"] + id_cols
    # context columns — appended when the table actually has them
    for ctx_col in ("date_creation", "date_last_modified"):
        if ctx_col in existing and ctx_col not in id_cols:
            id_cols.append(ctx_col)
    fail_cols = [c for c in _failing_columns(table, existing) if c not in id_cols]
    if not fail_cols:
        return None

    id_sel   = ", ".join(f'"{c}"' for c in id_cols)
    fail_sel = ", ".join(f'"{c}"' for c in fail_cols)
    data_sel = ", ".join(f'"{c}"' for c in (id_cols + fail_cols))

    lb_clause = ""
    if valid_le_books:
        codes     = ", ".join(f"'{lb}'" for lb in sorted(valid_le_books))
        lb_clause = f'AND "le_book" IN ({codes})'
    # optional extra row filter (e.g. a date_last_modified month) applied to every branch
    if extra_where:
        lb_clause += f' AND ({extra_where})'

    sq = f'"{schema}"."{table}"'
    branches: list[str] = [] #list of individual SQL queeries for different rules and dimensions (completeness, accuracy, validity, uniqueness, timeliness, referential integrity)

  #faling rows for completness, accuracy, validity, uniqueness, referential integrity
    issue_cols: dict[str, list[str]] = {}
    rule_conds: list[tuple[str, str]] = []
    if table not in _SKIP_COMPLETENESS_TABLES:
        for col in MANDATORY_COLUMNS.get(table, []):
            if col in existing:
                label = f"Missing {col}"
                rule_conds.append((f'"{col}" IS NULL', label))
                issue_cols[label] = [col]
    for rid in ACC_TABLE_RULES.get(table, []):
        if rid in _REL_RULE_IDS:
            continue  # referential integrity — own branch below, needs a real JOIN
        r = _acc_invalid(rid, existing)
        if r:
            label = "{}: {}".format(rid, ACC_RULE_META[rid].get("name", rid))
            rule_conds.append((r[0], label))
            issue_cols[label] = r[1]
    for rid in VAL_TABLE_RULES.get(table, []):
        r = _val_invalid(rid, existing)
        if r:
            label = "{}: {}".format(rid, VAL_RULE_META[rid].get("name", rid))
            rule_conds.append((r[0], label))
            issue_cols[label] = r[1]
    for rid in TIM_TABLE_RULES.get(table, []):
        r = _tim_invalid(rid, existing)
        if r:
            label = "{}: {}".format(rid, TIM_RULE_META[rid].get("name", rid))
            rule_conds.append((r[0], label))
            issue_cols[label] = r[1]

    if rule_conds:
        case_arr = ",\n            ".join(
            f'CASE WHEN ({cond}) THEN {_sqlstr(label)} END' for cond, label in rule_conds)
        any_fail = " OR ".join(f'({cond})' for cond, _ in rule_conds)
        branches.append(
            f'SELECT {id_sel}, _u.issue_type AS issue_type, {fail_sel}\n'
            f'FROM (SELECT {data_sel} FROM {sq} WHERE ({any_fail}) {lb_clause}) _f\n'
            f'CROSS JOIN LATERAL unnest(ARRAY[\n            {case_arr}\n        ]::text[]) AS _u(issue_type)\n'
            f'WHERE _u.issue_type IS NOT NULL')

    # uniqueness rules — one branch per rule, ROW_NUMBER() to find duplicates
    for rid in UNI_TABLE_RULES.get(table, []):
        fields = UNI_RULE_META[rid]["fields"]
        keys   = [c for c in fields if c in existing]
        anchor = fields[0]
        if anchor in existing and keys:
            part = ", ".join(f'"{c}"' for c in (["le_book"] + keys))
            issue = f"{rid}: {UNI_RULE_META[rid]['name']}"
            issue_cols[issue] = keys
            sub = (f'SELECT {data_sel}, '
                   f'ROW_NUMBER() OVER (PARTITION BY {part} ORDER BY {part}) AS _rn '
                   f'FROM {sq} WHERE "{anchor}" IS NOT NULL {lb_clause}')
            branches.append(
                f'SELECT {id_sel}, {_sqlstr(issue)} AS issue_type, {fail_sel} '
                f'FROM ({sub}) _q WHERE _rn > 1')

    #rel rules
    for rid in ACC_TABLE_RULES.get(table, []):
        if rid not in _REL_RULE_IDS:
            continue
        meta    = ACC_RULE_META[rid]
        child_c = meta["child_col"]
        if child_c not in existing:
            continue
        parent_t, parent_c = meta["parent_table"], meta["parent_col"]
        issue = f"{rid}: {meta['name']}"
        issue_cols[issue] = [child_c]
        branches.append(
            f'SELECT {id_sel}, {_sqlstr(issue)} AS issue_type, {fail_sel}\n'
            f'FROM {sq}\n'
            f'LEFT JOIN (SELECT DISTINCT "{parent_c}" AS _parent_key '
            f'FROM "{schema}"."{parent_t}") p ON "{child_c}" = p._parent_key\n'
            f'WHERE "{child_c}" IS NOT NULL AND p._parent_key IS NULL {lb_clause}')

    if not branches:
        return None

    output_cols = id_cols + ["issue_type"] + fail_cols
    inner = " UNION ALL ".join(branches)

    if per_issue_cap and per_issue_cap > 0:
        # Cap rows per (institution, issue) server-side so the client receives a
        # bounded set (one Excel sheet's worth) instead of streaming millions.
        cols_q = ", ".join("issue_type" if c == "issue_type" else f'"{c}"'
                           for c in output_cols)
        sql = (f'SELECT {cols_q} FROM ('
               f'SELECT {cols_q}, ROW_NUMBER() OVER (PARTITION BY "le_book", issue_type) AS _srn '
               f'FROM ({inner}) _w) _c WHERE _srn <= {per_issue_cap} '
               f'ORDER BY "le_book", issue_type')
    else:
        # order by institution then issue so the writer can stream one
        # (institution, issue) group at a time → one Excel sheet per issue.
        sql = inner + ' ORDER BY "le_book", issue_type'

    if limit and limit > 0:
        sql += f" LIMIT {limit}"

    return sql, output_cols, issue_cols


_TABLE_TO_COMP_RULE: dict[str, str] = {
    "customers_expanded":     "COMP-001",
    "accounts":               "COMP-002",
    "contracts_disburse":     "COMP-003",
    "contract_loans":         "COMP-004",
    "contract_schedules":     "COMP-005",
    "contracts_expanded":     "COMP-006",
    "loan_applications_2":    "COMP-007",
    "prev_loan_applications": "COMP-008",
}
_RULE_PREFIX_RE = re.compile(r"^([A-Z]+-\d+)")

#write failing rows to excel sheet for each institution
def write_institution_zips(engine, schema: str, table: str,
                           valid_le_books: frozenset, categories: dict,
                           month: str, limit: int = 0,
                           max_rows_per_sheet: int = 50000,
                           extra_where: str = "") -> None:
    
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Font, PatternFill

    RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    RED_FONT = Font(color="9C0006")
    HDR_FONT = Font(bold=True)
    HDR_RED  = Font(bold=True, color="9C0006")
    DATE_FONT = Font(bold=True, color="1F4E79")

    # Load per-issue detected_at and sla_deadline from Greenplum so each sheet
    # carries the correct dates rather than a single export-run timestamp.
    # Keyed by (le_book, table_name, rule_id) so the same rule_id firing on two
    # different tables for the same institution doesn't overwrite the other's dates.
    _issue_dates: dict[tuple[str, str, str], tuple[str, str]] = {}
    lb_list = list(valid_le_books)
    if lb_list:
        try:
            from sqlalchemy import bindparam, text as _text
            from storage.postgres.connection import get_engine
            _sql = _text(
                "SELECT le_book, table_name, rule_id, detected_at, sla_deadline "
                "FROM dq_open_issues "
                "WHERE table_name=:tbl AND le_book IN :lbs "
                "AND status IN ('open','pending_resolution')"
            ).bindparams(bindparam("lbs", expanding=True))
            with get_engine().connect() as _con:
                for _r in _con.execute(_sql, {"tbl": table, "lbs": lb_list}).mappings():
                    _issue_dates[(_r["le_book"], _r["table_name"], _r["rule_id"])] = (
                        _r["detected_at"] or "", _r["sla_deadline"] or ""
                    )
        except Exception as _exc:
            log.warning("Could not load issue dates for %s: %s — sheets will have blank dates", table, _exc)

    def _dates_for(lb: str, issue_label: str) -> tuple[str, str]:
        """Return (detected_at, sla_deadline) for the given issue label and institution."""
        if issue_label.startswith("Missing "):
            rid = _TABLE_TO_COMP_RULE.get(table)
        else:
            _m = _RULE_PREFIX_RE.match(issue_label)
            rid = _m.group(1) if _m else None
        if rid is None:
            return ("", "")
        return _issue_dates.get((lb, table, rid), ("", ""))

    # Fetch column list once; no row data needed yet.
    with engine.connect() as _cc:
        existing = all_columns(_cc, schema, table)

    # Probe with all institutions to get out_cols/issue_cols (same regardless of le_book).
    _probe = build_failing_union(schema, table, existing, valid_le_books, limit,
                                 extra_where=extra_where, per_issue_cap=0)
    if not _probe:
        return
    _, out_cols, issue_cols = _probe

    # sheet columns = output cols minus issue_type; insert enrichment after le_book
    sheet_cols = [c for c in out_cols if c != "issue_type"]
    header: list[str] = []
    for c in sheet_cols:
        header.append(c)
        if c == "le_book":
            header += ["stakeholder_name", "category_type"]
    header.append("issue_type")
    header.append("detected_at")
    header.append("sla_deadline")

    from storage.evidence_store import store_rows as _store_evidence
    run_date = month + "-01"   # label stored with evidence rows

    # Mutable state shared by the helper closures; reset per institution.
    st: dict = {"wb": None, "path": None, "ws": None, "issue": None,
                "affected": set(), "n": 0, "used": set(), "part": 1,
                "detected_at": "", "sla_deadline": "", "lb": None}
    evidence_buf: dict[str, list[dict]] = {}

    def _sheet_title_part(issue: str, part: int) -> str:
        sfx  = f" ({part})" if part > 1 else ""
        base = re.sub(r"[\[\]:*?/\\]", " ", issue)
        name = (base[:31 - len(sfx)] + sfx).strip() or "issue"
        i    = 2
        while name.lower() in st["used"]:
            sfx2 = f" ({i})"
            name = (base[:31 - len(sfx2)] + sfx2).strip()
            i   += 1
        st["used"].add(name.lower())
        return name

    def _start_sheet(issue: str, lb: str, part: int = 1):
        ws       = st["wb"].create_sheet(title=_sheet_title_part(issue, part))
        affected = set(issue_cols.get(issue, []))
        det, sla = _dates_for(lb, issue)
        cells    = []
        for col in header:
            cell = WriteOnlyCell(ws, value=col)
            if col in affected:
                cell.font, cell.fill = HDR_RED, RED_FILL
            elif col in ("detected_at", "sla_deadline"):
                cell.font = DATE_FONT
            else:
                cell.font = HDR_FONT
            cells.append(cell)
        ws.append(cells)
        st.update(ws=ws, issue=issue, affected=affected, n=0, part=part,
                  detected_at=det, sla_deadline=sla)

    def _write_row(m: dict, name: str, ctype: str):
        if st["n"] >= _EXCEL_MAX_ROWS:
            _start_sheet(st["issue"], st["lb"], part=st["part"] + 1)
        affected = st["affected"]
        cells = []
        for col in sheet_cols:
            cell = WriteOnlyCell(st["ws"], value=_coerce(m.get(col)))
            if col in affected:
                cell.fill, cell.font = RED_FILL, RED_FONT
            cells.append(cell)
            if col == "le_book":
                cells.append(WriteOnlyCell(st["ws"], value=name))
                cells.append(WriteOnlyCell(st["ws"], value=ctype))
        cells.append(WriteOnlyCell(st["ws"], value=st["issue"]))
        cells.append(WriteOnlyCell(st["ws"], value=st["detected_at"]))
        cells.append(WriteOnlyCell(st["ws"], value=st["sla_deadline"]))
        st["ws"].append(cells)
        st["n"] += 1
        # buffer raw row for evidence store
        evidence_buf.setdefault(st["issue"], []).append(
            {col: m.get(col) for col in sheet_cols}
        )

    def _new_workbook():
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        st.update(wb=Workbook(write_only=True), path=path, ws=None,
                  issue=None, used=set())

    def _flush_issue_evidence(issue_label: str, lb: str) -> None:
        """Flush buffered rows for one issue to the evidence store and free RAM."""
        rows = evidence_buf.pop(issue_label, [])
        if not rows:
            return
        if issue_label.startswith("Missing "):
            rule_id = _TABLE_TO_COMP_RULE.get(table)
        else:
            _m = _RULE_PREFIX_RE.match(issue_label or "")
            rule_id = _m.group(1) if _m else None
        if rule_id:
            try:
                _store_evidence(lb, rule_id, table, run_date, rows)
            except Exception as _exc:
                log.warning("evidence store failed %s/%s/%s: %s",
                            lb, rule_id, table, _exc)

    def _close(lb: str):
        if st["wb"] is None:
            return
        st["wb"].save(st["path"])
        zip_path = ISSUE_REPORTS_DIR / f"{lb}_{month}.zip"
        arcname  = f"{table}.xlsx"
        # If the ZIP already contains this table's sheet, rebuild it without
        # the stale entry before appending the fresh one (guards against
        # duplicate entries from partial re-runs).
        if zip_path.exists():
            with zipfile.ZipFile(zip_path, "r") as _rz:
                _entries = _rz.namelist()
                if arcname in _entries:
                    kept = [(n, _rz.read(n)) for n in _entries if n != arcname]
                else:
                    kept = None
            if kept is not None:
                tmp = zip_path.with_suffix(".zip.tmp")
                with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as _wz:
                    for n, data in kept:
                        _wz.writestr(n, data)
                tmp.replace(zip_path)
        with zipfile.ZipFile(zip_path, "a", zipfile.ZIP_DEFLATED) as zf:
            zf.write(st["path"], arcname=arcname)
        os.unlink(st["path"])
        log.info("  ZIP %-6s  %s.xlsx", lb, table)
        st.update(wb=None, path=None, ws=None)
        # Flush any remaining issue (the last one — all others were flushed at transition).
        for issue_label in list(evidence_buf.keys()):
            _flush_issue_evidence(issue_label, lb)

    # Query one institution at a time so each Greenplum cursor covers only that
    # institution's rows — avoids the multi-hour single-query timeout that kills
    # large tables (e.g. customers_8expanded at ~4 M rows across all le_books).
    for lb in sorted(valid_le_books):
        built = build_failing_union(schema, table, existing, frozenset({lb}), limit,
                                    extra_where=extra_where, per_issue_cap=0)
        if not built:
            continue
        sql, _, _ = built  # issue_cols/out_cols are institution-independent

        st.clear()
        st.update(wb=None, path=None, ws=None, issue=None,
                  affected=set(), n=0, used=set(), part=1,
                  detected_at="", sla_deadline="", lb=lb)
        evidence_buf.clear()

        with engine.connect() as conn:
            try:
                # Referential-integrity branches join against a deduplicated
                # parent table that can be tens of millions of rows — default
                # work_mem spills that to disk.
                conn.execute(text("SET work_mem = '512MB'"))
                conn.execute(text("SET statement_timeout = '30min'"))
            except Exception:
                conn.rollback()

            result = conn.execution_options(stream_results=True).execute(text(sql))
            cur_issue: str | None = None
            info  = categories.get(lb, {})
            name  = info.get("name") or lb
            name  = name.title() if isinstance(name, str) else name
            ctype = info.get("category_type") or ""

            for row in result:
                m     = dict(row._mapping)
                issue = m.get("issue_type")
                if cur_issue is None:
                    _new_workbook()  # defer until first row so empty institutions skip ZIP
                    _start_sheet(issue, lb)
                    cur_issue = issue
                elif issue != cur_issue:
                    _flush_issue_evidence(cur_issue, lb)
                    _start_sheet(issue, lb)
                    cur_issue = issue
                _write_row(m, name, ctype)

            if st["wb"] is not None:
                _close(lb)


# def _copy_detection_sheets(
#     wb,
#     det_zip: Path,
#     table: str,
#     rule_id: str,
#     xlsx_cache: dict,
# ) -> None:
#     """Append failing-row sheet(s) for rule_id from the detection ZIP into wb.

#     xlsx_cache is a caller-owned dict keyed by (zip_path, table) → bytes | None,
#     so the same table XLSX is only read from disk once per resolved-ZIP build.
#     """
#     import io as _io
#     from openpyxl import load_workbook as _load_wb

#     cache_key = (det_zip, table)
#     if cache_key not in xlsx_cache:
#         try:
#             with zipfile.ZipFile(det_zip) as _outer:
#                 xlsx_cache[cache_key] = (
#                     _outer.read(f"{table}.xlsx")
#                     if f"{table}.xlsx" in _outer.namelist() else None
#                 )
#         except Exception as exc:
#             log.warning("Cannot open detection ZIP %s: %s", det_zip.name, exc)
#             xlsx_cache[cache_key] = None

#     xlsx_bytes = xlsx_cache[cache_key]
#     if xlsx_bytes is None:
#         return

#     try:
#         src_wb = _load_wb(_io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
#     except Exception as exc:
#         log.warning("Cannot parse detection XLSX %s/%s: %s", det_zip.name, table, exc)
#         return

#     is_comp = rule_id.startswith("COMP-")
#     try:
#         for sname in src_wb.sheetnames:
#             if is_comp:
#                 if not sname.startswith("Missing "):
#                     continue
#                 dst_title = sname[:31]
#             else:
#                 if not sname.startswith(rule_id):
#                     continue
#                 dst_title = f"{rule_id} — Records"[:31]

#             src_ws = src_wb[sname]
#             part   = 1
#             dst_ws = None
#             n      = 0
#             for row_vals in src_ws.iter_rows(values_only=True):
#                 if dst_ws is None or n >= _EXCEL_MAX_ROWS:
#                     sfx    = f" ({part})" if part > 1 else ""
#                     base   = re.sub(r"[\[\]:*?/\\]", " ", dst_title)
#                     title  = (base[:31 - len(sfx)] + sfx).strip() or "records"
#                     dst_ws = wb.create_sheet(title=title)
#                     n      = 0
#                     part  += 1
#                 dst_ws.append(list(row_vals))
#                 n += 1
#     except Exception as exc:
#         log.warning("Error copying detection sheet %s/%s/%s: %s",
#                     det_zip.name, table, rule_id, exc)
#     finally:
#         src_wb.close()


def write_resolved_institution_zip(le_book: str, month: str, resolved_issues: list[dict]) -> bool:
    """Build issue_reports/{le_book}_{month}_resolved.zip from Greenplum issue metadata.
    Returns True if the ZIP was written, False if resolved_issues is empty.
    """
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Alignment, Font, PatternFill

    if not resolved_issues:
        return False

    dst_zip = ISSUE_REPORTS_DIR / f"{le_book}_{month}_resolved.zip"

    #styles
    HDR_FONT      = Font(bold=True, size=10, color="FFFFFF")
    HDR_FILL      = PatternFill("solid", fgColor="1F4E79")   # dark blue header
    HDR_FILL_GRN  = PatternFill("solid", fgColor="16A34A")   # green header for resolved cols
    HDR_FILL_RED  = PatternFill("solid", fgColor="B91C1C")   # red header for remaining cols
    HDR_ALIGN     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    DATA_FONT     = Font(size=10)
    DATA_FILL_ALT = PatternFill("solid", fgColor="F0F4F8")   # alternating row
    GRN_FILL      = PatternFill("solid", fgColor="DCFCE7")
    GRN_FILL_ALT  = PatternFill("solid", fgColor="BBF7D0")
    GRN_FONT      = Font(size=10, color="15803D", bold=True)
    RED_DATA_FILL = PatternFill("solid", fgColor="FEE2E2")
    RED_DATA_FONT = Font(size=10, color="B91C1C", bold=True)
    AMB_DATA_FILL = PatternFill("solid", fgColor="FEF9C3")
    AMB_DATA_FONT = Font(size=10, color="92400E", bold=True)

    # Column layout:
    #   0  INSTITUTION        blue hdr
    #   1  TABLE              blue hdr
    #   2  RULE ID            blue hdr
    #   3  RULE NAME          blue hdr
    #   4  DIMENSION          blue hdr
    #   5  ROWS AT DETECTION  blue hdr
    #   6  ROWS RESOLVED      green hdr  ← how many were fixed
    #   7  ROWS REMAINING     red hdr    ← still failing (0 = fully resolved)
    #   8  RESOLUTION         green hdr  ← "Full" or "Partial"
    #   9  DETECTED           blue hdr
    #  10  RESOLVED           green hdr
    #  11  SLA DEADLINE       blue hdr
    #  12  ON TIME            green hdr
    HEADERS = ["INSTITUTION", "TABLE", "RULE ID", "RULE NAME", "DIMENSION",
               "ROWS AT DETECTION", "ROWS RESOLVED", "ROWS REMAINING", "RESOLUTION",
               "DETECTED", "RESOLVED", "SLA DEADLINE", "ON TIME"]
    COL_W   = [22, 24, 10, 38, 14, 18, 16, 16, 12, 12, 12, 14, 10]
    GREEN_IDX     = {6, 8, 10, 12}   # ROWS RESOLVED, RESOLUTION, RESOLVED date, ON TIME
    RED_IDX       = {7}              # ROWS REMAINING (red when > 0)
    PARTIAL_IDX   = {8}              # RESOLUTION cell uses amber when partial

    # group by table then rule
    by_table: dict[str, list[dict]] = {}
    for iss in resolved_issues:
        by_table.setdefault(iss["table_name"], []).append(iss)

    tmp_files: list[tuple[str, str]] = []
    _xlsx_cache: dict = {}   # (det_zip_path, table) -> bytes | None
    try:
        for table, issues in sorted(by_table.items()):
            wb = Workbook(write_only=True)

            # one sheet per rule (sorted by rule_id for consistency)
            for iss in sorted(issues, key=lambda x: x.get("rule_id", "")):
                rule_id       = iss.get("rule_id", "")
                rule_name     = iss.get("rule_name") or rule_id
                dimension     = (iss.get("dimension") or "").title()
                inst_name     = (iss.get("institution_name") or le_book).title()
                rows_at_det   = iss.get("original_failing_rows") or iss.get("last_failing_rows") or iss.get("failing_rows", 0)
                # Issues reach write_resolved_zips only after two consecutive
                # clean scans (invalid==0 twice). mark_pending_resolution()
                # never zeroes failing_rows in dq_open_issues, so the field
                # still holds the original detection count — ignore it here.
                rows_remaining = 0
                rows_resolved  = rows_at_det or 0
                is_partial     = False
                resolution     = "Full"
                detected      = iss.get("detected_at", "")
                resolved      = iss.get("resolved_at", "")
                deadline      = iss.get("sla_deadline", "")
                on_time       = ("On Time" if resolved and deadline and resolved <= deadline
                                 else "Late" if resolved and deadline else "")

                sheet_title = re.sub(r"[\[\]:*?/\\]", " ", rule_id)[:31].strip() or "issue"
                ws = wb.create_sheet(title=sheet_title)

                # column widths
                from openpyxl.utils import get_column_letter
                for ci, w in enumerate(COL_W, start=1):
                    ws.column_dimensions[get_column_letter(ci)].width = w
                ws.row_dimensions[1].height = 28

                # header row
                hdr = []
                for ci, h in enumerate(HEADERS):
                    c = WriteOnlyCell(ws, value=h)
                    c.font      = HDR_FONT
                    c.alignment = HDR_ALIGN
                    if ci in RED_IDX:
                        c.fill = HDR_FILL_RED
                    elif ci in GREEN_IDX:
                        c.fill = HDR_FILL_GRN
                    else:
                        c.fill = HDR_FILL
                    hdr.append(c)
                ws.append(hdr)

                # rule description row
                desc_cell = WriteOnlyCell(ws, value=f"{rule_id}  —  {rule_name}")
                desc_cell.font = Font(size=10, italic=True, color="444444")
                ws.append([desc_cell])

                # data row
                values = [inst_name, table, rule_id, rule_name, dimension,
                          rows_at_det, rows_resolved, rows_remaining, resolution,
                          detected, resolved, deadline, on_time]
                cells = []
                for ci, val in enumerate(values):
                    c = WriteOnlyCell(ws, value=val)
                    if ci in RED_IDX and rows_remaining > 0:
                        c.fill = RED_DATA_FILL
                        c.font = RED_DATA_FONT
                    elif ci in PARTIAL_IDX and is_partial:
                        c.fill = AMB_DATA_FILL
                        c.font = AMB_DATA_FONT
                    elif ci in GREEN_IDX:
                        c.fill = GRN_FILL_ALT
                        c.font = GRN_FONT
                    else:
                        c.fill = DATA_FILL_ALT
                        c.font = DATA_FONT
                    cells.append(c)
                ws.append(cells)

                # append the original failing rows as a second sheet.
                # Primary source: evidence store (always available regardless
                # of detection month). Fallback: detection ZIP (legacy / pre-store runs).
                from storage.evidence_store import load_rows as _load_evidence
                ev_run_date, ev_rows = _load_evidence(le_book, rule_id, table)
                if ev_rows:
                    ev_ws = wb.create_sheet(title=(rule_id + " — Records")[:31])
                    if ev_rows:
                        ev_hdr = list(ev_rows[0].keys())
                        _all_meta = {**VAL_RULE_META, **ACC_RULE_META, **UNI_RULE_META}
                        rule_key_fields = set(_all_meta.get(rule_id, {}).get("fields", []))
                        ev_hdr_cells = []
                        for col_name in ev_hdr:
                            c = WriteOnlyCell(ev_ws, value=col_name)
                            c.font      = HDR_FONT
                            c.alignment = HDR_ALIGN
                            c.fill      = HDR_FILL_GRN if col_name in rule_key_fields else HDR_FILL
                            ev_hdr_cells.append(c)
                        ev_ws.append(ev_hdr_cells)
                        for ev_row in ev_rows:
                            ev_ws.append([
                                WriteOnlyCell(ev_ws, value=_coerce(ev_row.get(c)))
                                for c in ev_hdr
                            ])
                    log.debug("  evidence from store: %s/%s — %d rows (run %s)",
                              rule_id, table, len(ev_rows), ev_run_date)
                else:
                    log.debug("  no evidence rows for %s/%s/%s — records sheet omitted",
                              rule_id, table, le_book)

            fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
            os.close(fd)
            wb.save(tmp_path)
            tmp_files.append((f"{table}.xlsx", tmp_path))
            log.info("  resolved sheet built: %s  (%d rule(s))", table, len(issues))

        if tmp_files:
            ISSUE_REPORTS_DIR.mkdir(exist_ok=True)
            with zipfile.ZipFile(dst_zip, "w", zipfile.ZIP_DEFLATED) as zf:
                for arcname, tmp_path in tmp_files:
                    zf.write(tmp_path, arcname=arcname)
            log.info("Written resolved ZIP: %s  (%d table(s))", dst_zip.name, len(tmp_files))
            return True

    finally:
        for _, tmp_path in tmp_files:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass

    return False
