"""
dq_monthly_detection.py — Monthly DQ detection pipeline.

Runs a full-table scan (no date window) across all dimensions:
completeness, accuracy, timeliness, validity, uniqueness, relationship.

For every failing (rule, institution, table) combination an issue is written
to dq_open_issues with sla_deadline = today + 30 days.  Issues that were
open and now pass are marked pending_resolution (confirmed by the daily
resolution scanner).

Run this once per month.  The daily resolution pipeline
(dq_resolution_pipeline.py) then re-checks open issues each day and
marks them resolved when the underlying data is fixed, or overdue when the
SLA deadline passes without a fix.
"""
from __future__ import annotations
import argparse
import json
import logging
import os
import sys
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv
from sqlalchemy import text

from db_utils import CATEGORY_TYPES

SCRIPT_DIR = Path(__file__).parent
ISSUE_REPORTS_DIR = SCRIPT_DIR / "issue_reports"
HISTORY_FILE = SCRIPT_DIR / "dq_history.json"
sys.path.insert(0, str(SCRIPT_DIR))
logging.basicConfig(
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
    level=logging.INFO,
)
log = logging.getLogger("dq_monthly")

TABLES = [
    "accounts",
    "customers_expanded",
    "contracts_disburse",
    "contract_loans",
    "contract_schedules",
    "contracts_expanded",
    "loan_applications_2",
    "prev_loan_applications",
]

# db helper functions
def _db_columns(conn, schema: str, table: str) -> set[str]:
    rows = conn.execute(text("""
        SELECT column_name FROM information_schema.columns
        WHERE table_schema = :s AND table_name = :t
    """), {"s": schema, "t": table}).fetchall()
    return {r[0].lower() for r in rows}


# ── history aggregation (moved here from dq_pipeline_2m to decouple) ───────────
_SCORE_KEYS = {
    "comp": ("completeness", "overall_completeness_score", "completeness_score"),
    "acc":  ("accuracy",     "overall_accuracy_score",     "accuracy_score"),
    "tim":  ("timeliness",   "overall_timeliness_score",   "timeliness_score"),
    "val":  ("validity",     "overall_validity_score",     "validity_score"),
    "rel":  ("_rel",         "overall_ri_score",           "ri_score"),
}

DIMS = ["completeness", "accuracy", "timeliness", "validity"]


def _merge_rel(scores: dict) -> dict:
    """Average RI score into accuracy and drop the temporary _rel key."""
    if "_rel" not in scores:
        return scores
    acc = float(scores.get("accuracy") or 0.0)
    rel = float(scores.pop("_rel"))
    scores["accuracy"] = round((acc + rel) / 2, 2)
    return scores


def _inst_scores_from_report(report: dict, lb_score_key: str) -> dict[str, float]:
    """Average each le_book's per-table scores across all evaluated tables."""
    lb_table_scores: dict[str, list[float]] = {}
    for tbl_data in report.get("tables", {}).values():
        if tbl_data.get("status") != "evaluated":
            continue
        for lb, lb_data in tbl_data.get("le_book_breakdown", {}).items():
            s = lb_data.get(lb_score_key)
            if s is not None:
                lb_table_scores.setdefault(lb, []).append(float(s))
    return {
        lb: round(sum(scores) / len(scores), 2)
        for lb, scores in lb_table_scores.items()
        if scores
    }


def _build_history_entry(run_date: str, R: dict, categories: dict,
                         dup_counts: dict | None = None,
                         cat_dup_counts: dict | None = None) -> dict:
    """
    Aggregate engine results into a single history entry:
      overall        — one score per dimension (4 dims; RI averaged into accuracy)
      by_category    — average scores per category type (B, MF, OSACCO)
      by_institution — per-le_book scores across all 4 dimensions
    """
    # overall scores — extract all engines then merge RI into accuracy
    overall: dict[str, float] = {}
    for rkey, (dim, overall_key, _) in _SCORE_KEYS.items():
        esummary = (R.get(rkey) or {}).get("executive_summary") or {}
        # Relationship not run (no report) → don't fold a phantom 0 into accuracy
        # (that would halve it via _merge_rel). Skip rel entirely when absent.
        if rkey == "rel" and not esummary:
            continue
        overall[dim] = round(float(esummary.get(overall_key) or 0.0), 2)
    overall = _merge_rel(overall)

    # per-institution scores per dimension (extract then merge RI into accuracy)
    lb_dim_scores: dict[str, dict[str, float]] = {}
    for rkey, (dim, _, lb_score_key) in _SCORE_KEYS.items():
        inst_scores = _inst_scores_from_report(R.get(rkey) or {}, lb_score_key)
        for lb, score in inst_scores.items():
            lb_dim_scores.setdefault(lb, {})[dim] = score
    for lb in lb_dim_scores:
        lb_dim_scores[lb] = _merge_rel(lb_dim_scores[lb])

    # enrich with category metadata; compute per-institution overall (4 dims)
    _dups     = dup_counts or {}
    _cat_dups = cat_dup_counts or {}
    by_institution: dict = {}
    for lb, dim_scores in lb_dim_scores.items():
        cat_info    = categories.get(lb, {})
        inst_scores = [dim_scores[d] for d in DIMS if d in dim_scores]
        by_institution[lb] = {
            "name":               cat_info.get("name", lb),
            "category_type":      cat_info.get("category_type", ""),
            "overall":            round(sum(inst_scores) / len(inst_scores), 2) if inst_scores else 0.0,
            "customer_duplicates": _dups.get(lb, 0),
            **{d: dim_scores.get(d, 0.0) for d in DIMS},
        }

    # Institutions with dup counts but no windowed DQ data still need an entry
    for lb, dup in _dups.items():
        if lb not in by_institution:
            cat_info = categories.get(lb, {})
            by_institution[lb] = {
                "name":               cat_info.get("name", lb),
                "category_type":      cat_info.get("category_type", ""),
                "overall":            0.0,
                "customer_duplicates": dup,
                **{d: 0.0 for d in DIMS},
            }

    # by_category: average institution scores grouped by category_type
    cat_buckets: dict[str, list[dict]] = {}
    for inst_data in by_institution.values():
        ct = inst_data.get("category_type", "")
        if ct in CATEGORY_TYPES:
            cat_buckets.setdefault(ct, []).append(inst_data)

    by_category: dict = {}
    for ct, institutions in cat_buckets.items():
        cat_scores: dict[str, float] = {}
        for dim in DIMS:
            scores = [i[dim] for i in institutions if i.get(dim, 0) > 0]
            cat_scores[dim] = round(sum(scores) / len(scores), 2) if scores else 0.0
        # Use pre-computed DISTINCT count when available to avoid
        # double-counting customers with duplicates at multiple institutions.
        cat_scores["customer_duplicates"] = _cat_dups.get(
            ct, sum(i.get("customer_duplicates", 0) for i in institutions)
        )
        by_category[ct] = cat_scores

    # per-table failing rule counts and null column detail
    table_fail_counts = _table_fail_counts(R)
    table_null_cols   = _table_null_cols((R.get("comp") or {}))

    return {
        "date":              run_date,
        "overall":           overall,
        "by_category":       by_category,
        "by_institution":    by_institution,
        "table_fail_counts": table_fail_counts,
        "table_null_cols":   table_null_cols,
    }


def _table_fail_counts(R: dict) -> dict[str, int]:
    """Count failing rules (invalid > 0) per table across all engines."""
    from dq_rules import REL_RULE_META
    counts: dict[str, int] = {}

    # Completeness: 1 rule per table — fails when any null_cells > 0
    for tbl, tdata in (R.get("comp") or {}).get("tables", {}).items():
        if tdata.get("status") == "evaluated" and tdata.get("null_cells", 0) > 0:
            counts[tbl] = counts.get(tbl, 0) + 1

    # Accuracy, Timeliness, Validity, Uniqueness: rule-level breakdown
    for rkey in ("acc", "tim", "val", "uni"):
        for tbl, tdata in (R.get(rkey) or {}).get("tables", {}).items():
            if tdata.get("status") == "evaluated":
                for rdata in tdata.get("rules", {}).values():
                    if rdata.get("invalid", 0) > 0:
                        counts[tbl] = counts.get(tbl, 0) + 1

    # Relationship: credit child table
    for rid, rdata in (R.get("rel") or {}).get("rules", {}).items():
        if rdata.get("invalid", 0) > 0:
            child_tbl = REL_RULE_META.get(rid, {}).get("child_table")
            if child_tbl:
                counts[child_tbl] = counts.get(child_tbl, 0) + 1

    return counts


def _table_null_cols(comp_report: dict) -> dict[str, dict[str, int]]:
    """Return {table: {col: null_count}} for columns that have nulls."""
    result: dict[str, dict[str, int]] = {}
    for tbl, tdata in comp_report.get("tables", {}).items():
        if tdata.get("status") == "evaluated":
            nulls = {k: v for k, v in tdata.get("null_counts", {}).items() if v > 0}
            if nulls:
                result[tbl] = nulls
    return result


def _append_history(entry: dict) -> None:
    """Append (or replace same-date entry) in the history log."""
    history: list = []
    if HISTORY_FILE.exists():
        try:
            history = json.loads(HISTORY_FILE.read_text())
        except Exception:
            history = []
    # idempotent: replace if same date was already written today
    history = [e for e in history if e.get("date") != entry["date"]]
    history.append(entry)
    history.sort(key=lambda e: e.get("date", ""))
    HISTORY_FILE.write_text(json.dumps(history, indent=2, default=str))
    log.info("History log updated → %s  (%d entries total)", HISTORY_FILE, len(history))


def _zip_failing_rows_sql(engine, schema: str, table: str,
                          valid_le_books: frozenset, categories: dict,
                          month: str, limit: int = 0,
                          max_rows_per_sheet: int = 50000,
                          extra_where: str = "") -> None:
    """Stream FAILING rows (all dimensions) for `table` from SQL and write a
    per-institution {table}.xlsx into their monthly ZIP — ONE SHEET PER ISSUE
    (sheet name = issue type), with the affected column(s) highlighted red.

    Memory-safe: rows are capped per (institution, issue) server-side, streamed
    ORDER BY le_book, issue_type, and written through an openpyxl write_only
    workbook saved to a temp file per institution. Excel caps a sheet at ~1.05M
    rows; max_rows_per_sheet bounds it well below that (truncation noted on-sheet).
    """
    import os
    import re
    import tempfile
    import zipfile
    from datetime import date as _date, datetime as _datetime
    from decimal import Decimal
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Font, PatternFill
    from failing_rows_sql import build_failing_union

    RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    RED_FONT = Font(color="9C0006")
    HDR_FONT = Font(bold=True)
    HDR_RED  = Font(bold=True, color="9C0006")

    def _coerce(v):
        if v is None or isinstance(v, (str, int, float, bool, _date, _datetime)):
            return v
        if isinstance(v, Decimal):
            return float(v)
        return str(v)

    with engine.connect() as conn:
        existing = _db_columns(conn, schema, table)
        built = build_failing_union(schema, table, existing, valid_le_books, limit,
                                    extra_where=extra_where,
                                    per_issue_cap=max_rows_per_sheet + 1)
        if not built:
            return
        sql, out_cols, issue_cols = built

        # sheet columns = output cols minus issue_type; insert enrichment after le_book
        sheet_cols = [c for c in out_cols if c != "issue_type"]
        header: list[str] = []
        for c in sheet_cols:
            header.append(c)
            if c == "le_book":
                header += ["stakeholder_name", "category_type"]

        def _sheet_title(label: str, used: set) -> str:
            name = re.sub(r"[\[\]:*?/\\]", " ", label)[:31].strip() or "issue"
            base, i = name, 1
            while name.lower() in used:
                sfx = f" ({i})"
                name = base[:31 - len(sfx)] + sfx
                i += 1
            used.add(name.lower())
            return name

        result = conn.execution_options(stream_results=True).execute(text(sql))
        st = {"wb": None, "path": None, "ws": None, "issue": None,
              "affected": set(), "n": 0, "used": set(), "trunc": False}

        def _finish_sheet():
            if st["ws"] is not None and st["trunc"]:
                st["ws"].append([WriteOnlyCell(
                    st["ws"], value=f"... truncated at {max_rows_per_sheet:,} rows")])

        def _start_sheet(issue: str):
            _finish_sheet()
            ws = st["wb"].create_sheet(title=_sheet_title(issue, st["used"]))
            affected = set(issue_cols.get(issue, []))
            cells = []
            for col in header:
                cell = WriteOnlyCell(ws, value=col)
                if col in affected:
                    cell.font, cell.fill = HDR_RED, RED_FILL
                else:
                    cell.font = HDR_FONT
                cells.append(cell)
            ws.append(cells)
            st.update(ws=ws, issue=issue, affected=affected, n=0, trunc=False)

        def _write_row(m: dict, name: str, ctype: str):
            if st["n"] >= max_rows_per_sheet:
                st["trunc"] = True
                return
            affected = st["affected"]
            cells = []
            for col in sheet_cols:
                cell = WriteOnlyCell(st["ws"], value=_coerce(m.get(col)))
                if col in affected:
                    cell.fill, cell.font = RED_FILL, RED_FONT
                cells.append(cell)
                if col == "le_book":
                    cells.append(WriteOnlyCell(st["ws"], value=name))
                    cells.append(WriteOnlyCell(st["ws"], value=ctype))
            st["ws"].append(cells)
            st["n"] += 1

        def _new_workbook():
            fd, path = tempfile.mkstemp(suffix=".xlsx")
            os.close(fd)
            st.update(wb=Workbook(write_only=True), path=path, ws=None,
                      issue=None, used=set())

        def _close(lb: str):
            if st["wb"] is None:
                return
            _finish_sheet()
            st["wb"].save(st["path"])
            zip_path = ISSUE_REPORTS_DIR / f"{lb}_{month}.zip"
            with zipfile.ZipFile(zip_path, "a", zipfile.ZIP_DEFLATED) as zf:
                zf.write(st["path"], arcname=f"{table}.xlsx")
            os.unlink(st["path"])
            log.info("  ZIP %-6s  %s.xlsx", lb, table)
            st.update(wb=None, path=None, ws=None)

        cur_lb: str | None = None
        for row in result:
            m  = dict(row._mapping)
            lb = str(m["le_book"]).strip() if m.get("le_book") is not None else None
            if lb is None:
                continue
            issue = m.get("issue_type")
            info  = categories.get(lb, {})
            name  = info.get("name") or lb
            name  = name.title() if isinstance(name, str) else name
            ctype = info.get("category_type") or ""
            if lb != cur_lb:
                if cur_lb is not None:
                    _close(cur_lb)
                cur_lb = lb
                _new_workbook()
                _start_sheet(issue)
            elif issue != st["issue"]:
                _start_sheet(issue)
            _write_row(m, name, ctype)
        if cur_lb is not None:
            _close(cur_lb)


# orchestration

def month_filter(month: str) -> tuple[str, str]:
    """For a 'YYYY-MM' reporting month, return (extra_where, month_end_date):

      • extra_where   — a date_last_modified predicate scoping to that calendar
                        month (the same slice the --month detection run uses)
      • month_end_date — 'YYYY-MM-DD' month-end, used to tag the run's detected_at

    Shared by the --month CLI and the month-aware resolution scan so detection and
    resolution always re-evaluate the SAME data slice.
    """
    from calendar import monthrange
    y, m   = (int(p) for p in month.split("-")[:2])
    last   = monthrange(y, m)[1]
    ny, nm = (y + (m == 12), (m % 12) + 1)
    extra_where = (f"\"date_last_modified\" >= '{y:04d}-{m:02d}-01' "
                   f"AND \"date_last_modified\" < '{ny:04d}-{nm:02d}-01'")
    return extra_where, f"{y:04d}-{m:02d}-{last:02d}"


def run_monthly_detection(engine, schema: str, run_date: str,
                          tables: list[str] | None = None,
                          limit: int = 0,
                          le_books: frozenset | None = None,
                          extra_where: str = "") -> None:
    """
    Full-table DQ scan across all dimensions, run entirely in SQL (memory-safe).

    Pipeline:
      1. comp/acc/val/uni evaluate_from_sql with window_days=0 (full-table scan) → R
      2. detect_and_update_issues(R)  → dq_open_issues
      3. _build_history_entry / _append_history  → dq_history.json (dashboard trends)
      4. per-institution failing-row ZIPs streamed from SQL → issue_reports/

    tables:   restrict to a subset of tables (testing)
    limit:    row cap per table, 0 = no limit (testing)
    le_books: restrict to these institution codes (testing); default = all in-scope
    """
    import completeness_check as comp_eng
    import accuracy_check     as acc_eng
    import validity_check     as val_eng
    import uniqueness_check    as uni_eng
    from db_utils import get_valid_le_books, get_le_book_categories, customer_dup_counts
    from dq_issue_tracker import detect_and_update_issues, ensure_tables

    ensure_tables()

    log.info("Fetching institution metadata …")
    valid_le_books = le_books if le_books is not None else get_valid_le_books(engine, schema)
    categories     = get_le_book_categories(engine, schema)
    if le_books is not None:
        log.info("Institution filter: %s", ", ".join(sorted(valid_le_books)))

    if tables:
        log.info("Table filter: %s", ", ".join(tables))
    if limit:
        log.info("Row limit: %d per table (TEST MODE)", limit)

    # ── dimension scoring: pure-SQL engines, full-table scan (window_days=0) ───
    FULL_SCAN = 0
    wm: dict = {}
    log.info("Running completeness …")
    comp_report = comp_eng.evaluate_from_sql(engine, schema, valid_le_books, FULL_SCAN, wm,
                                             str(SCRIPT_DIR / "dq_report.json"),
                                             row_limit=limit, tables=tables, extra_where=extra_where)
    log.info("Running accuracy …")
    acc_report  = acc_eng.evaluate_from_sql(engine, schema, valid_le_books, FULL_SCAN, wm,
                                            str(SCRIPT_DIR / "dq_accuracy_report.json"),
                                            row_limit=limit, tables=tables, extra_where=extra_where)
    log.info("Running validity …")
    val_report  = val_eng.evaluate_from_sql(engine, schema, valid_le_books, FULL_SCAN, wm,
                                            str(SCRIPT_DIR / "dq_validity_report.json"),
                                            row_limit=limit, tables=tables, extra_where=extra_where)
    log.info("Running uniqueness …")
    uni_report  = uni_eng.evaluate_from_sql(engine, schema, valid_le_books, FULL_SCAN, wm,
                                            str(SCRIPT_DIR / "dq_uniqueness_report.json"),
                                            row_limit=limit, tables=tables, extra_where=extra_where)

    R = {"comp": comp_report, "acc": acc_report, "val": val_report, "uni": uni_report}

    log.info("Writing issues to tracker …")
    detect_and_update_issues(R, categories, run_date)

    log.info("Computing customer duplicate counts …")
    dup_counts, cat_dup_counts = customer_dup_counts(engine, schema, valid_le_books)
    log.info("  %d institution(s) with duplicate customers", len(dup_counts))

    log.info("Writing history entry …")
    entry = _build_history_entry(run_date, R, categories, dup_counts, cat_dup_counts)
    if entry.get("by_institution"):
        _append_history(entry)
    else:
        log.warning("No institution data in scope (empty run) — skipping history entry "
                    "for run_date=%s (would otherwise render as a 0%% cliff / blank overview)",
                    run_date)

    # ── per-institution failing-row ZIPs: streamed from SQL (memory-safe) ──────
    month = run_date[:7]
    ISSUE_REPORTS_DIR.mkdir(exist_ok=True)
    for old in ISSUE_REPORTS_DIR.glob(f"*_{month}.zip"):
        old.unlink()
        log.info("Removed stale ZIP: %s", old.name)
    log.info("Streaming per-institution failing-row ZIPs …")
    for table in (tables or TABLES):
        _zip_failing_rows_sql(engine, schema, table, valid_le_books, categories, month, limit,
                              extra_where=extra_where)

    log.info("Monthly detection complete — run_date=%s", run_date)


# cli

if __name__ == "__main__":
    load_dotenv(SCRIPT_DIR / ".env")

    parser = argparse.ArgumentParser(
        description="BNR DQ Monthly Detection."
    )
    parser.add_argument("--schema", default=os.environ.get("DQ_SCHEMA", "dqp"),
                        help="PostgreSQL schema (default: dqp)")
    parser.add_argument("--date", default=datetime.now().strftime("%Y-%m-%d"),
                        metavar="YYYY-MM-DD",
                        help="Run date used as detected_at (default: today)")
    parser.add_argument("--tables", nargs="+", default=None,
                        metavar="TABLE",
                        help="Restrict to these tables only (testing)")
    parser.add_argument("--limit", type=int, default=0,
                        help="Row cap per table, 0 = no limit (testing)")
    parser.add_argument("--le-book", nargs="+", default=None, metavar="CODE",
                        help="Restrict to these institution codes only (testing)")
    parser.add_argument("--month", default=None, metavar="YYYY-MM",
                        help="Scope detection to one reporting month by "
                             "date_last_modified (e.g. 2026-05). Sets detected_at "
                             "to the month-end date.")
    args = parser.parse_args()

    # --month: restrict data to that calendar month (by date_last_modified) and
    # date the run at month-end so issues are tagged to the reporting period.
    extra_where = ""
    if args.month:
        extra_where, args.date = month_filter(args.month)

    from db_utils import get_engine, build_connection_string
    engine = get_engine(build_connection_string())

    log.info("=" * 60)
    log.info("Monthly detection pipeline started  schema=%s  date=%s",
             args.schema, args.date)
    if args.tables:
        log.info("  Tables : %s", ", ".join(args.tables))
    if args.limit:
        log.info("  Limit  : %d rows/table  [TEST MODE]", args.limit)
    log.info("=" * 60)

    run_monthly_detection(engine, args.schema, args.date,
                          tables=args.tables, limit=args.limit,
                          le_books=frozenset(args.le_book) if args.le_book else None,
                          extra_where=extra_where)

    log.info("=" * 60)
    log.info("Done.")
    log.info("=" * 60)
    sys.exit(0)
