"""
Generate per-institution DQ issue XLSX reports from the 7-day DataFrame window.

Called from dq_pipeline_2m.py after the DQ engines finish.
Output: dashboard/reports/{le_book}_{name}.xlsx — one file per institution.

Each XLSX has 7 sheets:
  Info          — institution metadata + run timestamp
  Summary       — issue counts per dimension
  Completeness  — rows with NULL mandatory fields
  Accuracy      — rows failing accuracy rules
  Timeliness    — rows failing timeliness rules
  Validity      — rows failing validity rules
  Relationship  — child rows with orphaned FK values
"""
from __future__ import annotations

import logging
import re
from datetime import datetime
from pathlib import Path

import pandas as pd

import accuracy_check
import timeliness_check
import validity_check
from completeness_check import MANDATORY_COLUMNS
from accuracy_check    import RULE_META as ACC_META,  TABLE_RULES as ACC_TABLE_RULES
from timeliness_check  import RULE_META as TIM_META,  TABLE_RULES as TIM_TABLE_RULES
from validity_check    import RULE_META as VAL_META,  TABLE_RULES as VAL_TABLE_RULES
from relationship_check import RULE_META as REL_META

log = logging.getLogger("dq_issue_export")

# ── primary key per table ─────────────────────────────────────────────────────
TABLE_PK: dict[str, str] = {
    "customers_expanded":    "customer_id",
    "accounts":              "account_no",
    "contracts_expanded":    "contract_sequence_number",
    "contracts_disburse":    "contract_id",
    "contract_loans":        "contract_sequence_number",
    "contract_schedules":    "contract_sequence_number",
    "loan_applications_2":   "loan_application_id",
    "prev_loan_applications": "loan_application_id",
}

# ── human-readable context columns per table ──────────────────────────────────
# Each entry is (column_label, df_column_name).  These appear in the "Record Info"
# column so the reader can identify the record without a DB lookup.
TABLE_CONTEXT: dict[str, list[tuple[str, str]]] = {
    "customers_expanded":    [("Name",         "customer_name"),
                              ("Opened",        "customer_open_date")],
    "accounts":              [("Account Name",  "account_name"),
                              ("Customer ID",   "customer_id")],
    "contracts_expanded":    [("Customer ID",   "customer_id"),
                              ("Start Date",    "start_date")],
    "contracts_disburse":    [("Business Date", "business_date")],
    "contract_loans":        [("Perf. Class",   "performance_class"),
                              ("Created",       "date_creation")],
    "contract_schedules":    [("Schedule Date", "schedule_date")],
    "loan_applications_2":   [("Customer Name", "customer_name"),
                              ("Applied",       "application_date")],
    "prev_loan_applications": [("Business Date","business_date")],
}

# ── plain-English issue labels for RI rules ───────────────────────────────────
REL_ISSUE_LABELS: dict[str, str] = {
    "REL-001": "Account has no matching Customer",
    "REL-002": "Contract has no matching Customer",
    "REL-003": "Loan Application has no matching Customer",
    "REL-004": "Contract-Loan has no matching Contract",
    "REL-005": "Payment Schedule has no matching Contract",
    "REL-006": "Disbursement has no matching Contract",
    "REL-007": "Previous Application has no matching current Loan Application",
    "REL-008": "Contract references unknown Loan Application",
}

_HDR_FILL = "1A3A6B"   # BNR navy
_HDR_FONT = "FFFFFF"
_ALT_FILL = "F4F6F9"


def _pk(df: pd.DataFrame, table: str) -> pd.Series:
    col = TABLE_PK.get(table, "")
    return df[col].astype(str) if (col and col in df.columns) else pd.Series("—", index=df.index)

def _make_record_info(df: pd.DataFrame, table: str) -> pd.Series:
    """
    Return a Series of 'Label: value | Label: value' strings for each row.
    Uses TABLE_CONTEXT to pick which columns to surface per table.
    """
    context = [(lbl, col) for lbl, col in TABLE_CONTEXT.get(table, [])
               if col in df.columns]
    if not context:
        return pd.Series("—", index=df.index)

    pieces = []
    for lbl, col in context:
        s     = df[col].fillna("").astype(str).str.strip()
        valid = ~s.isin(["", "nan", "NaT", "None", "NaN"])
        pieces.append((lbl + ": " + s).where(valid, ""))

    combined = pd.concat(pieces, axis=1)
    return combined.apply(
        lambda row: " | ".join(v for v in row if v) or "—",
        axis=1,
    )


# ── per-dimension issue collectors ────────────────────────────────────────────

def _completeness_df(inst_frames: dict) -> pd.DataFrame:
    COLS = ["Row ID", "Record Info", "Table", "Field", "Issue"]
    chunks = []
    for table, df in inst_frames.items():
        if df.empty:
            continue
        pk   = _pk(df, table)
        info = _make_record_info(df, table)
        for col in MANDATORY_COLUMNS.get(table, []):
            if col not in df.columns:
                continue
            mask = df[col].isna()
            if not mask.any():
                continue
            chunks.append(pd.DataFrame({
                "Row ID":      pk[mask].values,
                "Record Info": info[mask].values,
                "Table":       table,
                "Field":       col,
                "Issue":       "Missing value (NULL)",
            }))
    return pd.concat(chunks, ignore_index=True) if chunks else pd.DataFrame(columns=COLS)


def _rule_issues_df(inst_frames: dict, engine_mod, table_rules: dict,
                    rule_meta: dict, score_key: str) -> pd.DataFrame:
    COLS = ["Row ID", "Record Info", "Table", "Rule", "Rule Name", "Field(s)", "Bad Value"]
    chunks = []
    for table, df in inst_frames.items():
        if df.empty:
            continue
        pk   = _pk(df, table)
        info = _make_record_info(df, table)
        for rule_id in table_rules.get(table, []):
            mask = engine_mod.run_rule_mask(rule_id, df)
            if not mask.any():
                continue
            meta   = rule_meta[rule_id]
            fields = [f for f in meta["fields"] if f in df.columns]
            bad    = df[mask]
            chunks.append(pd.DataFrame({
                "Row ID":      pk[mask].values,
                "Record Info": info[mask].values,
                "Table":       table,
                "Rule":        rule_id,
                "Rule Name":   meta["name"],
                "Field(s)":    ", ".join(meta["fields"]),
                "Bad Value":   (bad[fields].astype(str).agg(" | ".join, axis=1).values
                                if fields else ""),
            }))
    return pd.concat(chunks, ignore_index=True) if chunks else pd.DataFrame(columns=COLS)


def _relationship_df(le_book: str, dataframes: dict) -> pd.DataFrame:
    COLS = ["Row ID", "Record Info", "Issue",
            "Checked Table", "FK Column", "Orphaned Value (not found in parent)",
            "Expected In Table", "Rule"]
    chunks = []
    for rule_id, meta in REL_META.items():
        child_t  = meta["child_table"]
        child_c  = meta["child_col"]
        parent_t = meta["parent_table"]
        parent_c = meta["parent_col"]

        child_full = dataframes.get(child_t,  pd.DataFrame())
        parent_df  = dataframes.get(parent_t, pd.DataFrame())
        if child_full.empty or parent_df.empty:
            continue
        if child_c not in child_full.columns or parent_c not in parent_df.columns:
            continue

        # filter child to this institution; parent used unfiltered to avoid false orphans
        child_df = (child_full[child_full["le_book"].astype(str) == le_book].copy()
                    if "le_book" in child_full.columns else child_full.copy())
        child_df = child_df[child_df[child_c].notna()].copy()
        if child_df.empty:
            continue

        parent_keys     = frozenset(parent_df[parent_c].dropna().astype(str))
        child_df["_fk"] = child_df[child_c].astype(str)
        orphan_mask     = ~child_df["_fk"].isin(parent_keys)
        if not orphan_mask.any():
            continue

        bad    = child_df[orphan_mask]
        pk_col = TABLE_PK.get(child_t, child_c)
        info   = _make_record_info(bad, child_t)
        chunks.append(pd.DataFrame({
            "Row ID":                            (bad[pk_col].astype(str).values
                                                  if pk_col in bad.columns else "—"),
            "Record Info":                       info.values,
            "Issue":                             REL_ISSUE_LABELS.get(rule_id, meta["name"]),
            "Checked Table":                     child_t,
            "FK Column":                         child_c,
            "Orphaned Value (not found in parent)": bad["_fk"].values,
            "Expected In Table":                 parent_t,
            "Rule":                              rule_id,
        }))
    return pd.concat(chunks, ignore_index=True) if chunks else pd.DataFrame(columns=COLS)


# ── XLSX styling ──────────────────────────────────────────────────────────────

def _style_sheet(ws, n_cols: int) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils  import get_column_letter

    hdr_fill = PatternFill("solid", fgColor=_HDR_FILL)
    alt_fill = PatternFill("solid", fgColor=_ALT_FILL)
    wht_fill = PatternFill("solid", fgColor="FFFFFF")

    for cell in ws[1]:
        cell.font      = Font(bold=True, color=_HDR_FONT, size=10)
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = alt_fill if row_idx % 2 == 0 else wht_fill
        for cell in row:
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.font      = Font(size=9)

    for col_idx in range(1, n_cols + 1):
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or ""))
             for r in range(1, min(ws.max_row + 1, 102))),
            default=8,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 55)

    ws.freeze_panes  = "A2"
    ws.row_dimensions[1].height = 18


# ── single-institution XLSX writer ────────────────────────────────────────────

def _write_institution_xlsx(
    le_book: str,
    cat_info: dict,
    inst_frames: dict,
    dataframes: dict,
    output_dir: Path,
) -> None:
    inst_name = (cat_info.get("name") or le_book).title()
    safe      = re.sub(r"[^\w]", "_", inst_name)[:30].strip("_")
    path      = output_dir / f"{le_book}_{safe}.xlsx"

    comp_df = _completeness_df(inst_frames)
    acc_df  = _rule_issues_df(inst_frames, accuracy_check,   ACC_TABLE_RULES, ACC_META, "accuracy_score")
    tim_df  = _rule_issues_df(inst_frames, timeliness_check, TIM_TABLE_RULES, TIM_META, "timeliness_score")
    val_df  = _rule_issues_df(inst_frames, validity_check,   VAL_TABLE_RULES, VAL_META, "validity_score")
    rel_df  = _relationship_df(le_book, dataframes)

    run_ts  = datetime.now().strftime("%Y-%m-%d %H:%M UTC")
    info_df = pd.DataFrame({
        "Item":  ["Institution", "LE Book", "Category", "Generated At"],
        "Value": [inst_name, le_book, cat_info.get("category_type", ""), run_ts],
    })
    summary_df = pd.DataFrame({
        "Dimension":   ["Completeness", "Accuracy", "Timeliness", "Validity", "Accuracy (Referential)"],
        "Issue Count": [len(comp_df), len(acc_df), len(tim_df), len(val_df), len(rel_df)],
    })

    sheets = [
        ("Info",          info_df),
        ("Summary",       summary_df),
        ("Completeness",  comp_df),
        ("Accuracy",      acc_df),
        ("Timeliness",    tim_df),
        ("Validity",      val_df),
        ("Relationship",  rel_df),
    ]

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in sheets:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        wb = writer.book
        for sheet_name, df in sheets:
            _style_sheet(wb[sheet_name], len(df.columns))

    total = len(comp_df) + len(acc_df) + len(tim_df) + len(val_df) + len(rel_df)
    log.info("  ✓ %-6s  %-30s  %d issues → %s", le_book, inst_name[:30], total, path.name)


# ── main entry point ──────────────────────────────────────────────────────────

def export_institution_issues(
    dataframes: dict,
    le_book_categories: dict,
    valid_le_books: frozenset,
    output_dir: Path,
) -> None:
    """
    Write one XLSX per institution with row-level DQ issues across all 5 dimensions.

    Args:
        dataframes:         Pre-loaded DataFrames from the pipeline window.
        le_book_categories: {le_book: {"name": ..., "category_type": ...}}
        valid_le_books:     frozenset of le_book codes in scope.
        output_dir:         Directory to write XLSX files into (created if absent).
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(exist_ok=True)

    # Group DataFrames by le_book once — avoids scanning each table N times
    grouped: dict[str, dict[str, pd.DataFrame]] = {}
    for table, df in dataframes.items():
        if df.empty or "le_book" not in df.columns:
            continue
        for le_val, sub in df.groupby("le_book"):
            lb = str(le_val)
            if valid_le_books and lb not in valid_le_books:
                continue
            grouped.setdefault(lb, {})[table] = sub.reset_index(drop=True)

    if not grouped:
        log.warning("No institution data found — no XLSX files written.")
        return

    log.info("Writing issue reports for %d institution(s) → %s", len(grouped), output_dir)
    for le_book, inst_frames in sorted(grouped.items()):
        try:
            _write_institution_xlsx(
                le_book,
                le_book_categories.get(le_book, {}),
                inst_frames,
                dataframes,
                output_dir,
            )
        except Exception as exc:
            log.error("  ✗ %s — %s", le_book, exc)

    log.info("Institution issue reports complete.")
