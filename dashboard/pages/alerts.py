# Alerts page
from __future__ import annotations
import json
from dash import dcc, html
from dashboard.components import _dim_pill, _score_color
from dashboard.data import CATEGORIES_FILE, WATERMARK_FILE, _inst_scores, _today_entry
from dashboard.theme import (
    BG, BRAND, CARD, CAT_LABELS, C_GREEN, C_RED, DIMS, DIVIDER, MUTED,
    RESOLVED_GREEN, RESOLVED_GREEN_BG, TABLE_NAMES_PRETTY, TEXT, _URGENCY_COLORS,
)


def _table_dim_matrix(
    impact: dict[str, dict[str, dict]],
    pipeline_lbs: set[str] | None = None,
) -> html.Div:
    """Compact dimension-impact matrix: rows=tables, columns=active dimensions.

    impact — {table: {dimension: {failing_rows, inst_count}}}
    Returns an empty Div if no open-issue data is present.
    """
    if not impact:
        return html.Div()

    # Only show dimensions that have at least one issue
    all_dims_in_data = sorted({d for cells in impact.values() for d in cells})
    # Use canonical display order where possible
    _dim_order = ["completeness", "accuracy", "validity", "uniqueness", "timeliness"]
    active_dims = [d for d in _dim_order if d in all_dims_in_data] + \
                  [d for d in all_dims_in_data if d not in _dim_order]

    if not active_dims:
        return html.Div()

    # Sort tables by total failing rows descending
    def _table_total(cells: dict) -> int:
        return sum(c["failing_rows"] for c in cells.values())

    tables_sorted = sorted(impact.keys(), key=lambda t: _table_total(impact[t]), reverse=True)

    # Per-dimension max for scaling bars
    dim_max: dict[str, int] = {}
    for d in active_dims:
        dim_max[d] = max(
            (impact[t].get(d, {}).get("failing_rows", 0) for t in tables_sorted),
            default=1,
        ) or 1

    # Dimension color palette (matches existing theme use)
    _DIM_COLORS = {
        "completeness": "#753918",
        "accuracy":     "#2563EB",
        "validity":     "#7C3D1E",
        "uniqueness":   "#16A34A",
        "timeliness":   "#D97706",
    }
    _H = {"fontSize": "10px", "fontWeight": "900", "color": MUTED,
          "textTransform": "uppercase", "letterSpacing": "0.04em",
          "padding": "7px 14px", "whiteSpace": "nowrap"}

    CELL_W = "140px"

    # Header row
    hdr_cells = [html.Div("Table", style={**_H, "flex": "1"})]
    for d in active_dims:
        clr = _DIM_COLORS.get(d, MUTED)
        hdr_cells.append(html.Div(d[:4].title() + ".", style={
            **_H, "width": CELL_W, "textAlign": "center",
            "color": clr, "borderLeft": f"1px solid {DIVIDER}",
        }))

    hdr = html.Div(hdr_cells, style={
        "display": "flex", "background": BG,
        "borderRadius": "8px 8px 0 0",
        "borderBottom": f"2px solid {DIVIDER}",
    })

    # Data rows
    data_rows = []
    for i, tbl in enumerate(tables_sorted):
        cells_data = impact[tbl]
        tbl_label  = TABLE_NAMES_PRETTY.get(tbl, tbl.replace("_", " ").title())
        bg         = "rgba(244,246,249,0.7)" if i % 2 == 0 else CARD

        row_cells = [html.Div(
            html.Span(tbl_label, style={"fontSize": "12px", "fontWeight": "700",
                                        "color": TEXT}),
            style={"flex": "1", "padding": "10px 14px",
                   "display": "flex", "alignItems": "center"},
        )]

        for d in active_dims:
            clr  = _DIM_COLORS.get(d, MUTED)
            cell = cells_data.get(d)
            if cell and cell["failing_rows"] > 0:
                fr   = cell["failing_rows"]
                ni   = cell["inst_count"]
                pct  = min(100, round(fr / dim_max[d] * 100))
                bar  = html.Div(
                    html.Div(style={
                        "width":        f"{pct}%",
                        "height":       "5px",
                        "background":   clr,
                        "borderRadius": "3px",
                        "minWidth":     "4px",
                    }),
                    style={"height": "5px", "background": DIVIDER,
                           "borderRadius": "3px", "marginBottom": "4px"},
                )
                count_line = html.Div([
                    html.Span(f"{fr:,}", style={
                        "fontSize": "11px", "fontWeight": "700", "color": clr,
                    }),
                    html.Span(
                        f"  {ni} inst{'s' if ni != 1 else ''}",
                        style={"fontSize": "9px", "color": MUTED, "marginLeft": "4px"},
                    ),
                ], style={"display": "flex", "alignItems": "baseline"})
                cell_content = html.Div([bar, count_line], style={"width": "100%"})
            else:
                cell_content = html.Span("—", style={"color": DIVIDER,
                                                     "fontSize": "13px"})

            row_cells.append(html.Div(
                cell_content,
                style={
                    "width": CELL_W, "padding": "8px 14px",
                    "borderLeft": f"1px solid {DIVIDER}",
                    "display": "flex", "alignItems": "center",
                },
            ))

        data_rows.append(html.Div(row_cells, style={
            "display": "flex", "alignItems": "center",
            "background": bg, "borderBottom": f"1px solid {DIVIDER}",
        }))

    total_fr = sum(_table_total(impact[t]) for t in tables_sorted)
    n_tables  = len(tables_sorted)
    n_dims    = len(active_dims)

    return html.Div([
        html.Div([
            html.H3("Dimension Impact by Table", style={
                "fontSize": "14px", "fontWeight": "900", "color": TEXT,
                "margin": "0", "display": "inline",
            }),
            html.Span(
                f"  {n_tables} table{'s' if n_tables != 1 else ''} · "
                f"{n_dims} dimension{'s' if n_dims != 1 else ''} · "
                f"{total_fr:,} failing rows (open issues)",
                style={"fontSize": "11px", "color": MUTED, "marginLeft": "10px"},
            ),
        ], style={"marginBottom": "10px"}),
        html.Div([hdr, *data_rows], style={
            "background": CARD, "borderRadius": "8px",
            "border": f"1px solid {DIVIDER}",
            "overflowX": "auto",
        }),
    ], style={"marginBottom": "28px"})


def _issues_to_xlsx(issues: list) -> bytes:
    """Build an in-memory XLSX workbook from a list of resolved issue dicts."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import date as _date

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resolved Issues"

    hdr_fill       = PatternFill("solid", fgColor="753918")
    hdr_fill_green = PatternFill("solid", fgColor="16A34A")
    hdr_font       = Font(color="FFFFFF", bold=True, size=10)
    hdr_align      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    res_font       = Font(color="15803D", bold=True, size=10)
    res_fill       = PatternFill("solid", fgColor="DCFCE7")
    res_fill_alt   = PatternFill("solid", fgColor="BBF7D0")

    headers    = ["Institution", "Table", "Rule ID", "Dimension",
                  "Failing Rows", "Detected", "SLA Deadline", "Resolved", "Days to Fix", "On Time"]
    col_widths = [30, 28, 12, 14, 12, 12, 14, 12, 12, 10]
    resolved_ci = headers.index("Resolved") + 1

    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill_green if ci == resolved_ci else hdr_fill
        cell.alignment = hdr_align
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22

    for ri, iss in enumerate(issues, start=2):
        try:
            days_fix = (_date.fromisoformat(iss.get("resolved_at", "")) -
                        _date.fromisoformat(iss.get("detected_at", ""))).days
            days_str = f"{days_fix}d"
        except Exception:
            days_str = ""

        resolved_at  = iss.get("resolved_at", "")
        sla_deadline = iss.get("sla_deadline", "")
        on_time = ("On Time" if resolved_at and sla_deadline and resolved_at <= sla_deadline
                   else "Late" if resolved_at and sla_deadline else "")
        ws.append([
            (iss.get("institution_name") or iss["le_book"]).title(),
            iss.get("table_name", ""),
            iss.get("rule_id", ""),
            iss.get("dimension", "").title(),
            iss.get("last_failing_rows") or iss.get("failing_rows", ""),
            iss.get("detected_at", ""),
            sla_deadline,
            resolved_at,
            days_str,
            on_time,
        ])
        row_fill = PatternFill("solid", fgColor="F2EDE9") if ri % 2 == 0 else None
        for ci in range(1, len(headers) + 1):
            cell = ws.cell(row=ri, column=ci)
            if ci == resolved_ci:
                cell.fill = res_fill_alt if ri % 2 == 0 else res_fill
                cell.font = res_font
            elif row_fill:
                cell.fill = row_fill

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _freshness_section() -> html.Div:
    """Read watermark.json and render a per-table data freshness summary."""
    from datetime import date as _date
    import json as _json

    try:
        wm = _json.loads(WATERMARK_FILE.read_text())
    except Exception:
        return html.Div()

    today = _date.today()
    rows  = []
    for table, wm_val in sorted(wm.items()):
        try:
            last = _date.fromisoformat(str(wm_val)[:10])
            days = (today - last).days
        except Exception:
            continue

        if days <= 1:
            color, badge, bg = "#B8860B", "Fresh", "rgba(184,134,11,.07)"
        elif days <= 3:
            color, badge, bg = "#A0784A", f"{days}d ago", "rgba(160,120,74,.07)"
        else:
            color, badge, bg = "#7C3D1E", f"{days}d ago", "rgba(124,61,30,.07)"

        rows.append(html.Div([
            html.Span(table, style={
                "flex": "1", "fontSize": "12px", "color": TEXT,
                "fontFamily": "monospace", "padding": "7px 12px",
            }),
            html.Span(str(wm_val)[:10], style={
                "width": "110px", "fontSize": "11px", "color": MUTED,
                "padding": "7px 10px",
            }),
            html.Span(badge, style={
                "width": "80px", "fontSize": "11px", "fontWeight": "700",
                "color": color, "padding": "7px 10px", "textAlign": "center",
            }),
        ], style={
            "display": "flex", "alignItems": "center",
            "background": bg, "borderBottom": f"1px solid {DIVIDER}",
        }))

    any_stale = any(
        (today - _date.fromisoformat(str(v)[:10])).days > 1
        for v in wm.values()
        if str(v)[:10]
    )

    H = {"fontSize": "11px", "fontWeight": "900", "color": MUTED,
         "textTransform": "uppercase", "letterSpacing": "0.05em", "padding": "8px 12px"}
    header = html.Div([
        html.Span("Table",        style={**H, "flex": "1"}),
        html.Span("Last Updated", style={**H, "width": "110px"}),
        html.Span("Freshness",    style={**H, "width": "80px", "textAlign": "center"}),
    ], style={
        "display": "flex", "background": BG,
        "borderRadius": "8px 8px 0 0", "borderBottom": f"2px solid {DIVIDER}",
    })

    title_color = "#7C3D1E" if any_stale else "#B8860B"
    title_label = "⚠ Some tables have stale data" if any_stale else "✓ All tables up to date"

    return html.Div([
        html.Div([
            html.H3("Data Freshness", style={
                "fontSize": "14px", "fontWeight": "900", "color": TEXT,
                "margin": "0", "display": "inline",
            }),
            html.Span(f"  {title_label}", style={
                "fontSize": "11px", "color": title_color,
                "marginLeft": "10px", "fontWeight": "700",
            }),
        ], style={"marginBottom": "10px"}),
        html.Div([header, *rows], style={
            "background": CARD, "borderRadius": "8px",
            "border": f"1px solid {DIVIDER}", "marginBottom": "28px",
        }),
    ])


def _row_delta(current: int, previous: int | None, pending: bool = False,
               style: dict | None = None) -> html.Div:
    """Render failing-row count with a struck-through previous value when progress exists."""
    col = MUTED if pending else C_RED
    if previous and previous != current:
        return html.Div([
            html.Span(f"{previous:,}", style={"textDecoration": "line-through",
                                              "color": MUTED, "fontSize": "10px",
                                              "marginRight": "3px"}),
            html.Span(f"{current:,}", style={"fontWeight": "700", "color": col,
                                             "fontSize": "12px"}),
        ], style={"display": "flex", "alignItems": "center",
                  "justifyContent": "flex-end", **(style or {})})
    return html.Div(
        html.Span(f"{current:,}", style={"fontWeight": "700", "color": col,
                                         "fontSize": "12px"}),
        style={"textAlign": "right", **(style or {})},
    )


def _progress_bar(original: int, current: int,
                  le_book: str, rule_id: str, table: str) -> html.Div:
    """Mini inline progress bar for partial resolution.

    Shows: ████░░░ 82% fixed (450/550) + download remaining button.
    Only call when original > current (i.e. measurable progress).
    """
    fixed = max(0, original - current)
    pct   = min(100, round(fixed / original * 100)) if original else 0
    bar   = html.Div(
        html.Div(style={
            "width":        f"{pct}%",
            "height":       "4px",
            "background":   RESOLVED_GREEN,
            "borderRadius": "2px",
        }),
        style={
            "width":        "72px",
            "height":       "4px",
            "background":   DIVIDER,
            "borderRadius": "2px",
            "marginRight":  "6px",
            "flexShrink":   "0",
        },
    )
    label = html.Span(
        f"{pct}%  {fixed:,}/{original:,}",
        style={"fontSize": "9px", "color": RESOLVED_GREEN,
               "fontWeight": "700", "whiteSpace": "nowrap"},
    )
    dl_btn = html.A(
        "⬇ remaining",
        href=f"/download/remaining/{le_book}/{rule_id}/{table}",
        target="_blank",
        style={
            "fontSize":       "9px",
            "fontWeight":     "700",
            "color":          BRAND,
            "textDecoration": "none",
            "background":     "rgba(117,57,24,.08)",
            "border":         f"1px solid rgba(117,57,24,.25)",
            "borderRadius":   "3px",
            "padding":        "1px 6px",
            "marginLeft":     "8px",
            "whiteSpace":     "nowrap",
            "flexShrink":     "0",
        },
    )
    return html.Div(
        [bar, label, dl_btn],
        style={"display": "flex", "alignItems": "center", "padding": "4px 12px"},
    )


def _urgency_dot(band: str) -> html.Span:
    col = _URGENCY_COLORS.get(band, MUTED)
    label = {"new": "New", "attention": "Attention", "urgent": "Urgent",
             "critical": "Critical", "overdue": "⚠ Overdue"}.get(band, band.title())
    return html.Span([
        html.Span("●", style={"color": col, "fontSize": "9px", "marginRight": "4px"}),
        html.Span(label, style={"color": col, "fontSize": "11px", "fontWeight": "700"}),
    ])


def _build_table_issue_sections(issues_by_table: dict, status: str,
                                cat_filter: str = "") -> html.Div:
    """
    Render Alerts page: Table → Institution (mini score card) → Rules.
    cat_filter: '' = all, 'B' = banks, 'MF' = MFIs, 'SACCO' = SACCOs+OSACCOs.
    """
    if not issues_by_table:
        label = "open issues" if status == "open" else "resolved issues"
        return html.Div(f"No {label}.", style={"color": MUTED, "padding": "24px"})

    # load categories for type-lookup
    try:
        _cats = json.loads(CATEGORIES_FILE.read_text())
    except Exception:
        _cats = {}
    cat_type_map = {str(lb): (info.get("category_type") or "") for lb, info in _cats.items()}

    today_entry = _today_entry()
    _BO = ["new", "attention", "urgent", "critical", "overdue"]

    # issue_id → {cr_id, status} for active CRs
    try:
        from remediation.change_requests import get_issue_cr_map
        _cr_map = get_issue_cr_map()
    except Exception:
        _cr_map = {}

    _CR_STATUS_CLR = {
        "open":        "#2563EB",
        "in_progress": "#D97706",
        "submitted":   "#7C3D1E",
        "approved":    "#16A34A",
        "rejected":    "#DC2626",
    }

    def _worst(a: str, b: str) -> str:
        return a if _BO.index(a) >= _BO.index(b) else b

    sections = []
    for table, rules in issues_by_table.items():
        # ── pivot: table → {le_book: {name, rules, total_rows, worst_urgency}} ──
        inst_map: dict[str, dict] = {}
        for rule in rules:
            for inst in rule["institutions"]:
                lb  = inst["le_book"]
                ct  = cat_type_map.get(str(lb), "")
                # category filter  (SACCO covers OSACCO)
                if cat_filter:
                    if cat_filter == "SACCO" and ct not in ("SACCO", "OSACCO"):
                        continue
                    elif cat_filter != "SACCO" and ct != cat_filter:
                        continue
                if lb not in inst_map:
                    inst_map[lb] = {
                        "name":          inst["institution_name"],
                        "rules":         [],
                        "total_rows":    0,
                        "worst_urgency": "new",
                        "any_pending":   False,
                    }
                inst_map[lb]["rules"].append({
                    "rule_id":          rule["rule_id"],
                    "rule_name":        rule["rule_name"],
                    "dimension":        rule["dimension"],
                    "failing_rows":     inst["failing_rows"],
                    "last_failing_rows": inst.get("last_failing_rows"),
                    "urgency_band":     inst["urgency_band"],
                    "days_left":        inst["days_left"],
                    "issue_id":         inst.get("issue_id", ""),
                    "recurrence_count": inst.get("recurrence_count", 0),
                    "pending":          inst.get("pending", False),
                })
                inst_map[lb]["total_rows"]   += inst["failing_rows"]
                inst_map[lb]["worst_urgency"] = _worst(
                    inst_map[lb]["worst_urgency"], inst["urgency_band"])
                if inst.get("pending"):
                    inst_map[lb]["any_pending"] = True

        if not inst_map:
            continue

        total_rows  = sum(d["total_rows"] for d in inst_map.values())
        worst_tbl   = max(inst_map.values(), key=lambda x: _BO.index(x["worst_urgency"]))["worst_urgency"]
        band_col    = _URGENCY_COLORS.get(worst_tbl, MUTED)
        tbl_label   = TABLE_NAMES_PRETTY.get(table, table.replace("_", " ").title())
        n_inst      = len(inst_map)

        # ── table section header ─────────────────────────────────────────────
        tbl_header = html.Div([
            html.Span("●", style={"color": band_col, "fontSize": "10px",
                                   "marginRight": "8px"}),
            html.Span(tbl_label, style={
                "fontSize": "14px", "fontWeight": "900", "color": TEXT,
                "letterSpacing": "0.02em",
            }),
            html.Span(
                f"{n_inst} institution{'s' if n_inst != 1 else ''}",
                style={"fontSize": "11px", "color": MUTED, "marginLeft": "12px"},
            ),
            html.Span(f"{total_rows:,} failing rows", style={
                "fontSize": "11px", "color": C_RED, "fontWeight": "700",
                "marginLeft": "10px",
            }),
        ], style={
            "display": "flex", "alignItems": "center",
            "padding": "11px 18px",
            "background": BG,
            "borderLeft": f"4px solid {band_col}",
            "borderRadius": "6px 6px 0 0",
        })

        # ── institution rows (sorted by worst urgency then failing rows) ─────
        inst_divs = []
        sort_key  = lambda kv: (_BO.index(kv[1]["worst_urgency"]), -kv[1]["total_rows"])
        for lb, idata in sorted(inst_map.items(), key=sort_key, reverse=True):
            urg      = idata["worst_urgency"]
            urg_col  = _URGENCY_COLORS.get(urg, MUTED)
            cidx     = f"ait_{table}__{lb}".replace("-", "_")

            # mini dimension score chips
            i_scores  = _inst_scores(today_entry, lb)
            score_chips = []
            for dim in DIMS:
                s = float(i_scores.get(dim, 0))
                score_chips.append(html.Span([
                    html.Span(dim[:4].title(),
                              style={"fontSize": "9px", "color": MUTED,
                                     "marginRight": "2px"}),
                    html.Span(f"{s:.0f}%",
                              style={"fontSize": "11px", "fontWeight": "700",
                                     "color": _score_color(s)}),
                ], style={"marginRight": "12px", "whiteSpace": "nowrap"}))

            # badges: recurrence + pending-resolution
            badges = []
            max_recurrence = max(
                (r.get("recurrence_count", 0) for r in idata["rules"]), default=0
            )
            if max_recurrence > 0:
                badges.append(html.Span(
                    f"↺ #{max_recurrence + 1} recurrence",
                    style={"fontSize": "9px", "fontWeight": "700",
                           "color": C_RED, "background": "rgba(220,38,38,.10)",
                           "border": "1px solid rgba(220,38,38,.30)",
                           "borderRadius": "3px", "padding": "1px 5px",
                           "marginRight": "6px", "whiteSpace": "nowrap"},
                ))
            if idata.get("any_pending"):
                badges.append(html.Span(
                    "⟳ Confirming",
                    style={"fontSize": "9px", "fontWeight": "700",
                           "color": "#D97706", "background": "rgba(217,119,6,.10)",
                           "border": "1px solid rgba(217,119,6,.30)",
                           "borderRadius": "3px", "padding": "1px 5px",
                           "marginRight": "6px", "whiteSpace": "nowrap"},
                ))
            # CR badge: pick the most advanced active CR for any issue in this inst+table
            inst_issue_ids = [r.get("issue_id") for r in idata["rules"] if r.get("issue_id")]
            cr_info = next((_cr_map[iid] for iid in inst_issue_ids if iid in _cr_map), None)
            if cr_info:
                cr_clr = _CR_STATUS_CLR.get(cr_info["status"], MUTED)
                badges.append(html.Span(
                    f"CR: {cr_info['status'].replace('_', ' ').title()}",
                    style={"fontSize": "9px", "fontWeight": "700",
                           "color": cr_clr, "background": f"rgba(0,0,0,.05)",
                           "border": f"1px solid {cr_clr}",
                           "borderRadius": "3px", "padding": "1px 5px",
                           "marginRight": "6px", "whiteSpace": "nowrap"},
                ))

            inst_header = html.Div([
                _urgency_dot(urg),
                html.Span(idata["name"], style={
                    "fontSize": "13px", "fontWeight": "700",
                    "color": TEXT, "marginLeft": "8px",
                }),
                html.Span(f"({lb})", style={
                    "fontSize": "10px", "color": MUTED,
                    "fontFamily": "monospace", "marginLeft": "6px",
                    "marginRight": "10px",
                }),
                html.Div(badges, style={"display": "flex", "alignItems": "center",
                                        "flexShrink": "0"}),
                html.Span(style={"flex": "1"}),
                html.Div(score_chips, style={
                    "display": "flex", "alignItems": "center",
                    "marginRight": "18px",
                }),
                html.Span(f"{idata['total_rows']:,} rows", style={
                    "fontSize": "12px", "fontWeight": "700",
                    "color": C_RED, "marginRight": "14px",
                }),
                html.Div("⬇ ZIP",
                    id={"type": "open-issue-dl-btn", "index": lb},
                    n_clicks=0,
                    title=f"Download issue report for {idata['name']}",
                    style={
                        "cursor": "pointer", "background": BRAND, "color": CARD,
                        "fontSize": "10px", "fontWeight": "700",
                        "padding": "3px 10px", "borderRadius": "4px",
                        "userSelect": "none", "marginRight": "10px",
                        "flexShrink": "0",
                    },
                ),
                html.Div("▶", id={"type": "alert-inst-toggle", "index": cidx},
                         n_clicks=0,
                         style={"fontSize": "10px", "color": MUTED,
                                "cursor": "pointer", "userSelect": "none",
                                "width": "18px", "textAlign": "center"}),
            ], style={
                "display": "flex", "alignItems": "center",
                "padding": "9px 18px",
                "background": CARD,
                "borderBottom": f"1px solid {DIVIDER}",
                "borderLeft": f"3px solid {urg_col}",
                "cursor": "pointer",
            })

            # collapsed rule detail panel
            _RH = {"fontSize": "10px", "fontWeight": "900", "color": MUTED,
                   "textTransform": "uppercase", "letterSpacing": "0.04em",
                   "padding": "5px 12px", "whiteSpace": "nowrap"}
            rule_hdr = html.Div([
                html.Span("Rule",      style={**_RH, "width": "90px"}),
                html.Span("Issue",     style={**_RH, "flex": "1"}),
                html.Span("Dimension", style={**_RH, "width": "105px"}),
                html.Span("Rows",      style={**_RH, "width": "70px",
                                              "textAlign": "right"}),
                html.Span("Progress",  style={**_RH, "width": "200px"}),
                html.Span("SLA",       style={**_RH, "width": "90px",
                                              "textAlign": "right"}),
            ], style={"display": "flex", "background": BG,
                      "borderBottom": f"1px solid {DIVIDER}"})

            rule_rows = [rule_hdr]
            for j, r in enumerate(
                sorted(idata["rules"],
                       key=lambda x: _BO.index(x["urgency_band"]), reverse=True)
            ):
                dl  = r["days_left"]
                ov  = r["urgency_band"] == "overdue"
                dlc = C_RED if dl <= 5 or ov else TEXT
                dls = f"⚠ {abs(dl)}d over" if ov else f"{dl}d left"
                # per-rule: pending indicator + CR badge
                rule_badges = []
                if r.get("pending"):
                    rule_badges.append(html.Span(
                        "⟳", title="Pending full-scan confirmation",
                        style={"color": "#D97706", "fontSize": "10px",
                               "marginRight": "4px"},
                    ))
                rule_cr = _cr_map.get(r.get("issue_id", ""))
                if rule_cr:
                    rcr_clr = _CR_STATUS_CLR.get(rule_cr["status"], MUTED)
                    rule_badges.append(html.Span(
                        rule_cr["cr_id"],
                        style={"fontSize": "9px", "fontWeight": "700",
                               "color": rcr_clr, "marginRight": "4px"},
                    ))

                # progress bar cell — only when original > current
                orig = r.get("original_failing_rows")
                curr = r["failing_rows"]
                if orig and int(orig) > curr:
                    prog_cell = _progress_bar(int(orig), curr,
                                              lb, r["rule_id"], table)
                else:
                    prog_cell = html.Div(style={"width": "200px"})

                rule_rows.append(html.Div([
                    html.Span(r["rule_id"], style={
                        "width": "90px", "fontSize": "11px", "fontWeight": "900",
                        "color": BRAND, "fontFamily": "monospace",
                        "padding": "6px 12px",
                    }),
                    html.Div([
                        html.Span(r["rule_name"], style={
                            "fontSize": "12px", "color": TEXT,
                        }),
                        html.Div(rule_badges, style={"display": "inline-flex",
                                                      "alignItems": "center",
                                                      "marginLeft": "6px"}),
                    ], style={"flex": "1", "padding": "6px 12px",
                              "display": "flex", "alignItems": "center"}),
                    html.Div(_dim_pill(r["dimension"]),
                             style={"width": "105px", "padding": "4px 12px"}),
                    _row_delta(r["failing_rows"], r.get("last_failing_rows"),
                               pending=r.get("pending", False),
                               style={"width": "70px", "textAlign": "right",
                                      "padding": "6px 12px"}),
                    html.Div(prog_cell, style={"width": "200px"}),
                    html.Span(dls, style={
                        "width": "90px", "fontSize": "11px", "fontWeight": "700",
                        "color": dlc, "textAlign": "right", "padding": "6px 12px",
                    }),
                ], style={
                    "display": "flex", "alignItems": "center",
                    "background": "rgba(244,246,249,0.8)" if j % 2 == 0 else CARD,
                    "borderBottom": f"1px solid {DIVIDER}",
                    "opacity": "0.75" if r.get("pending") else "1",
                }))

            inst_divs.append(html.Div([
                inst_header,
                html.Div(
                    html.Div(rule_rows),
                    id={"type": "alert-inst-collapse", "index": cidx},
                    style={"display": "none"},
                ),
            ]))

        sections.append(html.Div(
            [tbl_header, *inst_divs],
            style={"marginBottom": "16px", "border": f"1px solid {DIVIDER}",
                   "borderRadius": "6px", "overflow": "hidden"},
        ))

    if not sections:
        return html.Div("No issues for the selected category.",
                        style={"color": MUTED, "padding": "24px"})
    return html.Div(sections)


def _build_resolved_rows(issues: list) -> html.Div:
    """Flat list for resolved issues."""
    from datetime import date as _date
    today = _date.today()
    if not issues:
        return html.Div("No resolved issues.", style={"color": MUTED, "padding": "20px"})

    H = {"fontSize": "11px", "fontWeight": "900", "color": MUTED,
         "textTransform": "uppercase", "letterSpacing": "0.05em", "padding": "8px 10px"}
    hdr = html.Div([
        html.Span("Institution", style={**H, "flex": "1"}),
        html.Span("Table",       style={**H, "width": "160px"}),
        html.Span("Rule",        style={**H, "width": "80px"}),
        html.Span("Dimension",   style={**H, "width": "100px"}),
        html.Span("Detected",    style={**H, "width": "96px"}),
        html.Span("Resolved",    style={**H, "width": "96px"}),
        html.Span("Days to Fix", style={**H, "width": "80px", "textAlign": "center"}),
        html.Span("Timing",      style={**H, "width": "76px", "textAlign": "center"}),
    ], style={"display": "flex", "background": BG,
              "borderRadius": "8px 8px 0 0", "borderBottom": f"2px solid {DIVIDER}"})

    rows = []
    for i, iss in enumerate(issues):
        detected  = iss.get("detected_at", "—")
        resolved  = iss.get("resolved_at",  "—")
        deadline  = iss.get("sla_deadline", "")
        try:
            days_fix = (_date.fromisoformat(resolved) - _date.fromisoformat(detected)).days
            fix_str  = f"{days_fix}d"
            fix_clr  = C_GREEN if days_fix <= 7 else (C_RED if days_fix >= 20 else TEXT)
        except Exception:
            fix_str, fix_clr = "—", MUTED
        try:
            on_time = _date.fromisoformat(resolved) <= _date.fromisoformat(deadline)
            timing_label = "On Time"
            timing_color = C_GREEN
        except Exception:
            on_time = None
            timing_label, timing_color = "—", MUTED
        if on_time is False:
            timing_label = "Late"
            timing_color = C_RED
        name = (iss.get("institution_name") or iss["le_book"]).title()
        rows.append(html.Div([
            html.Div([html.Span(name, style={"fontSize": "12px", "color": TEXT})],
                     style={"flex": "1", "padding": "7px 10px"}),
            html.Span(iss["table_name"],       style={"width": "160px", "fontSize": "11px", "color": MUTED, "padding": "7px 10px"}),
            html.Span(iss["rule_id"],           style={"width": "80px",  "fontSize": "11px", "fontWeight": "700", "color": TEXT, "padding": "7px 10px"}),
            html.Span(iss["dimension"].title(), style={"width": "100px", "fontSize": "11px", "color": MUTED, "padding": "7px 10px"}),
            html.Span(detected,                style={"width": "96px",  "fontSize": "11px", "color": MUTED, "padding": "7px 10px"}),
            html.Span(resolved,                style={"width": "96px",  "fontSize": "11px", "color": C_GREEN, "padding": "7px 10px"}),
            html.Span(fix_str,                 style={"width": "80px",  "fontSize": "12px", "fontWeight": "700", "color": fix_clr, "textAlign": "center", "padding": "7px 10px"}),
            html.Span(timing_label,            style={"width": "76px",  "fontSize": "11px", "fontWeight": "700", "color": timing_color, "textAlign": "center", "padding": "7px 10px"}),
        ], style={
            "display": "flex", "alignItems": "center",
            "background": "#C9956C" if i % 2 == 0 else BG,
            "borderBottom": f"1px solid {DIVIDER}",
        }))
    return html.Div([hdr, *rows], style={
        "background": CARD, "borderRadius": "8px",
        "border": f"1px solid {DIVIDER}", "marginBottom": "8px",
    })


def _build_resolved_by_institution(issues: list, cat_filter: str = "") -> html.Div:
    """Resolved issues grouped by institution with collapsible rule detail and per-inst ZIP download."""
    from datetime import date as _date
    from collections import defaultdict

    if not issues:
        return html.Div("No resolved issues.", style={"color": MUTED, "padding": "24px"})

    try:
        _cats = json.loads(CATEGORIES_FILE.read_text())
    except Exception:
        _cats = {}

    if cat_filter:
        _SACCO_TYPES = {"SACCO", "OSACCO"}
        def _matches(lb: str) -> bool:
            ct = (_cats.get(str(lb), {}).get("category_type") or "").upper()
            return ct in _SACCO_TYPES if cat_filter == "SACCO" else ct == cat_filter
        issues = [i for i in issues if _matches(i["le_book"])]

    if not issues:
        return html.Div("No resolved issues for the selected category.",
                        style={"color": MUTED, "padding": "24px"})

    by_inst: dict[str, list] = defaultdict(list)
    for iss in issues:
        by_inst[iss["le_book"]].append(iss)

    inst_order = sorted(
        by_inst.keys(),
        key=lambda lb: max((i.get("resolved_at") or "") for i in by_inst[lb]),
        reverse=True,
    )

    H = {"fontSize": "11px", "fontWeight": "900", "color": MUTED,
         "textTransform": "uppercase", "letterSpacing": "0.05em", "padding": "8px 10px"}

    _CAT_CLR = {"B": "#2563EB", "MF": "#753918", "SACCO": "#16A34A", "OSACCO": "#16A34A"}

    cards = []
    for cidx, lb in enumerate(inst_order):
        inst_issues = sorted(by_inst[lb],
                             key=lambda x: x.get("resolved_at") or "", reverse=True)
        inst_info   = _cats.get(str(lb), {})
        name        = (inst_issues[0].get("institution_name") or lb).title()
        cat_type    = (inst_info.get("category_type") or "").upper()
        tables      = sorted({i["table_name"] for i in inst_issues})
        latest_res  = max((i.get("resolved_at") or "") for i in inst_issues)

        fix_days = []
        for i in inst_issues:
            try:
                fix_days.append(
                    (_date.fromisoformat(i["resolved_at"]) -
                     _date.fromisoformat(i["detected_at"])).days
                )
            except Exception:
                pass
        avg_fix = f"{sum(fix_days) // len(fix_days)}d avg" if fix_days else "—"

        cat_clr = _CAT_CLR.get(cat_type, MUTED)

        header = html.Div([
            html.Div([
                html.Span("▶",
                          id={"type": "res-inst-toggle", "index": cidx},
                          n_clicks=0,
                          style={"cursor": "pointer", "color": MUTED, "fontSize": "10px",
                                 "marginRight": "10px", "userSelect": "none",
                                 "flexShrink": "0"}),
                html.Span(name, style={"fontWeight": "700", "fontSize": "13px",
                                       "color": TEXT}),
                html.Span(cat_type, style={
                    "fontSize": "10px", "fontWeight": "700", "color": cat_clr,
                    "border": f"1px solid {cat_clr}", "borderRadius": "3px",
                    "padding": "1px 6px", "marginLeft": "8px",
                }),
                html.Span(f"le_book: {lb}", style={
                    "fontSize": "10px", "color": MUTED, "marginLeft": "8px",
                }),
            ], style={"display": "flex", "alignItems": "center", "flex": "1",
                      "minWidth": "0"}),

            html.Div([
                html.Span(f"{len(inst_issues)} issue{'s' if len(inst_issues) != 1 else ''}",
                          style={"fontSize": "11px", "fontWeight": "700",
                                 "color": C_GREEN, "whiteSpace": "nowrap"}),
                html.Span("  ·  ", style={"color": DIVIDER, "fontSize": "11px"}),
                html.Span(
                    ", ".join(t.replace("_", " ") for t in tables[:3])
                    + ("…" if len(tables) > 3 else ""),
                    style={"fontSize": "11px", "color": MUTED, "whiteSpace": "nowrap"},
                ),
                html.Span("  ·  ", style={"color": DIVIDER, "fontSize": "11px"}),
                html.Span(f"resolved {latest_res}",
                          style={"fontSize": "11px", "color": C_GREEN,
                                 "whiteSpace": "nowrap"}),
                html.Span("  ·  ", style={"color": DIVIDER, "fontSize": "11px"}),
                html.Span(avg_fix, style={"fontSize": "11px", "color": MUTED,
                                          "whiteSpace": "nowrap"}),
            ], style={"display": "flex", "alignItems": "center", "gap": "2px",
                      "flexWrap": "nowrap", "overflow": "hidden"}),

            html.Div("⬇ ZIP",
                     id={"type": "res-dl-btn", "index": lb},
                     n_clicks=0,
                     style={"cursor": "pointer", "background": BRAND, "color": CARD,
                            "fontSize": "10px", "fontWeight": "700",
                            "padding": "4px 12px", "borderRadius": "4px",
                            "userSelect": "none", "marginLeft": "16px",
                            "flexShrink": "0"}),
        ], style={
            "display": "flex", "alignItems": "center",
            "padding": "10px 14px", "background": CARD, "gap": "12px",
        })

        tbl_hdr = html.Div([
            html.Span("Table",        style={**H, "flex": "1"}),
            html.Span("Rule",         style={**H, "width": "90px"}),
            html.Span("Dimension",    style={**H, "width": "110px"}),
            html.Span("Failing Rows", style={**H, "width": "100px",
                                             "textAlign": "right"}),
            html.Span("Detected",     style={**H, "width": "96px"}),
            html.Span("Resolved",     style={**H, "width": "96px"}),
            html.Span("Days to Fix",  style={**H, "width": "80px",
                                             "textAlign": "center"}),
            html.Span("Timing",       style={**H, "width": "70px",
                                             "textAlign": "center"}),
        ], style={"display": "flex", "background": BG,
                  "borderTop": f"1px solid {DIVIDER}"})

        detail_rows = []
        for j, iss in enumerate(inst_issues):
            try:
                days_fix = (_date.fromisoformat(iss["resolved_at"]) -
                            _date.fromisoformat(iss["detected_at"])).days
                fix_str  = f"{days_fix}d"
                fix_clr  = C_GREEN if days_fix <= 7 else (C_RED if days_fix >= 20 else TEXT)
            except Exception:
                fix_str, fix_clr = "—", MUTED
            try:
                on_time = (_date.fromisoformat(iss["resolved_at"]) <=
                           _date.fromisoformat(iss["sla_deadline"]))
                timing_label = "On Time"
                timing_color = C_GREEN
            except Exception:
                on_time = None
                timing_label, timing_color = "—", MUTED
            if on_time is False:
                timing_label = "Late"
                timing_color = C_RED

            detail_rows.append(html.Div([
                html.Span(iss["table_name"],
                          style={"flex": "1", "fontSize": "11px", "color": TEXT,
                                 "padding": "6px 10px", "fontFamily": "monospace"}),
                html.Span(iss["rule_id"],
                          style={"width": "90px", "fontSize": "11px", "fontWeight": "700",
                                 "color": TEXT, "padding": "6px 10px"}),
                html.Span(iss["dimension"].title(),
                          style={"width": "110px", "fontSize": "11px",
                                 "color": MUTED, "padding": "6px 10px"}),
                html.Span(f"{(iss.get('last_failing_rows') or iss.get('failing_rows', 0)):,}",
                          style={"width": "100px", "fontSize": "11px", "color": MUTED,
                                 "textAlign": "right", "padding": "6px 10px"}),
                html.Span(iss.get("detected_at", "—"),
                          style={"width": "96px", "fontSize": "11px",
                                 "color": MUTED, "padding": "6px 10px"}),
                html.Span(iss.get("resolved_at", "—"),
                          style={"width": "96px", "fontSize": "11px",
                                 "color": C_GREEN, "padding": "6px 10px"}),
                html.Span(fix_str,
                          style={"width": "80px", "fontSize": "12px", "fontWeight": "700",
                                 "color": fix_clr, "textAlign": "center",
                                 "padding": "6px 10px"}),
                html.Span(timing_label,
                          style={"width": "70px", "fontSize": "11px", "fontWeight": "700",
                                 "color": timing_color, "textAlign": "center",
                                 "padding": "6px 10px"}),
            ], style={
                "display": "flex", "alignItems": "center",
                "background": BG if j % 2 == 0 else CARD,
                "borderTop": f"1px solid {DIVIDER}",
            }))

        collapse = html.Div(
            [tbl_hdr, *detail_rows],
            id={"type": "res-inst-collapse", "index": cidx},
            style={"display": "none"},
        )

        cards.append(html.Div(
            [header, collapse],
            style={
                "background": CARD, "border": f"1px solid {DIVIDER}",
                "borderRadius": "8px", "marginBottom": "8px",
                "overflow": "hidden",
            },
        ))

    return html.Div(cards)


def _build_issue_rows(issues: list, status: str) -> html.Div:
    """Render issue table rows for the given status category."""
    from datetime import date as _date
    today = _date.today()

    if not issues:
        label = {"open": "open issues", "resolved": "resolved issues"}.get(status, "issues")
        return html.Div(f"No {label}.", style={"color": MUTED, "padding": "20px"})

    H = {"fontSize": "11px", "fontWeight": "900", "color": MUTED,
         "textTransform": "uppercase", "letterSpacing": "0.05em", "padding": "8px 10px"}

    if status == "resolved":
        hdr = html.Div([
            html.Span("Institution",  style={**H, "flex": "1"}),
            html.Span("Table",        style={**H, "width": "160px"}),
            html.Span("Rule",         style={**H, "width": "90px"}),
            html.Span("Dimension",    style={**H, "width": "100px"}),
            html.Span("Detected",     style={**H, "width": "96px"}),
            html.Span("Resolved",     style={**H, "width": "96px"}),
            html.Span("Days to Fix",  style={**H, "width": "80px", "textAlign": "center"}),
        ], style={"display": "flex", "background": BG,
                  "borderRadius": "8px 8px 0 0", "borderBottom": f"2px solid {DIVIDER}"})
    else:
        hdr = html.Div([
            html.Span("Institution",  style={**H, "flex": "1"}),
            html.Span("Table",        style={**H, "width": "160px"}),
            html.Span("Rule",         style={**H, "width": "90px"}),
            html.Span("Dimension",    style={**H, "width": "100px"}),
            html.Span("Failing Rows", style={**H, "width": "90px", "textAlign": "right"}),
            html.Span("Detected",     style={**H, "width": "90px"}),
            html.Span("Deadline",     style={**H, "width": "90px"}),
            html.Span("Remaining",    style={**H, "width": "76px", "textAlign": "center"}),
            html.Span("Notify",       style={**H, "width": "52px", "textAlign": "center"}),
        ], style={"display": "flex", "background": BG,
                  "borderRadius": "8px 8px 0 0", "borderBottom": f"2px solid {DIVIDER}"})

    rows = []
    for i, iss in enumerate(issues):
        band  = iss.get("urgency_band", "new")
        lb    = iss["le_book"]
        name  = (iss.get("institution_name") or lb).title()

        if status == "resolved":
            row_clr = RESOLVED_GREEN
            bg      = "rgba(22,163,74,0.15)" if i % 2 == 0 else RESOLVED_GREEN_BG
        else:
            row_clr = _URGENCY_COLORS.get(band, MUTED)
            bg      = "#C9956C" if i % 2 == 0 else BG

        clr = row_clr
        inst_cell = html.Div([
            html.Span("●", style={"color": clr, "fontSize": "9px", "marginRight": "5px"}),
            html.Span(name, style={"fontSize": "12px", "color": TEXT}),
        ], style={"flex": "1", "display": "flex", "alignItems": "center",
                  "padding": "7px 10px", "borderLeft": f"3px solid {clr}"})

        if status == "resolved":
            detected  = iss.get("detected_at", "—")
            resolved  = iss.get("resolved_at",  "—")
            try:
                days_fix = (_date.fromisoformat(resolved) - _date.fromisoformat(detected)).days
                fix_str  = f"{days_fix}d"
                fix_clr  = C_GREEN if days_fix <= 7 else (C_RED if days_fix >= 20 else TEXT)
            except Exception:
                fix_str, fix_clr = "—", MUTED
            row_children = [
                inst_cell,
                html.Span(iss["table_name"],       style={"width": "160px", "fontSize": "11px", "color": MUTED, "padding": "7px 10px"}),
                html.Span(iss["rule_id"],           style={"width": "90px",  "fontSize": "11px", "fontWeight": "700", "color": TEXT, "padding": "7px 10px"}),
                html.Span(iss["dimension"].title(), style={"width": "100px", "fontSize": "11px", "color": MUTED, "padding": "7px 10px"}),
                html.Span(detected,                 style={"width": "96px",  "fontSize": "11px", "color": MUTED, "padding": "7px 10px"}),
                html.Span(resolved,                 style={"width": "96px",  "fontSize": "11px", "color": C_GREEN, "padding": "7px 10px"}),
                html.Span(fix_str,                  style={"width": "80px",  "fontSize": "12px", "fontWeight": "700", "color": fix_clr, "textAlign": "center", "padding": "7px 10px"}),
            ]
        else:
            try:
                days_left = (_date.fromisoformat(iss["sla_deadline"]) - today).days
            except Exception:
                days_left = None
            is_overdue  = band == "overdue" or (isinstance(days_left, int) and days_left < 0)
            if is_overdue:
                over_days  = abs(days_left) if isinstance(days_left, int) else "?"
                days_str   = f"⚠ {over_days}d over"
                days_color = _URGENCY_COLORS["overdue"]
            elif isinstance(days_left, int) and days_left <= 5:
                days_str   = f"{days_left}d left"
                days_color = C_RED
            elif isinstance(days_left, int):
                days_str   = f"{days_left}d left"
                days_color = TEXT
            else:
                days_str   = "?"
                days_color = MUTED
            row_children = [
                inst_cell,
                html.Span(iss["table_name"],           style={"width": "160px", "fontSize": "11px", "color": MUTED, "padding": "7px 10px"}),
                html.Span(iss["rule_id"],              style={"width": "90px",  "fontSize": "11px", "fontWeight": "700", "color": TEXT, "padding": "7px 10px"}),
                html.Span(iss["dimension"].title(),    style={"width": "100px", "fontSize": "11px", "color": MUTED, "padding": "7px 10px"}),
                html.Span(f"{iss['failing_rows']:,}",  style={"width": "90px",  "fontSize": "12px", "fontWeight": "700", "color": TEXT, "textAlign": "right", "padding": "7px 10px"}),
                html.Span(iss["detected_at"],          style={"width": "90px",  "fontSize": "11px", "color": MUTED, "padding": "7px 10px"}),
                html.Span(iss["sla_deadline"],         style={"width": "90px",  "fontSize": "11px", "color": C_RED if is_overdue else MUTED, "padding": "7px 10px"}),
                html.Span(days_str,                    style={"width": "76px",  "fontSize": "12px", "fontWeight": "700", "color": days_color, "textAlign": "center", "padding": "7px 10px"}),
                html.Div("🔔",
                    id={"type": "notify-btn", "index": lb},
                    n_clicks=0,
                    title=f"Send reminder to {name}",
                    style={"width": "52px", "textAlign": "center", "fontSize": "14px",
                           "cursor": "pointer", "color": clr, "padding": "7px 0",
                           "userSelect": "none"},
                ),
            ]

        rows.append(html.Div(row_children, style={
            "display": "flex", "alignItems": "center",
            "background": bg, "borderBottom": f"1px solid {DIVIDER}",
        }))

    return html.Div([hdr, *rows], style={
        "background": CARD, "borderRadius": "8px",
        "border": f"1px solid {DIVIDER}", "marginBottom": "8px",
    })


def _alerts_page(cat: str = "") -> html.Div:
    cat_label = CAT_LABELS.get(cat, "All Categories") if cat else "All Categories"

    header_row = html.Div([
        html.Div([
            html.H2("Resolved Issues", style={
                "fontSize": "18px", "fontWeight": "900", "color": TEXT,
                "margin": "0 0 2px",
            }),
            html.Span(f"Showing: {cat_label}", style={
                "fontSize": "11px", "color": MUTED,
            }),
        ]),
        html.Div("⬇ Export XLSX", id="issues-download-btn", n_clicks=0,
                 style={"cursor": "pointer", "background": BRAND, "color": CARD,
                        "fontSize": "11px", "fontWeight": "700",
                        "padding": "6px 14px", "borderRadius": "5px",
                        "userSelect": "none", "whiteSpace": "nowrap",
                        "alignSelf": "center"}),
    ], style={"display": "flex", "justifyContent": "space-between",
              "alignItems": "flex-start", "marginBottom": "20px"})

    return html.Div([
        header_row,
        html.Div(id="alerts-summary-bar", style={"marginBottom": "24px"}),
        html.Div(id="table-dim-matrix",   style={"marginBottom": "0"}),
        # Hidden RadioItems — preserves callback wiring; value drives filtered data
        dcc.RadioItems(
            id="alerts-cat-filter",
            options=[
                {"label": "All", "value": ""},
                {"label": "B",   "value": "B"},
                {"label": "MF",  "value": "MF"},
                {"label": "S",   "value": "SACCO"},
            ],
            value=cat or "",
            style={"display": "none"},
        ),
        html.Div(id="issue-list"),
        html.Div(id="notify-feedback"),
    ], style={"padding": "28px 32px", "maxWidth": "1380px", "margin": "0 auto"})
