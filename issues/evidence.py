"""Parse issue-report ZIPs to determine which rule_ids have evidence in the latest report."""
from __future__ import annotations

import logging
import re
import zipfile
from io import BytesIO
from pathlib import Path
from typing import Optional

log = logging.getLogger(__name__)

ISSUE_REPORTS_DIR = Path(__file__).resolve().parents[1] / "issue_reports"

# One COMP rule per table — inverse of _COMP_TABLE_NAMES in dq/rules/completeness.py
_TABLE_TO_COMP: dict[str, str] = {
    "customers_expanded":     "COMP-001",
    "accounts":               "COMP-002",
    "contracts_disburse":     "COMP-003",
    "contract_loans":         "COMP-004",
    "contract_schedules":     "COMP-005",
    "contracts_expanded":     "COMP-006",
    "loan_applications_2":    "COMP-007",
    "prev_loan_applications": "COMP-008",
}

_RID_RE = re.compile(r"^([A-Z]+-\d+)")

# Per-ZIP cache: zip_path -> (mtime, frozenset[rule_ids])
_zip_cache: dict[Path, tuple[float, frozenset]] = {}

# Full-scan cache: (dir_mtime, result_dict) — invalidated when issue_reports/ changes
_all_cache: Optional[tuple[float, dict]] = None


def _parse_zip(zp: Path) -> frozenset[str]:
    """Parse one report ZIP and return evidenced rule_ids. Cached by file mtime."""
    try:
        mtime = zp.stat().st_mtime
    except OSError:
        return frozenset()

    cached = _zip_cache.get(zp)
    if cached and cached[0] == mtime:
        return cached[1]

    evidenced: set[str] = set()
    try:
        with zipfile.ZipFile(zp) as zf:
            for entry in zf.namelist():
                if not entry.endswith(".xlsx"):
                    continue
                table = entry[:-5]
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(BytesIO(zf.read(entry)), read_only=True)
                    sheets = wb.sheetnames
                    wb.close()
                except Exception as exc:
                    log.warning("Could not parse %s in %s: %s", entry, zp.name, exc)
                    continue

                for sheet in sheets:
                    if sheet.startswith("Missing "):
                        comp = _TABLE_TO_COMP.get(table)
                        if comp:
                            evidenced.add(comp)
                    else:
                        m = _RID_RE.match(sheet)
                        if m:
                            evidenced.add(m.group(1))
    except Exception as exc:
        log.warning("Failed to read evidence from ZIP %s: %s", zp.name, exc)

    result = frozenset(evidenced)
    _zip_cache[zp] = (mtime, result)
    return result


def get_evidenced_rule_ids(le_book: str) -> frozenset[str]:
    """Return rule_ids that have at least one failing-row sheet in the institution's latest report.

    Completeness rules use "Missing {col}" sheet names, so they're detected by table name.
    All other rules use "{RULE_ID}  {name}" sheet names (colon replaced by space by openpyxl).
    Returns an empty frozenset when no report ZIP exists for the institution.
    """
    if not ISSUE_REPORTS_DIR.exists():
        return frozenset()

    zips = sorted(
        [z for z in ISSUE_REPORTS_DIR.glob(f"{le_book}_*.zip")
         if not z.name.endswith("_resolved.zip")],
        reverse=True,
    )
    if not zips:
        return frozenset()

    return _parse_zip(zips[0])


def get_all_evidence() -> dict[str, frozenset[str]]:
    """Return {le_book: evidenced_rule_ids} for every institution that has a report ZIP.

    Cached by issue_reports/ directory mtime — returns the same dict for the lifetime
    of a pipeline run without re-opening any ZIPs.  Individual ZIPs are also cached
    by their own mtime so a partial refresh (e.g. one institution's ZIP replaced) is
    handled correctly.
    """
    global _all_cache

    if not ISSUE_REPORTS_DIR.exists():
        return {}

    try:
        dir_mtime = ISSUE_REPORTS_DIR.stat().st_mtime
    except OSError:
        return {}

    if _all_cache is not None and _all_cache[0] == dir_mtime:
        return _all_cache[1]

    result: dict[str, frozenset[str]] = {}
    seen: set[str] = set()
    for zp in sorted(ISSUE_REPORTS_DIR.glob("*_*.zip"), reverse=True):
        if zp.name.endswith("_resolved.zip"):
            continue
        parts = zp.stem.split("_", 1)
        if len(parts) != 2:
            continue
        lb = parts[0]
        if lb in seen:
            continue
        seen.add(lb)
        result[lb] = _parse_zip(zp)

    _all_cache = (dir_mtime, result)
    return result
