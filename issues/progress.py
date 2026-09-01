# Progress monitoring for partially-resolved issues.
# Reads original_failing_rows vs current failing_rows and the dq_issue_progress
# scan log to compute how many rows have been fixed so far.
from __future__ import annotations
import io
import logging

log = logging.getLogger("issues.progress")


def get_progress_summary(issue_id_: str) -> dict:
    """Return progress stats for one open issue.

    Returns a dict with keys:
      original, current, fixed, pct (0-100), scans (list of {scan_date, failing_rows}).
    Returns {} if the issue is not found or has no original_failing_rows yet.
    """
    from issues.repositories import ensure_tables, get_issue_by_id
    from sqlalchemy import text
    from storage.postgres.connection import get_engine

    ensure_tables()
    iss = get_issue_by_id(issue_id_)
    if not iss:
        return {}

    original = iss.get("original_failing_rows")
    current  = int(iss.get("failing_rows") or 0)
    if not original:
        return {"current": current, "scans": []}

    original = int(original)
    fixed    = max(0, original - current)
    pct      = round(fixed / original * 100) if original else 0

    try:
        with get_engine().connect() as con:
            rows = con.execute(
                text(
                    "SELECT scan_date, failing_rows FROM dq_issue_progress "
                    "WHERE issue_id=:iid ORDER BY scan_date"
                ),
                {"iid": issue_id_},
            ).mappings().fetchall()
            scans = [{"scan_date": r["scan_date"], "failing_rows": int(r["failing_rows"])}
                     for r in rows]
    except Exception:
        scans = []

    return {
        "original": original,
        "current":  current,
        "fixed":    fixed,
        "pct":      pct,
        "scans":    scans,
    }


def build_remaining_xlsx(le_book: str, rule_id: str, table_name: str) -> bytes | None:
    """Build an Excel workbook from the latest evidence snapshot (remaining failing rows).

    Falls back to the original snapshot if no 'latest' exists yet
    (i.e. only one scan has run so far).
    Returns None if no evidence is stored at all.
    """
    from storage.evidence_store import load_rows
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    run_date, rows = load_rows(le_book, rule_id, table_name, snapshot_type="latest")
    if not rows:
        run_date, rows = load_rows(le_book, rule_id, table_name, snapshot_type="original")
    if not rows:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = table_name[:31]

    hdr_fill  = PatternFill("solid", fgColor="753918")
    hdr_font  = Font(color="FFFFFF", bold=True, size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = list(rows[0].keys())
    for ci, h in enumerate(headers, 1):
        cell            = ws.cell(row=1, column=ci, value=h)
        cell.font       = hdr_font
        cell.fill       = hdr_fill
        cell.alignment  = hdr_align
        ws.column_dimensions[cell.column_letter].width = max(12, len(str(h)) + 2)
    ws.row_dimensions[1].height = 20

    alt_fill = PatternFill("solid", fgColor="F2EDE9")
    for ri, row in enumerate(rows, 2):
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=ri, column=ci, value=row.get(h))
            if ri % 2 == 0:
                cell.fill = alt_fill

    buf = io.BytesIO()
    wb.save(buf)
    log.info("Built remaining-rows xlsx for %s/%s/%s (%d rows, run %s)",
             le_book, rule_id, table_name, len(rows), run_date)
    return buf.getvalue()
